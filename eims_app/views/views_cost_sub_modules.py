"""
造价咨询子模块视图 - 包含6个子模块的所有视图函数
项目信息、任务计划、任务实施、审核成果、收费情况、项目存档
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_GET
from ..models import (
    CostProjectInfo,
    CostTaskPlan,
    CostTaskImplementation,
    CostReviewResult,
    CostPaymentStatus,
    CostProjectArchive,
    CostRemunerationDistribution,
    CostRemunerationItem,
    # 统一表模型
    CostProjectUnified,
    CostUnifiedRemunerationItem,
    CostConsultingReminder,
)
from ..forms.form_cost_sub_modules import (
    CostProjectInfoForm,
    CostProjectUnifiedForm,
    CostTaskPlanForm,
    CostTaskPlanUnifiedForm,
    CostTaskImplementationForm,
    CostReviewResultForm,
    CostPaymentStatusForm,
    CostProjectArchiveForm,
    CostRemunerationDistributionForm,
    CostRemunerationItemForm,
)
from ..utils.tenant_utils import filter_queryset_by_tenant
import openpyxl
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.conf import settings
from datetime import datetime
import logging
import os


# ============================================================================
# 项目信息视图
# ============================================================================

@login_required
def cost_project_info_list(request):
    """项目信息列表 - 使用统一表"""
    # 使用统一表查询，只加载需要的字段以提高性能
    queryset = CostProjectUnified.objects.only(
        'id', 'project_code', 'project_name', 'project_type',
        'project_status', 'compilation_category', 'review_category',
        'client_unit', 'entrusting_unit', 'contact_person', 'contact_phone',
        'submission_time', 'start_time',
        'compilation_amount', 'submission_amount',
        'approved_amount', 'reduced_amount', 'total_fee',
        'received_fee', 'pending_fee', 'created_at', 'update_time'
    ).all()
    
    queryset = filter_queryset_by_tenant(queryset, request)
    
    # 搜索
    search_key = request.GET.get('search', '')
    if search_key:
        queryset = queryset.filter(
            Q(project_code__icontains=search_key) |
            Q(project_name__icontains=search_key)
        )
    
    # 筛选
    project_status = request.GET.get('project_status', '')
    if project_status:
        queryset = queryset.filter(project_status=project_status)
    
    project_type = request.GET.get('project_type', '')
    if project_type:
        queryset = queryset.filter(project_type=project_type)
    
    # 动态字段筛选（来自右键菜单）
    for param_key in request.GET.keys():
        if param_key.startswith('filter_') and not param_key.endswith('_op'):
            field_name = param_key.replace('filter_', '', 1)
            operator = request.GET.get(f'{param_key}_op', 'contains')
            filter_value = request.GET.get(param_key, '').strip()
            
            if filter_value:
                # 根据操作符构建不同的查询
                if operator == 'contains':
                    queryset = queryset.filter(**{f'{field_name}__icontains': filter_value})
                elif operator == 'equals':
                    queryset = queryset.filter(**{field_name: filter_value})
                elif operator == 'starts_with':
                    queryset = queryset.filter(**{f'{field_name}__istartswith': filter_value})
                elif operator == 'ends_with':
                    queryset = queryset.filter(**{f'{field_name}__iendswith': filter_value})
                elif operator == 'not_contains':
                    queryset = queryset.exclude(**{f'{field_name}__icontains': filter_value})
    
    # 排序 (Django Admin 风格 - 支持多字段排序)
    sort_fields_str = request.GET.get('sort_field', 'created_at')
    sort_orders_str = request.GET.get('sort_order', 'desc')
    
    # DEBUG: Print sorting parameters
    print(f"DEBUG SORT - sort_field: {sort_fields_str}, sort_order: {sort_orders_str}")
    
    # 清除默认排序，确保自定义排序生效
    queryset = queryset.order_by()
    
    # 解析多字段排序（逗号分隔）
    if ',' in sort_fields_str:
        fields = [f.strip() for f in sort_fields_str.split(',')]
        orders = [o.strip() for o in sort_orders_str.split(',')]
        # 补齐 order 数量
        while len(orders) < len(fields):
            orders.append('asc')
        
        order_list = []
        for field, order in zip(fields, orders):
            if order == 'desc':
                order_list.append(f'-{field}')
            else:
                order_list.append(field)
        queryset = queryset.order_by(*order_list)
        print(f"DEBUG SORT - Multi-field order: {order_list}")
    else:
        # 单字段排序
        field = sort_fields_str.strip()
        order = sort_orders_str.strip() if sort_orders_str else 'asc'
        if order == 'desc':
            queryset = queryset.order_by(f'-{field}')
        else:
            queryset = queryset.order_by(field)
        print(f"DEBUG SORT - Single-field order: {'-' if order == 'desc' else ''}{field}")
    
    # 分页
    paginator = Paginator(queryset, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 统计
    total_count = queryset.count()
    
    context = {
        'page_obj': page_obj,
        'search_key': search_key,
        'project_status': project_status,
        'project_type': project_type,
        'total_count': total_count,
        'PROJECT_STATUS_CHOICES': CostProjectUnified.PROJECT_STATUS_CHOICES,
        'PROJECT_TYPE_CHOICES': CostProjectUnified.PROJECT_TYPE_CHOICES,
    }
    return render(request, 'cost_consulting/project_info/list.html', context)


@login_required
def cost_project_info_add(request):
    """新增项目信息 - 使用统一表"""
    if request.method == 'POST':
        form = CostProjectUnifiedForm(request.POST, tenant=request.tenant if hasattr(request, 'tenant') else None)
        if form.is_valid():
            unified_obj = form.save(commit=False)
            # 自动设置租户
            if hasattr(request, 'tenant') and request.tenant:
                unified_obj.tenant = request.tenant
            unified_obj.save()
            
            # 触发提醒：通知管理者有新项目待分配任务
            # 这里假设管理者是超级用户或特定角色的用户，实际应用中可以根据角色过滤
            from django.contrib.auth import get_user_model
            User = get_user_model()
            managers = User.objects.filter(is_staff=True) # 简化处理，通知所有管理员
            for manager in managers:
                CostConsultingReminder.objects.create(
                    tenant=unified_obj.tenant,
                    project=unified_obj,
                    sender=request.user,
                    receiver=manager,
                    reminder_type='new_project',
                    title=f'新项目待分配: {unified_obj.project_name}',
                    content=f'项目编号: {unified_obj.project_code} 已录入，请及时分配任务计划。',
                    link_url=reverse('eims_app:cost_task_plan_edit', args=[unified_obj.pk])
                )
            
            messages.success(request, '✓ 项目信息添加成功！')
            return redirect(reverse('eims_app:cost_project_info_list'))
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CostProjectUnifiedForm(tenant=request.tenant if hasattr(request, 'tenant') else None)
    
    context = {
        'form': form,
        'title': '新增项目信息',
        'action': 'add',
    }
    return render(request, 'cost_consulting/project_info/form.html', context)


@login_required
def cost_project_info_detail(request, pk):
    """项目信息详情 - 使用统一表"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    context = {'object': obj}
    return render(request, 'cost_consulting/project_info/detail.html', context)


@login_required
def cost_project_info_edit(request, pk):
    """编辑项目信息 - 使用统一表"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    
    if request.method == 'POST':
        form = CostProjectUnifiedForm(request.POST, instance=obj, tenant=request.tenant if hasattr(request, 'tenant') else None)
        if form.is_valid():
            form.save()
            messages.success(request, '✓ 项目信息更新成功！')
            # 重定向到列表页，使用 reverse() 确保清除所有GET参数
            return redirect(reverse('eims_app:cost_project_info_list'))
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CostProjectUnifiedForm(instance=obj, tenant=request.tenant if hasattr(request, 'tenant') else None)
    
    context = {
        'form': form,
        'object': obj,
        'title': '编辑项目信息',
        'action': 'edit',
    }
    return render(request, 'cost_consulting/project_info/form.html', context)


@login_required
def cost_project_info_delete(request, pk):
    """删除项目信息 - 使用统一表"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    obj.delete()
    messages.success(request, '✓ 删除成功！')
    return redirect(reverse('eims_app:cost_project_info_list'))


@login_required
def cost_project_info_batch_delete(request):
    """批量删除项目信息 - 使用统一表"""
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if ids:
            CostProjectUnified.objects.filter(id__in=ids).delete()
            messages.success(request, f'✓ 成功删除 {len(ids)} 条记录！')
        else:
            messages.warning(request, '⚠️ 未选择要删除的记录')
    return redirect(reverse('eims_app:cost_project_info_list'))


@login_required
def cost_project_info_export(request):
    """导出项目信息 - 使用统一表"""
    queryset = CostProjectUnified.objects.all()
    queryset = filter_queryset_by_tenant(queryset, request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目信息"
    
    headers = ['项目编号', '项目名称', '项目类型', '编制类别', '审核类别', '项目状态',
               '建设单位', '委托单位', '联系人', '联系电话', '送审时间', '开始时间',
               '计划工期', '计划完成时间', '编制金额', '送审金额', '审定金额', '审减金额',
               '报告时间', '结果确认', '费用总额', '已收费用', '待收费用', '费用结清']
    ws.append(headers)
    
    for obj in queryset:
        ws.append([
            obj.project_code, obj.project_name, obj.get_project_type_display(),
            obj.get_compilation_category_display(), obj.get_review_category_display(),
            obj.get_project_status_display(), obj.client_unit, obj.entrusting_unit,
            obj.contact_person, obj.contact_phone, obj.submission_time, obj.start_time,
            obj.planned_duration, obj.planned_completion_time, obj.compilation_amount,
            obj.submission_amount, obj.approved_amount, obj.reduced_amount,
            obj.report_time, obj.get_result_confirm_display(), obj.total_fee,
            obj.received_fee, obj.pending_fee, obj.get_fee_settlement_display()
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="项目信息.xlsx"'
    wb.save(response)
    return response


@login_required
def cost_project_info_import(request):
    """造价咨询项目信息导入 - 从 Excel 导入数据到 CostProjectInfo 表"""
    
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        
        if not excel_file:
            messages.error(request, '请选择要导入的 Excel 文件')
            return redirect('eims_app:cost_project_info_import')
        
        try:
            # 读取 Excel 文件
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            # 读取表头
            headers = [cell.value for cell in sheet[1]]
            
            # 字段映射（中文 → 英文）
            field_mapping = {
                '项目编号': 'project_code',
                '项目名称': 'project_name',
                '项目类型': 'project_type',
                '编制类别': 'compilation_category',
                '审核类别': 'review_category',
                '项目状态': 'project_status',
                '建设单位': 'client_unit',
                '委托单位': 'entrusting_unit',
                '联系人': 'contact_person',
                '联系电话': 'contact_phone',
                '送审时间': 'submission_time',
                '开始时间': 'start_time',
                '计划工期(天)': 'planned_duration',
                '计划完成时间': 'planned_completion_time',
                '编制金额(万元)': 'compilation_amount',
                '送审金额(万元)': 'submission_amount',
                '审定金额(万元)': 'approved_amount',
                '审减金额(万元)': 'reduced_amount',
                '报告时间': 'report_time',
                '结果确认': 'result_confirm',
                '费用总额(万元)': 'total_fee',
                '已收费用(万元)': 'received_fee',
                '待收费用(万元)': 'pending_fee',
                '费用结清': 'fee_settlement',
            }
            
            # 枚举字段映射
            project_type_mapping = {
                '预算': 'budget',
                '结算': 'settlement',
                '审核': 'audit',
                '其他': 'other',
            }
            
            compilation_category_mapping = {
                '土建': 'civil',
                '安装': 'install',
                '市政': 'municipal',
                '装饰': 'decoration',
                '其他': 'other',
            }
            
            review_category_mapping = {
                '初审': 'initial',
                '中审': 'intermediate',
                '终审': 'final',
            }
            
            project_status_mapping = {
                '未开始': 'not_started',
                '进行中': 'in_progress',
                '已完成': 'completed',
                '已暂停': 'suspended',
            }
            
            result_confirm_mapping = {
                '已确认': 'confirmed',
                '未确认': 'unconfirmed',
                '待确认': 'pending',
            }
            
            fee_settlement_mapping = {
                '已结清': 'settled',
                '未结清': 'unsettled',
                '部分结清': 'partial',
            }
            
            success_count = 0
            error_count = 0
            error_rows = []
            logger = logging.getLogger(__name__)
            
            # 从第 2 行开始读取数据
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 跳过空行
                    if not any(row):
                        continue
                    
                    # 创建数据字典
                    data = {}
                    for col_idx, value in enumerate(row, start=1):
                        if col_idx <= len(headers):
                            header = headers[col_idx - 1]
                            if header in field_mapping:
                                field_name = field_mapping[header]
                                if value is not None and str(value).strip():
                                    # 特殊字段处理
                                    date_fields = ['submission_time', 'start_time', 'planned_completion_time', 'report_time']
                                    if field_name in date_fields:
                                        try:
                                            if hasattr(value, 'strftime'):
                                                if hasattr(value, 'date'):
                                                    data[field_name] = value.date()
                                                else:
                                                    data[field_name] = value
                                            else:
                                                date_str = str(value).strip()
                                                if ' ' in date_str:
                                                    date_str = date_str.split(' ')[0]
                                                data[field_name] = datetime.strptime(date_str, '%Y-%m-%d').date()
                                        except Exception as e:
                                            logger.warning(f'第{row_idx}行 {header} 字段日期格式错误：{value}, 错误：{e}')
                                            data[field_name] = None
                                    elif field_name == 'planned_duration':
                                        try:
                                            data[field_name] = int(float(value))
                                        except:
                                            data[field_name] = 0
                                    elif field_name == 'project_type':
                                        data[field_name] = project_type_mapping.get(str(value).strip(), 'budget')
                                    elif field_name == 'compilation_category':
                                        data[field_name] = compilation_category_mapping.get(str(value).strip(), 'civil')
                                    elif field_name == 'review_category':
                                        data[field_name] = review_category_mapping.get(str(value).strip(), 'initial')
                                    elif field_name == 'project_status':
                                        data[field_name] = project_status_mapping.get(str(value).strip(), 'not_started')
                                    elif field_name == 'result_confirm':
                                        data[field_name] = result_confirm_mapping.get(str(value).strip(), 'unconfirmed')
                                    elif field_name == 'fee_settlement':
                                        data[field_name] = fee_settlement_mapping.get(str(value).strip(), 'unsettled')
                                    else:
                                        data[field_name] = str(value).strip()
                    
                    # 必填字段检查
                    required_fields = ['project_code', 'project_name']
                    missing_fields = [f for f in required_fields if not data.get(f)]
                    
                    if missing_fields:
                        error_msg = f'第{row_idx}行：缺少必填字段 {missing_fields}'
                        error_rows.append(error_msg)
                        error_count += 1
                        continue
                    
                    # 修复金额字段
                    amount_fields = ['compilation_amount', 'submission_amount', 'approved_amount', 'reduced_amount',
                                     'total_fee', 'received_fee', 'pending_fee']
                    for amt_field in amount_fields:
                        if amt_field not in data:
                            data[amt_field] = 0.0
                        else:
                            try:
                                data[amt_field] = float(str(data[amt_field]).replace(',', ''))
                            except:
                                data[amt_field] = 0.0
                    
                    # 设置租户
                    if hasattr(request, 'tenant') and request.tenant:
                        data['tenant'] = request.tenant
                    
                    # 创建或更新记录（根据 project_code）
                    try:
                        project, created = CostProjectInfo.objects.update_or_create(
                            project_code=data['project_code'],
                            defaults=data
                        )
                        if created:
                            success_count += 1
                            logger.info(f'第{row_idx}行：成功创建项目 {data["project_code"]}')
                        else:
                            success_count += 1
                            logger.info(f'第{row_idx}行：成功更新项目 {data["project_code"]}')
                    except Exception as e:
                        error_msg = f'第{row_idx}行：{str(e)}'
                        error_rows.append(error_msg)
                        logger.error(f'导入失败 - 行{row_idx}: {str(e)}\n数据：{data}')
                        error_count += 1
                        continue
                        
                except Exception as e:
                    import traceback
                    error_msg = f'第{row_idx}行：{str(e)}'
                    error_rows.append(error_msg)
                    logger.error(f'导入失败 - 行{row_idx}: {str(e)}\n{traceback.format_exc()}')
                    error_count += 1
                    continue
            
            # 显示导入结果
            if success_count > 0:
                messages.success(request, f'✓ 成功导入 {success_count} 条记录')
            if error_count > 0:
                messages.warning(request, f'⚠ 导入失败 {error_count} 条记录')
                
                # 将错误信息写入日志文件
                error_log_path = os.path.join(settings.BASE_DIR, 'logs', f'cost_import_error_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt')
                os.makedirs(os.path.dirname(error_log_path), exist_ok=True)
                with open(error_log_path, 'w', encoding='utf-8') as f:
                    f.write(f'造价咨询项目信息导入错误详情 - {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
                    f.write(f'成功：{success_count} 条，失败：{error_count} 条\n\n')
                    f.write('=' * 80 + '\n\n')
                    for i, error in enumerate(error_rows, 1):
                        f.write(f'{i}. {error}\n\n')
                
                messages.error(request, f'错误详情已保存到日志文件：{error_log_path}')
                
                # 在消息中显示前 5 个错误
                for error in error_rows[:5]:
                    messages.error(request, error)
            
            return redirect('eims_app:cost_project_info_list')
            
        except Exception as e:
            messages.error(request, f'导入失败：{str(e)}')
            return redirect('eims_app:cost_project_info_import')
    
    # GET 请求显示导入页面
    context = {
        'title': '造价咨询项目信息导入',
        'template_fields': [
            '项目编号', '项目名称', '项目类型', '编制类别', '审核类别', '项目状态',
            '建设单位', '委托单位', '联系人', '联系电话', '送审时间', '开始时间',
            '计划工期(天)', '计划完成时间', '编制金额(万元)', '送审金额(万元)', '审定金额(万元)', '审减金额(万元)',
            '报告时间', '结果确认', '费用总额(万元)', '已收费用(万元)', '待收费用(万元)', '费用结清'
        ]
    }
    
    return render(request, 'cost_consulting/project_info/import.html', context)


@login_required
def cost_project_info_export_template(request):
    """导出造价咨询项目信息导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目信息导入模板"
    
    headers = [
        '项目编号*', '项目名称*', '项目类型', '编制类别', '审核类别', '项目状态',
        '建设单位', '委托单位', '联系人', '联系电话', '送审时间', '开始时间',
        '计划工期(天)', '计划完成时间', '编制金额(万元)', '送审金额(万元)', 
        '审定金额(万元)', '审减金额(万元)', '报告时间', '结果确认', 
        '费用总额(万元)', '已收费用(万元)', '待收费用(万元)', '费用结清'
    ]
    ws.append(headers)
    
    # 设置表头样式
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 设置列宽
    column_widths = [15, 30, 12, 12, 12, 12, 20, 20, 12, 15, 15, 15, 15, 15, 18, 18, 18, 18, 15, 15, 18, 18, 18, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # 添加示例数据
    example_data = [
        ['ZJ-2024-001', '某某大厦造价咨询项目', '预算', '土建', '初审', '进行中',
         '某某建设单位', '某某委托单位', '张三', '13800138000', '2024-01-15', '2024-01-20',
         '30', '2024-02-20', '5000.00', '4800.00', '4500.00', '300.00', '2024-02-25', '已确认',
         '50.00', '30.00', '20.00', '未结清'],
    ]
    for row_data in example_data:
        ws.append(row_data)
    
    # 添加说明工作表
    ws_instructions = wb.create_sheet('填写说明')
    instructions = [
        ['造价咨询项目信息导入模板 - 填写说明'],
        [''],
        ['一、必填字段（带*标记）：'],
        ['  1. 项目编号：项目唯一标识，如 ZJ-2024-001'],
        ['  2. 项目名称：项目全称'],
        [''],
        ['二、枚举字段可选值：'],
        ['  1. 项目类型：预算、结算、审核、其他'],
        ['  2. 编制类别：土建、安装、市政、装饰、其他'],
        ['  3. 审核类别：初审、中审、终审'],
        ['  4. 项目状态：未开始、进行中、已完成、已暂停'],
        ['  5. 结果确认：已确认、未确认、待确认'],
        ['  6. 费用结清：已结清、未结清、部分结清'],
        [''],
        ['三、日期字段格式：'],
        ['  格式：YYYY-MM-DD，如 2024-01-15'],
        [''],
        ['四、金额字段：'],
        ['  单位：万元，保留两位小数，如 5000.00'],
        [''],
        ['五、注意事项：'],
        ['  1. 第一行为表头，请勿修改'],
        ['  2. 从第二行开始填写数据'],
        ['  3. 如果项目编号已存在，将自动更新该项目的数据'],
        ['  4. 导入前请先备份数据'],
    ]
    for row_data in instructions:
        ws_instructions.append(row_data)
    
    ws_instructions.column_dimensions['A'].width = 60
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="造价咨询项目信息导入模板.xlsx"'
    wb.save(response)
    return response


# ============================================================================
# 任务计划视图
# ============================================================================

@login_required
def cost_task_plan_list(request):
    """任务计划列表"""
    queryset = CostProjectUnified.objects.only('id', 'project_code', 'project_name', 'plan_compiler', 'plan_first_reviewer', 'plan_second_reviewer', 'plan_third_reviewer', 'created_at').all()
    queryset = filter_queryset_by_tenant(queryset, request)
    
    search_key = request.GET.get('search', '')
    if search_key:
        queryset = queryset.filter(
            Q(project_code__icontains=search_key) |
            Q(project_name__icontains=search_key)
        )
    
    # 动态字段筛选（来自右键菜单）
    for param_key in request.GET.keys():
        if param_key.startswith('filter_') and not param_key.endswith('_op'):
            field_name = param_key.replace('filter_', '', 1)
            operator = request.GET.get(f'{param_key}_op', 'contains')
            filter_value = request.GET.get(param_key, '').strip()
            
            if filter_value:
                # 根据操作符构建不同的查询
                if operator == 'contains':
                    queryset = queryset.filter(**{f'{field_name}__icontains': filter_value})
                elif operator == 'equals':
                    queryset = queryset.filter(**{field_name: filter_value})
                elif operator == 'starts_with':
                    queryset = queryset.filter(**{f'{field_name}__istartswith': filter_value})
                elif operator == 'ends_with':
                    queryset = queryset.filter(**{f'{field_name}__iendswith': filter_value})
                elif operator == 'not_contains':
                    queryset = queryset.exclude(**{f'{field_name}__icontains': filter_value})
    
    # 排序 (Django Admin 风格 - 支持多字段排序)
    sort_fields_str = request.GET.get('sort_field', 'created_at')
    sort_orders_str = request.GET.get('sort_order', 'desc')
    
    # 清除默认排序，确保自定义排序生效
    queryset = queryset.order_by()
    
    # 解析多字段排序（逗号分隔）
    if ',' in sort_fields_str:
        fields = [f.strip() for f in sort_fields_str.split(',')]
        orders = [o.strip() for o in sort_orders_str.split(',')]
        # 补齐 order 数量
        while len(orders) < len(fields):
            orders.append('asc')
        
        order_list = []
        for field, order in zip(fields, orders):
            if order == 'desc':
                order_list.append(f'-{field}')
            else:
                order_list.append(field)
        queryset = queryset.order_by(*order_list)
    else:
        # 单字段排序
        field = sort_fields_str.strip()
        order = sort_orders_str.strip() if sort_orders_str else 'asc'
        if order == 'desc':
            queryset = queryset.order_by(f'-{field}')
        else:
            queryset = queryset.order_by(field)
    
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_key': search_key,
        'total_count': queryset.count(),
        'show_detail': False,
        'current_project': None,
    }
    return render(request, 'cost_consulting/task_plan/list.html', context)


@login_required
def cost_task_plan_add(request):
    """新增任务计划 - 使用统一表"""
    if request.method == 'POST':
        form = CostTaskPlanUnifiedForm(request.POST, tenant=getattr(request, 'tenant', None))
        if form.is_valid():
            obj = form.save(commit=False)
            if hasattr(request, 'tenant') and request.tenant:
                obj.tenant = request.tenant
            
            # 自动同步项目信息到冗余字段
            if obj.project:
                obj.project_code = obj.project.project_code
                obj.project_name = obj.project.project_name
                obj.project_type = obj.project.project_type
            
            obj.save()
            messages.success(request, '✓ 任务计划添加成功！')
            return redirect('eims_app:cost_task_plan_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CostTaskPlanUnifiedForm(tenant=getattr(request, 'tenant', None))
    
    context = {'form': form, 'title': '新增任务计划', 'action': 'add'}
    return render(request, 'cost_consulting/task_plan/form.html', context)


@login_required
def cost_task_plan_detail(request, pk):
    """任务计划详情"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    return render(request, 'cost_consulting/task_plan/detail.html', {'object': obj})


@login_required
def cost_task_plan_edit(request, pk):
    """编辑任务计划 - 使用统一表"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    
    if request.method == 'POST':
        form = CostTaskPlanUnifiedForm(request.POST, instance=obj, tenant=getattr(request, 'tenant', None))
        if form.is_valid():
            updated_obj = form.save(commit=False)
            
            # 自动同步项目信息到冗余字段
            if updated_obj.project:
                updated_obj.project_code = updated_obj.project.project_code
                updated_obj.project_name = updated_obj.project.project_name
                updated_obj.project_type = updated_obj.project.project_type
            
            updated_obj.save()
            
            # 触发提醒：通知编制人或审核人任务已分配
            from django.contrib.auth import get_user_model
            User = get_user_model()
            
            # 通知编制人
            if updated_obj.plan_compiler_personnel and updated_obj.plan_compiler_personnel.user:
                CostConsultingReminder.objects.create(
                    tenant=updated_obj.tenant,
                    project=updated_obj,
                    sender=request.user,
                    receiver=updated_obj.plan_compiler_personnel.user,
                    reminder_type='task_assigned',
                    title=f'任务已分配: {updated_obj.project_name}',
                    content=f'您被指定为该项目（{updated_obj.project_code}）的编制人，请开始编制工作。',
                    link_url=reverse('eims_app:cost_task_implementation_edit', args=[updated_obj.pk])
                )
            
            # 通知一审人
            if updated_obj.plan_first_reviewer_personnel and updated_obj.plan_first_reviewer_personnel.user:
                CostConsultingReminder.objects.create(
                    tenant=updated_obj.tenant,
                    project=updated_obj,
                    sender=request.user,
                    receiver=updated_obj.plan_first_reviewer_personnel.user,
                    reminder_type='review_start',
                    title=f'审核任务待处理: {updated_obj.project_name}',
                    content=f'您被指定为该项目（{updated_obj.project_code}）的一审人，请准备审核工作。',
                    link_url=reverse('eims_app:cost_review_result_edit', args=[updated_obj.pk])
                )
            
            messages.success(request, '✓ 任务计划更新成功！')
            return redirect('eims_app:cost_task_plan_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CostTaskPlanUnifiedForm(instance=obj, tenant=getattr(request, 'tenant', None))
    
    context = {'form': form, 'object': obj, 'title': '编辑任务计划', 'action': 'edit'}
    return render(request, 'cost_consulting/task_plan/form.html', context)


@login_required
def cost_task_plan_delete(request, pk):
    """删除任务计划"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    obj.delete()
    messages.success(request, '✓ 删除成功！')
    return redirect('eims_app:cost_task_plan_list')


@login_required
def cost_task_plan_batch_delete(request):
    """批量删除任务计划"""
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if ids:
            CostProjectUnified.objects.filter(id__in=ids).delete()
            messages.success(request, f'✓ 成功删除 {len(ids)} 条记录！')
    return redirect('eims_app:cost_task_plan_list')


@login_required
def cost_task_plan_export(request):
    """导出任务计划"""
    queryset = CostProjectUnified.objects.all()
    queryset = filter_queryset_by_tenant(queryset, request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "任务计划"
    
    headers = ['项目编号', '项目名称', '项目类型', '编制人', '编制金额',
               '一审人员', '一审开始时间', '一审计划工期', '一审计划完成时间',
               '二审人员', '二审开始时间', '二审计划工期', '二审计划完成时间',
               '三审人员', '三审开始时间', '三审计划工期', '三审计划完成时间']
    ws.append(headers)
    
    for obj in queryset:
        ws.append([
            obj.project_code, obj.project_name, obj.get_project_type_display(),
            obj.compiler, obj.compilation_amount,
            obj.first_reviewer, obj.first_review_start_time, obj.first_review_planned_duration, obj.first_review_planned_completion,
            obj.second_reviewer, obj.second_review_start_time, obj.second_review_planned_duration, obj.second_review_planned_completion,
            obj.third_reviewer, obj.third_review_start_time, obj.third_review_planned_duration, obj.third_review_planned_completion
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="任务计划.xlsx"'
    wb.save(response)
    return response


# ============================================================================
# 任务实施视图
# ============================================================================

@login_required
def cost_task_implementation_list(request):
    """任务实施列表"""
    queryset = CostProjectUnified.objects.only('id', 'project_code', 'project_name', 'impl_compiler', 'impl_first_reviewer_personnel', 'implementation_status', 'created_at').all()
    queryset = filter_queryset_by_tenant(queryset, request)
    
    search_key = request.GET.get('search', '')
    if search_key:
        queryset = queryset.filter(
            Q(project_code__icontains=search_key) |
            Q(project_name__icontains=search_key)
        )
    
    # 动态字段筛选（来自右键菜单）
    for param_key in request.GET.keys():
        if param_key.startswith('filter_') and not param_key.endswith('_op'):
            field_name = param_key.replace('filter_', '', 1)
            operator = request.GET.get(f'{param_key}_op', 'contains')
            filter_value = request.GET.get(param_key, '').strip()
            
            if filter_value:
                # 根据操作符构建不同的查询
                if operator == 'contains':
                    queryset = queryset.filter(**{f'{field_name}__icontains': filter_value})
                elif operator == 'equals':
                    queryset = queryset.filter(**{field_name: filter_value})
                elif operator == 'starts_with':
                    queryset = queryset.filter(**{f'{field_name}__istartswith': filter_value})
                elif operator == 'ends_with':
                    queryset = queryset.filter(**{f'{field_name}__iendswith': filter_value})
                elif operator == 'not_contains':
                    queryset = queryset.exclude(**{f'{field_name}__icontains': filter_value})
    
    # 排序 (Django Admin 风格 - 支持多字段排序)
    sort_fields_str = request.GET.get('sort_field', 'created_at')
    sort_orders_str = request.GET.get('sort_order', 'desc')
    
    # 清除默认排序，确保自定义排序生效
    queryset = queryset.order_by()
    
    # 解析多字段排序（逗号分隔）
    if ',' in sort_fields_str:
        fields = [f.strip() for f in sort_fields_str.split(',')]
        orders = [o.strip() for o in sort_orders_str.split(',')]
        # 补齐 order 数量
        while len(orders) < len(fields):
            orders.append('asc')
        
        order_list = []
        for field, order in zip(fields, orders):
            if order == 'desc':
                order_list.append(f'-{field}')
            else:
                order_list.append(field)
        queryset = queryset.order_by(*order_list)
    else:
        # 单字段排序
        field = sort_fields_str.strip()
        order = sort_orders_str.strip() if sort_orders_str else 'asc'
        if order == 'desc':
            queryset = queryset.order_by(f'-{field}')
        else:
            queryset = queryset.order_by(field)
    
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_key': search_key,
        'total_count': queryset.count(),
        'show_detail': False,
        'current_project': None,
    }
    return render(request, 'cost_consulting/task_implementation/list.html', context)


@login_required
def cost_task_implementation_add(request):
    """新增任务实施"""
    if request.method == 'POST':
        form = CostTaskImplementationForm(request.POST, tenant=getattr(request, 'tenant', None))
        if form.is_valid():
            # 获取选中的项目
            selected_project = form.cleaned_data.get('selected_project')
            if not selected_project:
                messages.error(request, '请选择项目！')
            else:
                # 更新项目的任务实施字段
                selected_project.impl_compiler = form.cleaned_data.get('impl_compiler', '')
                selected_project.impl_compiler_personnel = form.cleaned_data.get('impl_compiler_personnel')
                selected_project.impl_compilation_amount = form.cleaned_data.get('impl_compilation_amount', 0)
                selected_project.impl_compilation_start = form.cleaned_data.get('impl_compilation_start')
                selected_project.impl_compilation_end = form.cleaned_data.get('impl_compilation_end')
                selected_project.impl_compilation_actual_duration = form.cleaned_data.get('impl_compilation_actual_duration', 0)
                selected_project.impl_first_reviewer = form.cleaned_data.get('impl_first_reviewer', '')
                selected_project.impl_first_reviewer_personnel = form.cleaned_data.get('impl_first_reviewer_personnel')
                selected_project.impl_first_review_start = form.cleaned_data.get('impl_first_review_start')
                selected_project.impl_first_review_end = form.cleaned_data.get('impl_first_review_end')
                selected_project.impl_first_review_actual_duration = form.cleaned_data.get('impl_first_review_actual_duration', 0)
                selected_project.impl_first_review_progress_result = form.cleaned_data.get('impl_first_review_progress_result', '')
                selected_project.impl_second_reviewer = form.cleaned_data.get('impl_second_reviewer', '')
                selected_project.impl_second_reviewer_personnel = form.cleaned_data.get('impl_second_reviewer_personnel')
                selected_project.impl_second_review_start = form.cleaned_data.get('impl_second_review_start')
                selected_project.impl_second_review_end = form.cleaned_data.get('impl_second_review_end')
                selected_project.impl_second_review_actual_duration = form.cleaned_data.get('impl_second_review_actual_duration', 0)
                selected_project.impl_second_review_progress_result = form.cleaned_data.get('impl_second_review_progress_result', '')
                selected_project.impl_third_reviewer = form.cleaned_data.get('impl_third_reviewer', '')
                selected_project.impl_third_reviewer_personnel = form.cleaned_data.get('impl_third_reviewer_personnel')
                selected_project.impl_third_review_start = form.cleaned_data.get('impl_third_review_start')
                selected_project.impl_third_review_end = form.cleaned_data.get('impl_third_review_end')
                selected_project.impl_third_review_actual_duration = form.cleaned_data.get('impl_third_review_actual_duration', 0)
                selected_project.impl_third_review_progress_result = form.cleaned_data.get('impl_third_review_progress_result', '')
                selected_project.implementation_status = form.cleaned_data.get('implementation_status', 'not_started')
                selected_project.save()
                messages.success(request, '✓ 任务实施添加成功！')
                return redirect('eims_app:cost_task_implementation_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CostTaskImplementationForm(tenant=getattr(request, 'tenant', None))
    
    context = {'form': form, 'title': '新增任务实施', 'action': 'add'}
    return render(request, 'cost_consulting/task_implementation/form.html', context)


@login_required
def cost_task_implementation_detail(request, pk):
    """任务实施详情"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    return render(request, 'cost_consulting/task_implementation/detail.html', {'object': obj})


@login_required
def cost_task_implementation_edit(request, pk):
    """编辑任务实施"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    
    if request.method == 'POST':
        form = CostTaskImplementationForm(request.POST, instance=obj, tenant=getattr(request, 'tenant', None))
        if form.is_valid():
            # 更新项目的任务实施字段
            obj.impl_compiler = form.cleaned_data.get('impl_compiler', '')
            obj.impl_compiler_personnel = form.cleaned_data.get('impl_compiler_personnel')
            obj.impl_compilation_amount = form.cleaned_data.get('impl_compilation_amount', 0)
            obj.impl_compilation_start = form.cleaned_data.get('impl_compilation_start')
            obj.impl_compilation_end = form.cleaned_data.get('impl_compilation_end')
            obj.impl_compilation_actual_duration = form.cleaned_data.get('impl_compilation_actual_duration', 0)
            obj.impl_first_reviewer = form.cleaned_data.get('impl_first_reviewer', '')
            obj.impl_first_reviewer_personnel = form.cleaned_data.get('impl_first_reviewer_personnel')
            obj.impl_first_review_start = form.cleaned_data.get('impl_first_review_start')
            obj.impl_first_review_end = form.cleaned_data.get('impl_first_review_end')
            obj.impl_first_review_actual_duration = form.cleaned_data.get('impl_first_review_actual_duration', 0)
            obj.impl_first_review_progress_result = form.cleaned_data.get('impl_first_review_progress_result', '')
            obj.impl_second_reviewer = form.cleaned_data.get('impl_second_reviewer', '')
            obj.impl_second_reviewer_personnel = form.cleaned_data.get('impl_second_reviewer_personnel')
            obj.impl_second_review_start = form.cleaned_data.get('impl_second_review_start')
            obj.impl_second_review_end = form.cleaned_data.get('impl_second_review_end')
            obj.impl_second_review_actual_duration = form.cleaned_data.get('impl_second_review_actual_duration', 0)
            obj.impl_second_review_progress_result = form.cleaned_data.get('impl_second_review_progress_result', '')
            obj.impl_third_reviewer = form.cleaned_data.get('impl_third_reviewer', '')
            obj.impl_third_reviewer_personnel = form.cleaned_data.get('impl_third_reviewer_personnel')
            obj.impl_third_review_start = form.cleaned_data.get('impl_third_review_start')
            obj.impl_third_review_end = form.cleaned_data.get('impl_third_review_end')
            obj.impl_third_review_actual_duration = form.cleaned_data.get('impl_third_review_actual_duration', 0)
            obj.impl_third_review_progress_result = form.cleaned_data.get('impl_third_review_progress_result', '')
            obj.implementation_status = form.cleaned_data.get('implementation_status', 'not_started')
            obj.save()
            messages.success(request, '✓ 任务实施更新成功！')
            return redirect('eims_app:cost_task_implementation_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CostTaskImplementationForm(instance=obj, tenant=getattr(request, 'tenant', None))
    
    context = {'form': form, 'object': obj, 'title': '编辑任务实施', 'action': 'edit'}
    return render(request, 'cost_consulting/task_implementation/form.html', context)


@login_required
def cost_task_implementation_delete(request, pk):
    """删除任务实施"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    obj.delete()
    messages.success(request, '✓ 删除成功！')
    return redirect('eims_app:cost_task_implementation_list')


@login_required
def cost_task_implementation_batch_delete(request):
    """批量删除任务实施"""
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if ids:
            CostProjectUnified.objects.filter(id__in=ids).delete()
            messages.success(request, f'✓ 成功删除 {len(ids)} 条记录！')
    return redirect('eims_app:cost_task_implementation_list')


@login_required
def cost_task_implementation_export(request):
    """导出任务实施"""
    queryset = CostProjectUnified.objects.all()
    queryset = filter_queryset_by_tenant(queryset, request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "任务实施"
    
    headers = ['项目编号', '项目名称', '项目类型', '编制人', '编制金额',
               '一审计划工期', '一审计划完成时间', '一审实际完成时间', '一审实际工期', '一审进度结果',
               '二审计划工期', '二审计划完成时间', '二审实际完成时间', '二审实际工期', '二审进度结果',
               '三审计划工期', '三审计划完成时间', '三审实际完成时间', '三审实际工期', '三审进度结果']
    ws.append(headers)
    
    for obj in queryset:
        ws.append([
            obj.project_code, obj.project_name, obj.get_project_type_display(),
            obj.compiler, obj.compilation_amount,
            obj.first_review_planned_duration, obj.first_review_planned_completion, obj.first_review_actual_completion, obj.first_review_actual_duration, obj.first_review_progress_result,
            obj.second_review_planned_duration, obj.second_review_planned_completion, obj.second_review_actual_completion, obj.second_review_actual_duration, obj.second_review_progress_result,
            obj.third_review_planned_duration, obj.third_review_planned_completion, obj.third_review_actual_completion, obj.third_review_actual_duration, obj.third_review_progress_result
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="任务实施.xlsx"'
    wb.save(response)
    return response


# ============================================================================
# 审核成果视图
# ============================================================================

@login_required
@never_cache
def cost_review_result_list(request):
    """审核成果列表"""
    queryset = CostProjectUnified.objects.only('id', 'project_code', 'project_name', 'review_compiler', 'review_final_approved_amount', 'created_at').all()
    queryset = filter_queryset_by_tenant(queryset, request)
    
    search_key = request.GET.get('search', '')
    if search_key:
        queryset = queryset.filter(
            Q(project_code__icontains=search_key) |
            Q(project_name__icontains=search_key)
        )
    
    # 动态字段筛选（来自右键菜单）
    for param_key in request.GET.keys():
        if param_key.startswith('filter_') and not param_key.endswith('_op'):
            field_name = param_key.replace('filter_', '', 1)
            operator = request.GET.get(f'{param_key}_op', 'contains')
            filter_value = request.GET.get(param_key, '').strip()
            
            if filter_value:
                # 根据操作符构建不同的查询
                if operator == 'contains':
                    queryset = queryset.filter(**{f'{field_name}__icontains': filter_value})
                elif operator == 'equals':
                    queryset = queryset.filter(**{field_name: filter_value})
                elif operator == 'starts_with':
                    queryset = queryset.filter(**{f'{field_name}__istartswith': filter_value})
                elif operator == 'ends_with':
                    queryset = queryset.filter(**{f'{field_name}__iendswith': filter_value})
                elif operator == 'not_contains':
                    queryset = queryset.exclude(**{f'{field_name}__icontains': filter_value})
    
    # 排序 (Django Admin 风格 - 支持多字段排序)
    sort_fields_str = request.GET.get('sort_field', 'created_at')
    sort_orders_str = request.GET.get('sort_order', 'desc')
    
    # 清除默认排序，确保自定义排序生效
    queryset = queryset.order_by()
    
    # 解析多字段排序（逗号分隔）
    if ',' in sort_fields_str:
        fields = [f.strip() for f in sort_fields_str.split(',')]
        orders = [o.strip() for o in sort_orders_str.split(',')]
        # 补齐 order 数量
        while len(orders) < len(fields):
            orders.append('asc')
        
        order_list = []
        for field, order in zip(fields, orders):
            if order == 'desc':
                order_list.append(f'-{field}')
            else:
                order_list.append(field)
        queryset = queryset.order_by(*order_list)
    else:
        # 单字段排序
        field = sort_fields_str.strip()
        order = sort_orders_str.strip() if sort_orders_str else 'asc'
        if order == 'desc':
            queryset = queryset.order_by(f'-{field}')
        else:
            queryset = queryset.order_by(field)
    
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_key': search_key,
        'total_count': queryset.count(),
        'show_detail': False,
        'current_project': None,
        'page_title': '审核成果',
    }
    return render(request, 'cost_consulting/review_result/list.html', context)


@login_required
def cost_review_result_add(request):
    """新增审核成果"""
    if request.method == 'POST':
        form = CostReviewResultForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            if hasattr(request, 'tenant') and request.tenant:
                obj.tenant = request.tenant
            
            # 自动同步项目信息到冗余字段
            if obj.project:
                obj.project_code = obj.project.project_code
                obj.project_name = obj.project.project_name
                obj.project_type = obj.project.project_type
            
            obj.save()
            messages.success(request, '✓ 审核成果添加成功！')
            return redirect('eims_app:cost_review_result_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CostReviewResultForm()
    
    context = {'form': form, 'title': '新增审核成果', 'action': 'add'}
    return render(request, 'cost_consulting/review_result/form.html', context)


@login_required
def cost_review_result_detail(request, pk):
    """审核成果详情"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    return render(request, 'cost_consulting/review_result/detail.html', {'object': obj})


@login_required
def cost_review_result_edit(request, pk):
    """编辑审核成果"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    
    if request.method == 'POST':
        form = CostReviewResultForm(request.POST, instance=obj)
        if form.is_valid():
            updated_obj = form.save(commit=False)
            
            # 自动同步项目信息到冗余字段
            if updated_obj.project:
                updated_obj.project_code = updated_obj.project.project_code
                updated_obj.project_name = updated_obj.project.project_name
                updated_obj.project_type = updated_obj.project.project_type
            
            updated_obj.save()
            messages.success(request, '✓ 审核成果更新成功！')
            return redirect('eims_app:cost_review_result_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CostReviewResultForm(instance=obj)
    
    context = {'form': form, 'object': obj, 'title': '编辑审核成果', 'action': 'edit'}
    return render(request, 'cost_consulting/review_result/form.html', context)


@login_required
def cost_review_result_delete(request, pk):
    """删除审核成果"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    obj.delete()
    messages.success(request, '✓ 删除成功！')
    return redirect('eims_app:cost_review_result_list')


@login_required
def cost_review_result_batch_delete(request):
    """批量删除审核成果"""
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if ids:
            CostProjectUnified.objects.filter(id__in=ids).delete()
            messages.success(request, f'✓ 成功删除 {len(ids)} 条记录！')
    return redirect('eims_app:cost_review_result_list')


@login_required
def cost_review_result_export(request):
    """导出审核成果"""
    queryset = CostProjectUnified.objects.all()
    queryset = filter_queryset_by_tenant(queryset, request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "审核成果"
    
    headers = ['项目编号', '项目名称', '项目类型', '编制人', '编制金额',
               '一审送审', '一审结果', '一审审减', '一审减率', '一审评价',
               '二审送审', '二审结果', '二审减率', '二审人员', '二审评价',
               '三审送审', '三审结果', '三审减率', '三审人员', '三审评价', '审定金额']
    ws.append(headers)
    
    for obj in queryset:
        ws.append([
            obj.project_code, obj.project_name, obj.get_project_type_display(),
            obj.compiler, obj.compilation_amount,
            obj.first_submission, obj.first_result, obj.first_reduction, obj.first_reduction_rate, obj.first_review_evaluation,
            obj.second_submission, obj.second_result, obj.second_reduction_rate, obj.second_reviewer, obj.second_evaluation,
            obj.third_submission, obj.third_result, obj.third_reduction_rate, obj.third_reviewer, obj.third_evaluation,
            obj.final_approved_amount
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="审核成果.xlsx"'
    wb.save(response)
    return response


# ============================================================================
# 收费情况视图
# ============================================================================

@login_required
@never_cache
def cost_payment_status_list(request):
    """收费情况列表"""
    queryset = CostProjectUnified.objects.only('id', 'project_code', 'project_name', 'payment_invoice_amount', 'payment_is_invoiced', 'payment_is_settled', 'created_at').all()
    queryset = filter_queryset_by_tenant(queryset, request)
    
    search_key = request.GET.get('search', '')
    if search_key:
        queryset = queryset.filter(
            Q(project_code__icontains=search_key) |
            Q(project_name__icontains=search_key)
        )
    
    # 动态字段筛选（来自右键菜单）
    for param_key in request.GET.keys():
        if param_key.startswith('filter_') and not param_key.endswith('_op'):
            field_name = param_key.replace('filter_', '', 1)
            operator = request.GET.get(f'{param_key}_op', 'contains')
            filter_value = request.GET.get(param_key, '').strip()
            
            if filter_value:
                # 根据操作符构建不同的查询
                if operator == 'contains':
                    queryset = queryset.filter(**{f'{field_name}__icontains': filter_value})
                elif operator == 'equals':
                    queryset = queryset.filter(**{field_name: filter_value})
                elif operator == 'starts_with':
                    queryset = queryset.filter(**{f'{field_name}__istartswith': filter_value})
                elif operator == 'ends_with':
                    queryset = queryset.filter(**{f'{field_name}__iendswith': filter_value})
                elif operator == 'not_contains':
                    queryset = queryset.exclude(**{f'{field_name}__icontains': filter_value})
    
    # 排序 (Django Admin 风格 - 支持多字段排序)
    sort_fields_str = request.GET.get('sort_field', 'created_at')
    sort_orders_str = request.GET.get('sort_order', 'desc')
    
    # 清除默认排序，确保自定义排序生效
    queryset = queryset.order_by()
    
    # 解析多字段排序（逗号分隔）
    if ',' in sort_fields_str:
        fields = [f.strip() for f in sort_fields_str.split(',')]
        orders = [o.strip() for o in sort_orders_str.split(',')]
        # 补齐 order 数量
        while len(orders) < len(fields):
            orders.append('asc')
        
        order_list = []
        for field, order in zip(fields, orders):
            if order == 'desc':
                order_list.append(f'-{field}')
            else:
                order_list.append(field)
        queryset = queryset.order_by(*order_list)
    else:
        # 单字段排序
        field = sort_fields_str.strip()
        order = sort_orders_str.strip() if sort_orders_str else 'asc'
        if order == 'desc':
            queryset = queryset.order_by(f'-{field}')
        else:
            queryset = queryset.order_by(field)
    
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_key': search_key,
        'total_count': queryset.count(),
        'show_detail': False,
        'current_project': None,
    }
    return render(request, 'cost_consulting/payment_status/list.html', context)


@login_required
def cost_payment_status_add(request):
    """新增收费情况"""
    if request.method == 'POST':
        form = CostPaymentStatusForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            if hasattr(request, 'tenant') and request.tenant:
                obj.tenant = request.tenant
            
            # 自动同步项目信息到冗余字段
            if obj.project:
                obj.project_code = obj.project.project_code
                obj.project_name = obj.project.project_name
            
            obj.save()
            messages.success(request, '✓ 收费情况添加成功！')
            return redirect('eims_app:cost_payment_status_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CostPaymentStatusForm()
    
    context = {'form': form, 'title': '新增收费情况', 'action': 'add'}
    return render(request, 'cost_consulting/payment_status/form.html', context)


@login_required
def cost_payment_status_detail(request, pk):
    """收费情况详情"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    return render(request, 'cost_consulting/payment_status/detail.html', {'object': obj})


@login_required
def cost_payment_status_edit(request, pk):
    """编辑收费情况"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    
    if request.method == 'POST':
        form = CostPaymentStatusForm(request.POST, instance=obj)
        if form.is_valid():
            updated_obj = form.save(commit=False)
            
            # 自动同步项目信息到冗余字段
            if updated_obj.project:
                updated_obj.project_code = updated_obj.project.project_code
                updated_obj.project_name = updated_obj.project.project_name
            
            updated_obj.save()
            messages.success(request, '✓ 收费情况更新成功！')
            return redirect('eims_app:cost_payment_status_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CostPaymentStatusForm(instance=obj)
    
    context = {'form': form, 'object': obj, 'title': '编辑收费情况', 'action': 'edit'}
    return render(request, 'cost_consulting/payment_status/form.html', context)


@login_required
def cost_payment_status_delete(request, pk):
    """删除收费情况"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    obj.delete()
    messages.success(request, '✓ 删除成功！')
    return redirect('eims_app:cost_payment_status_list')


@login_required
def cost_payment_status_batch_delete(request):
    """批量删除收费情况"""
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if ids:
            CostProjectUnified.objects.filter(id__in=ids).delete()
            messages.success(request, f'✓ 成功删除 {len(ids)} 条记录！')
    return redirect('eims_app:cost_payment_status_list')


@login_required
def cost_payment_status_export(request):
    """导出收费情况"""
    queryset = CostProjectUnified.objects.all()
    queryset = filter_queryset_by_tenant(queryset, request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "收费情况"
    
    headers = ['项目编号', '项目名称', '开票金额', '是否开票',
               '业主方应付', '业主方已付', '业主方待付',
               '施工方应付', '施工方已付', '施工方待付', '是否结清']
    ws.append(headers)
    
    for obj in queryset:
        ws.append([
            obj.project_code, obj.project_name, obj.invoice_amount, obj.get_is_invoiced_display(),
            obj.owner_payable, obj.owner_paid, obj.owner_pending,
            obj.contractor_payable, obj.contractor_paid, obj.contractor_pending, obj.get_is_settled_display()
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="收费情况.xlsx"'
    wb.save(response)
    return response


# ============================================================================
# 项目存档视图
# ============================================================================

@login_required
@never_cache
def cost_project_archive_list(request):
    """项目存档列表"""
    queryset = CostProjectUnified.objects.only('id', 'project_code', 'project_name', 'archive_status', 'approval_status', 'archive_electronic', 'archive_paper', 'created_at').all()
    queryset = filter_queryset_by_tenant(queryset, request)
    
    search_key = request.GET.get('search', '')
    if search_key:
        queryset = queryset.filter(
            Q(project_code__icontains=search_key) |
            Q(project_name__icontains=search_key)
        )
    
    # 动态字段筛选（来自右键菜单）
    for param_key in request.GET.keys():
        if param_key.startswith('filter_') and not param_key.endswith('_op'):
            field_name = param_key.replace('filter_', '', 1)
            operator = request.GET.get(f'{param_key}_op', 'contains')
            filter_value = request.GET.get(param_key, '').strip()
            
            if filter_value:
                # 根据操作符构建不同的查询
                if operator == 'contains':
                    queryset = queryset.filter(**{f'{field_name}__icontains': filter_value})
                elif operator == 'equals':
                    queryset = queryset.filter(**{field_name: filter_value})
                elif operator == 'starts_with':
                    queryset = queryset.filter(**{f'{field_name}__istartswith': filter_value})
                elif operator == 'ends_with':
                    queryset = queryset.filter(**{f'{field_name}__iendswith': filter_value})
                elif operator == 'not_contains':
                    queryset = queryset.exclude(**{f'{field_name}__icontains': filter_value})
    
    # 排序 (Django Admin 风格 - 支持多字段排序)
    sort_fields_str = request.GET.get('sort_field', 'created_at')
    sort_orders_str = request.GET.get('sort_order', 'desc')
    
    # 清除默认排序，确保自定义排序生效
    queryset = queryset.order_by()
    
    # 解析多字段排序（逗号分隔）
    if ',' in sort_fields_str:
        fields = [f.strip() for f in sort_fields_str.split(',')]
        orders = [o.strip() for o in sort_orders_str.split(',')]
        # 补齐 order 数量
        while len(orders) < len(fields):
            orders.append('asc')
        
        order_list = []
        for field, order in zip(fields, orders):
            if order == 'desc':
                order_list.append(f'-{field}')
            else:
                order_list.append(field)
        queryset = queryset.order_by(*order_list)
    else:
        # 单字段排序
        field = sort_fields_str.strip()
        order = sort_orders_str.strip() if sort_orders_str else 'asc'
        if order == 'desc':
            queryset = queryset.order_by(f'-{field}')
        else:
            queryset = queryset.order_by(field)
    
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_key': search_key,
        'total_count': queryset.count(),
        'show_detail': False,
        'current_project': None,
    }
    return render(request, 'cost_consulting/project_archive/list.html', context)


@login_required
def cost_project_archive_add(request):
    """新增项目存档（发起归档审批申请）"""
    if request.method == 'POST':
        form = CostProjectArchiveForm(request.POST, request.FILES)
        if form.is_valid():
            obj = form.save(commit=False)
            if hasattr(request, 'tenant') and request.tenant:
                obj.tenant = request.tenant
            
            # 自动同步项目信息到冗余字段
            if obj.project:
                obj.project_code = obj.project.project_code
                obj.project_name = obj.project.project_name
            
            # 设置提交人和提交时间
            obj.submitter_user = request.user
            obj.submitter = request.user.get_full_name() or request.user.username
            from django.utils import timezone
            obj.submit_time = timezone.now()
            
            # 初始状态为草稿，如果用户点击了“提交审批”则改为待审核
            if 'submit_for_approval' in request.POST:
                obj.approval_status = 'pending_approval'
                obj.archive_status = 'archiving'  # 提交审批后，归档流程开始
            else:
                obj.approval_status = 'draft'
                obj.archive_status = 'not_archived'  # 默认未归档
            
            obj.save()
            messages.success(request, '✓ 项目存档申请已保存！')
            return redirect('eims_app:cost_project_archive_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CostProjectArchiveForm()
    
    context = {'form': form, 'title': '新增项目存档', 'action': 'add'}
    return render(request, 'cost_consulting/project_archive/form.html', context)


@login_required
def cost_project_archive_detail(request, pk):
    """项目存档详情"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    return render(request, 'cost_consulting/project_archive/detail.html', {'object': obj})


@login_required
def cost_project_archive_edit(request, pk):
    """编辑项目存档（支持审批流程）"""
    from django.utils import timezone
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    
    if request.method == 'POST':
        form = CostProjectArchiveForm(request.POST, request.FILES, instance=obj)
        if form.is_valid():
            updated_obj = form.save(commit=False)
            
            # 根据流程进展自动更新归档状态
            if 'receive_archive' in request.POST:
                # 档案管理员接收档案
                updated_obj.archive_status = 'archived'
                updated_obj.archive_time = timezone.now()
                updated_obj.archive_date = timezone.now().date()
                updated_obj.operator = request.user.get_full_name() or request.user.username
            elif 'approve_archive' in request.POST:
                # 审批人审批通过 - 状态变为"待接收"
                updated_obj.approval_status = 'approved'
                updated_obj.approval_time = timezone.now()
                updated_obj.approval_remark = request.POST.get('approval_remark', '')
                updated_obj.archive_status = 'pending_receive'  # 审批通过后，等待档案管理员接收
                
                # 触发提醒：通知档案管理员接收
                from django.contrib.auth import get_user_model
                User = get_user_model()
                # 简化处理：通知所有档案管理员（这里假设 is_staff 包含档案管理员，实际应根据角色）
                archivists = User.objects.filter(is_staff=True)
                for archivist in archivists:
                    CostConsultingReminder.objects.create(
                        tenant=updated_obj.tenant,
                        project=updated_obj,
                        sender=request.user,
                        receiver=archivist,
                        reminder_type='archive_receive',
                        title=f'档案待接收: {updated_obj.project_name}',
                        content=f'项目归档申请已审批通过，请确定存放位置并接收档案。',
                        link_url=reverse('eims_app:cost_project_archive_edit', args=[updated_obj.pk])
                    )
                    
            elif 'reject_archive' in request.POST:
                # 审批人退回 - 状态变为"退回"
                updated_obj.approval_status = 'rejected'
                updated_obj.approval_remark = request.POST.get('approval_remark', '')
                updated_obj.archive_status = 'rejected'
            elif 'submit_for_approval' in request.POST:
                # 提交人发起审批 - 状态变为"归档中"
                updated_obj.approval_status = 'pending_approval'
                updated_obj.submitter_user = request.user
                updated_obj.submitter = request.user.get_full_name() or request.user.username
                updated_obj.submit_time = timezone.now()
                updated_obj.archive_status = 'archiving'  # 提交审批后，归档流程开始
            
            updated_obj.save()
            messages.success(request, '✓ 项目存档更新成功！')
            return redirect('eims_app:cost_project_archive_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CostProjectArchiveForm(instance=obj)
    
    context = {'form': form, 'object': obj, 'title': '编辑项目存档', 'action': 'edit'}
    return render(request, 'cost_consulting/project_archive/form.html', context)


@login_required
def cost_project_archive_delete(request, pk):
    """删除项目存档"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    obj.delete()
    messages.success(request, '✓ 删除成功！')
    return redirect('eims_app:cost_project_archive_list')


@login_required
def cost_project_archive_batch_delete(request):
    """批量删除项目存档"""
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if ids:
            CostProjectUnified.objects.filter(id__in=ids).delete()
            messages.success(request, f'✓ 成功删除 {len(ids)} 条记录！')
    return redirect('eims_app:cost_project_archive_list')


@login_required
def cost_project_archive_export(request):
    """导出项目存档"""
    queryset = CostProjectUnified.objects.all()
    queryset = filter_queryset_by_tenant(queryset, request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目存档"
    
    headers = ['项目编号', '项目名称', '服务合同', '送审资料', '过程资料', '勘察记录',
               '审核报告', '业主确认', '其他文件', '提交人', '提交时间', '存档时间']
    ws.append(headers)
    
    for obj in queryset:
        ws.append([
            obj.project_code, obj.project_name,
            obj.service_contract.name if obj.service_contract else '',
            obj.submission_material.name if obj.submission_material else '',
            obj.process_material.name if obj.process_material else '',
            obj.inspection_record.name if obj.inspection_record else '',
            obj.audit_report.name if obj.audit_report else '',
            '是' if obj.owner_confirmation else '否',
            obj.other_document.name if obj.other_document else '',
            obj.submitter, obj.submit_time, obj.archive_time
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="项目存档.xlsx"'
    wb.save(response)
    return response


# ============================================================================
# 酬劳分配视图
# ============================================================================

@login_required
@never_cache
def cost_remuneration_distribution_list(request):
    """酬劳分配列表"""
    queryset = CostProjectUnified.objects.only('id', 'project_code', 'project_name', 'remuneration_total_remuneration', 'remuneration_distribution_status', 'created_at').all()
    queryset = filter_queryset_by_tenant(queryset, request)
    
    # 搜索
    search_key = request.GET.get('search', '')
    if search_key:
        queryset = queryset.filter(
            Q(project_code__icontains=search_key) |
            Q(project_name__icontains=search_key)
        )
    
    # 筛选
    calculation_type = request.GET.get('calculation_type', '')
    if calculation_type:
        queryset = queryset.filter(calculation_type=calculation_type)
    
    distribution_status = request.GET.get('distribution_status', '')
    if distribution_status:
        queryset = queryset.filter(distribution_status=distribution_status)
    
    # 动态字段筛选（来自右键菜单）
    for param_key in request.GET.keys():
        if param_key.startswith('filter_') and not param_key.endswith('_op'):
            field_name = param_key.replace('filter_', '', 1)
            operator = request.GET.get(f'{param_key}_op', 'contains')
            filter_value = request.GET.get(param_key, '').strip()
            
            if filter_value:
                # 根据操作符构建不同的查询
                if operator == 'contains':
                    queryset = queryset.filter(**{f'{field_name}__icontains': filter_value})
                elif operator == 'equals':
                    queryset = queryset.filter(**{field_name: filter_value})
                elif operator == 'starts_with':
                    queryset = queryset.filter(**{f'{field_name}__istartswith': filter_value})
                elif operator == 'ends_with':
                    queryset = queryset.filter(**{f'{field_name}__iendswith': filter_value})
                elif operator == 'not_contains':
                    queryset = queryset.exclude(**{f'{field_name}__icontains': filter_value})
    
    # 排序 (Django Admin 风格 - 支持多字段排序)
    sort_fields_str = request.GET.get('sort_field', 'created_at')
    sort_orders_str = request.GET.get('sort_order', 'desc')
    
    # 清除默认排序，确保自定义排序生效
    queryset = queryset.order_by()
    
    # 解析多字段排序（逗号分隔）
    if ',' in sort_fields_str:
        fields = [f.strip() for f in sort_fields_str.split(',')]
        orders = [o.strip() for o in sort_orders_str.split(',')]
        # 补齐 order 数量
        while len(orders) < len(fields):
            orders.append('asc')
        
        order_list = []
        for field, order in zip(fields, orders):
            if order == 'desc':
                order_list.append(f'-{field}')
            else:
                order_list.append(field)
        queryset = queryset.order_by(*order_list)
    else:
        # 单字段排序
        field = sort_fields_str.strip()
        order = sort_orders_str.strip() if sort_orders_str else 'asc'
        if order == 'desc':
            queryset = queryset.order_by(f'-{field}')
        else:
            queryset = queryset.order_by(field)
    
    # 分页
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 统计
    total_count = queryset.count()
    
    context = {
        'page_obj': page_obj,
        'search_key': search_key,
        'calculation_type': calculation_type,
        'distribution_status': distribution_status,
        'total_count': total_count,
        'CALC_TYPE_CHOICES': CostRemunerationDistribution.CALC_TYPE_CHOICES,
        'DISTRIBUTION_STATUS_CHOICES': CostRemunerationDistribution.DISTRIBUTION_STATUS_CHOICES,
        'show_detail': False,
        'current_project': None,
    }
    return render(request, 'cost_consulting/remuneration_distribution/list.html', context)


@login_required
def cost_remuneration_distribution_add(request):
    """新增酬劳分配"""
    if request.method == 'POST':
        form = CostRemunerationDistributionForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            if hasattr(request, 'tenant') and request.tenant:
                obj.tenant = request.tenant
            
            # 自动同步项目信息到冗余字段
            if obj.project:
                obj.project_code = obj.project.project_code
                obj.project_name = obj.project.project_name
                obj.project_type = obj.project.project_type
            
            obj.save()
            messages.success(request, '✓ 酬劳分配添加成功！')
            return redirect('eims_app:cost_remuneration_distribution_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CostRemunerationDistributionForm()
    
    context = {'form': form, 'title': '新增酬劳分配', 'action': 'add'}
    return render(request, 'cost_consulting/remuneration_distribution/form.html', context)


@login_required
def cost_remuneration_distribution_detail(request, pk):
    """酬劳分配详情"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    items = obj.remuneration_items.all().order_by('role', 'person_name')
    
    context = {
        'object': obj,
        'items': items,
    }
    return render(request, 'cost_consulting/remuneration_distribution/detail.html', context)


@login_required
def cost_remuneration_distribution_edit(request, pk):
    """编辑酬劳分配"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    
    if request.method == 'POST':
        form = CostRemunerationDistributionForm(request.POST, instance=obj)
        if form.is_valid():
            updated_obj = form.save()
            messages.success(request, '✓ 修改成功！')
            return redirect('eims_app:cost_remuneration_distribution_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CostRemunerationDistributionForm(instance=obj)
    
    context = {'form': form, 'object': obj, 'title': '编辑酬劳分配', 'action': 'edit'}
    return render(request, 'cost_consulting/remuneration_distribution/form.html', context)


@login_required
def cost_remuneration_distribution_delete(request, pk):
    """删除酬劳分配"""
    obj = get_object_or_404(CostProjectUnified, pk=pk)
    obj.delete()
    messages.success(request, '✓ 删除成功！')
    return redirect('eims_app:cost_remuneration_distribution_list')


@login_required
def cost_remuneration_distribution_batch_delete(request):
    """批量删除酬劳分配"""
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if ids:
            CostProjectUnified.objects.filter(id__in=ids).delete()
            messages.success(request, f'✓ 成功删除 {len(ids)} 条记录！')
    return redirect('eims_app:cost_remuneration_distribution_list')


@login_required
def cost_remuneration_distribution_export(request):
    """导出酬劳分配"""
    queryset = CostProjectUnified.objects.all()
    queryset = filter_queryset_by_tenant(queryset, request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "酬劳分配"
    
    headers = ['项目编号', '项目名称', '计算类型', '计算基准', '工程总造价(万元)',
               '审减金额(万元)', '酬劳总额(万元)', '计算式', '分配状态']
    ws.append(headers)
    
    for obj in queryset:
        ws.append([
            obj.project_code, obj.project_name,
            obj.get_calculation_type_display(),
            obj.get_calculation_base_display(),
            obj.total_cost, obj.reduced_amount, obj.total_remuneration,
            obj.calculation_formula, obj.get_distribution_status_display()
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="酬劳分配.xlsx"'
    wb.save(response)
    return response


# ============================================================================
# API 视图 - 用于 AJAX 请求
# ============================================================================

@login_required
def cost_project_info_api(request, pk):
    """获取项目信息 API - 用于表单自动填充"""
    try:
        project = get_object_or_404(CostProjectInfo, pk=pk)
        
        # 检查租户权限
        if hasattr(request, 'tenant') and request.tenant:
            if project.tenant != request.tenant:
                return JsonResponse({
                    'success': False,
                    'error': '无权访问该项目'
                }, status=403)
        
        data = {
            'success': True,
            'project_code': project.project_code,
            'project_name': project.project_name,
            'project_type': project.project_type,
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def get_unread_reminder_count(request):
    """获取未读提醒数量及详情"""
    from django.utils import timezone
    now = timezone.now()
    
    # 查询未读且不在延迟期内的提醒
    reminders = CostConsultingReminder.objects.filter(
        receiver=request.user,
        status='unread'
    ).filter(
        Q(snooze_until__isnull=True) | Q(snooze_until__lte=now)
    ).order_by('-created_at')[:5]  # 最多返回5条
    
    reminder_list = []
    for r in reminders:
        reminder_list.append({
            'id': r.pk,
            'title': r.title,
            'content': r.content[:100] + '...' if len(r.content) > 100 else r.content,
            'type': r.reminder_type,
            'link_url': r.link_url,
            'created_at': r.created_at.strftime('%Y-%m-%d %H:%M')
        })
    
    return JsonResponse({
        'count': reminders.count(),
        'reminders': reminder_list
    })


@login_required
def snooze_reminder(request):
    """延迟提醒（5分钟后再次提醒）"""
    import json
    from django.utils import timezone
    from datetime import timedelta
    
    try:
        data = json.loads(request.body)
        reminder_ids = data.get('reminder_ids', [])
        
        if not reminder_ids:
            return JsonResponse({'success': False, 'error': '未提供提醒ID'}, status=400)
        
        # 设置延迟时间为5分钟后
        snooze_time = timezone.now() + timedelta(minutes=5)
        
        updated_count = CostConsultingReminder.objects.filter(
            pk__in=reminder_ids,
            receiver=request.user
        ).update(snooze_until=snooze_time)
        
        return JsonResponse({'success': True, 'updated': updated_count})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
def ignore_reminder(request):
    """忽略提醒（本次登录期间不再提醒）"""
    import json
    from django.utils import timezone
    
    try:
        data = json.loads(request.body)
        reminder_ids = data.get('reminder_ids', [])
        session_id = data.get('session_id', '')
        
        if not reminder_ids:
            return JsonResponse({'success': False, 'error': '未提供提醒ID'}, status=400)
        
        # 更新提醒状态为已忽略，并记录会话ID
        updated_count = CostConsultingReminder.objects.filter(
            pk__in=reminder_ids,
            receiver=request.user
        ).update(
            status='ignored',
            ignored_session=session_id,
            read_at=timezone.now()
        )
        
        return JsonResponse({'success': True, 'updated': updated_count})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
def mark_reminder_read(request, pk):
    """标记提醒为已读"""
    try:
        reminder = CostConsultingReminder.objects.get(pk=pk, receiver=request.user)
        reminder.status = 'read'
        from django.utils import timezone
        reminder.read_at = timezone.now()
        reminder.save()
        return JsonResponse({'success': True})
    except CostConsultingReminder.DoesNotExist:
        return JsonResponse({'success': False, 'error': '提醒不存在'}, status=404)
