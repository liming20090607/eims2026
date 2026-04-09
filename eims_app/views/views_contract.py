import logging
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, FileResponse
from django.utils import timezone
from django.db.models import Q
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
import json
import datetime
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from eims_app.models.model_project_detail import ProjectDetail  # 改用 ProjectDetail 模型
from eims_app.forms.form_contract_management import ContractManagementForm  # 使用合同管理表单
from eims_app.utils.tenant_utils import filter_queryset_by_tenant
from django.urls import reverse 
from django.db import transaction
from django.db.models.deletion import ProtectedError
import csv
from django.http import HttpResponse
from django.contrib.auth.decorators import user_passes_test, login_required

def is_superuser(user):
    return user.is_superuser

def contract_detail(request, pk):
    """合同详情 - 改用 ProjectDetail 模型"""
    project = get_object_or_404(ProjectDetail, pk=pk)
    return render(request, 'contract_management/detail.html', {'project': project})

def contract_list(request):
    """合同管理列表 - 改用 ProjectDetail 模型"""
    # 1. 获取筛选参数（保留原有逻辑）
    status = request.GET.get('status', '')
    contract_type = request.GET.get('contract_type', '')
    keyword = request.GET.get('keyword', '')

    # 2. 基础查询集（使用辅助函数按租户过滤）
    queryset = ProjectDetail.objects.select_related().all().order_by('project_code')
    queryset = filter_queryset_by_tenant(queryset, request)

    # 3. 多条件筛选
    if status:
        queryset = queryset.filter(contract_status=status)
    if contract_type:
        queryset = queryset.filter(contract_category=contract_type)
    if keyword:
        # 增强关键字搜索
        queryset = queryset.filter(
            Q(project_name__icontains=keyword) |
            Q(contract_code__icontains=keyword) |
            Q(project_code__icontains=keyword) |
            Q(contract_category__icontains=keyword) |
            Q(contract_status__icontains=keyword) |
            Q(contract_party_a__icontains=keyword) |
            Q(contract_party_b__icontains=keyword) |
            Q(project_address__icontains=keyword) |
            Q(project_scale__icontains=keyword) |
            Q(remark__icontains=keyword)
        ).distinct()

    # 4. 分页处理
    paginator = Paginator(queryset, 20)
    page = request.GET.get('page')
    try:
        page_obj = paginator.get_page(page)
    except PageNotAnInteger:
        page_obj = paginator.get_page(1)
    except EmptyPage:
        page_obj = paginator.get_page(paginator.num_pages)
        
    # 5. 获取总记录数
    total_count = queryset.count()
        
    context = {
        'page_obj': page_obj,
        'search_key': keyword,
        'contract_status_filter': status,
        'contract_type_filter': contract_type,
        'total_count': total_count,  # 添加总记录数
    }
    
    return render(request, 'contract_management/list.html', context)


@user_passes_test(is_superuser)
def contract_add(request):
    """新增合同 - 改用 ProjectDetail 模型"""
    if request.method == 'POST':
        form = ContractManagementForm(request.POST, request.FILES)
        if form.is_valid():
            contract = form.save(commit=False)
            # 自动分配租户
            if hasattr(contract, 'tenant') and hasattr(request, 'tenant'):
                contract.tenant = request.tenant
            contract.save()
            messages.success(request, "✓ 合同添加成功！")
            return redirect('eims_app:contract_management_list')
        else:
            messages.error(request, "合同添加失败，请检查红色标注的输入项！")
    else:
        form = ContractManagementForm()
    return render(request, 'contract_management/form.html', {'form': form})

def contract_detail(request, pk):
    """合同详情 - 改用 ProjectDetail 模型"""
    project = get_object_or_404(ProjectDetail, pk=pk)
    field_data = []
    for field in project._meta.get_fields():
        if field.concrete and not field.many_to_many and not field.one_to_many:
            if field.name in ['id', 'created_at', 'updated_at']:
                continue
            if hasattr(field, 'choices') and field.choices:
                display_value = getattr(project, f'get_{field.name}_display')()
            else:
                display_value = getattr(project, field.name, None)
            label = getattr(field, 'verbose_name', field.name).title()
            field_type = field.get_internal_type()
            field_data.append({
                'name': field.name,
                'label': label,
                'value': display_value,
                'type': field_type,
                'has_choices': bool(getattr(field, 'choices', None))
            })
    return render(request, 'contract_management/detail.html', {
        'project': project,
        'field_data': field_data
    })

@user_passes_test(is_superuser)
def contract_edit(request, pk):
    """编辑合同 - 改用 ProjectDetail 模型"""
    project = get_object_or_404(ProjectDetail, pk=pk)
    if request.method == "POST":
        form = ContractManagementForm(request.POST, request.FILES, instance=project)
        if form.is_valid():
            form.save()
            messages.success(request, "✓ 合同修改成功！")
            return redirect("eims_app:contract_management_list")
        else:
            messages.error(request, "合同修改失败，请检查红色标注的输入项！")
    else:
        form = ContractManagementForm(instance=project)
    return render(request, "contract_management/form.html", {
        "form": form,
        "page_title": f"编辑合同：{project.contract_code}",
        "project": project
    })

@user_passes_test(is_superuser)
def contract_delete(request, pk):
    """删除合同 - 改用 ProjectDetail 模型"""
    project = get_object_or_404(ProjectDetail, pk=pk)
    contract_code = project.contract_code
    project.delete()
    messages.success(request, f"合同【{contract_code}】删除成功！")
    return redirect("eims_app:contract_management_list")

@user_passes_test(is_superuser)
def contract_batch_delete(request):
    """批量删除合同 - 改用 ProjectDetail 模型"""
    if request.method == 'POST':
        contract_ids = request.POST.getlist('contract_ids')
    else:
        contract_ids = request.GET.get('ids', '').split(',')
    
    contract_ids = [cid for cid in contract_ids if cid.isdigit()]
    
    if not contract_ids:
        messages.warning(request, '未选择任何合同记录')
        return redirect('eims_app:contract_management_list')
    
    try:
        with transaction.atomic():
            # 获取要删除的记录（用于显示成功消息）
            projects_to_delete = ProjectDetail.objects.filter(id__in=contract_ids)
            count = projects_to_delete.count()
            
            if count == 0:
                messages.warning(request, '未找到要删除的合同记录')
                return redirect('eims_app:contract_management_list')
            
            # 执行删除
            projects_to_delete.delete()
            
            messages.success(request, f'✓ 成功删除 {count} 条合同记录')
            
    except Exception as e:
        messages.error(request, f'删除失败：{str(e)}')
    
    return redirect('eims_app:contract_management_list')

@user_passes_test(is_superuser)
def contract_import(request):
    """合同导入 - 已弃用，重定向到合同管理导入"""
    messages.info(request, 'ℹ️ 请使用合同管理导入功能')
    return redirect('eims_app:contract_management_import')

@user_passes_test(is_superuser)
def contract_import_template(request):
    """
    下载导入模板文件
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "合同导入模板"
    
    headers = [
        "合同编号", "项目编号", "项目名称", "合同类型", "合同状态",
        "合同总价（元）", "合同甲方", "合同乙方", "合同文本",
        "付款约定", "项目规模", "项目投资（万元）", "项目地址",
        "约定人员配备", "服务期", "签订日期", "服务截止期",
        "计划开工时间", "预计竣工时间", "延期约定", "备注"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    sample_data = [
        "HT-20240001", "XM-20240001", "示例项目", "工程合同", "草稿",
        "1000000", "甲方公司", "乙方公司", "合同正文内容",
        "按进度付款", "1000平方米", "500", "北京市朝阳区",
        "5人", "12个月", "2024-01-01", "2024-12-31",
        "2024-03-01", "2024-12-31", "无", "这是示例备注"
    ]
    
    for col_idx, value in enumerate(sample_data, 1):
        ws.cell(row=2, column=col_idx, value=value)
    
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 18
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = FileResponse(
        buffer,
        as_attachment=True,
        filename='合同导入模板.xlsx',
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    return response

@user_passes_test(is_superuser)
def contract_export_selected(request):
    """
    导出选中的合同记录（修复版：精准过滤选中ID）
    支持GET/POST两种参数传递方式
    """
    # 【关键修复】正确解析选中ID
    if request.method == 'POST':
        contract_ids = request.POST.getlist('contract_ids')  # 复选框提交
    else:
        # GET请求：处理URL参数 ?ids=1,2,3
        ids_param = request.GET.get('ids', '')
        contract_ids = [cid.strip() for cid in ids_param.split(',') if cid.strip().isdigit()]
    
    # 【关键修复】验证ID有效性
    if not contract_ids:
        messages.warning(request, '⚠️ 未选择任何合同记录')
        return redirect('eims_app:contract_list')
    
    # 【关键修复】精准过滤选中记录
    contracts = Contract.objects.filter(id__in=contract_ids).order_by('id')
    
    # 验证实际查询结果
    if not contracts.exists():
        messages.warning(request, f'⚠️ 未找到有效合同记录 (请求ID: {len(contract_ids)}个)')
        return redirect('eims_app:contract_list')
    
    # 创建CSV响应
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')  # UTF-8 BOM for Excel
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'合同导出_{timestamp}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # 写入表头（根据实际模型字段调整）
    writer.writerow([
        '合同ID', '合同编号', '项目名称', '甲方单位', '乙方单位',
        '合同金额(元)', '签订日期', '状态', '创建时间'
    ])
    
    # 写入数据行
    for contract in contracts:
        writer.writerow([
            contract.id,
            contract.contract_code,
            contract.project.name if hasattr(contract, 'project') and contract.project else '无项目',
            contract.party_a,
            contract.party_b,
            f"{contract.amount:,.2f}" if contract.amount else '0.00',
            contract.sign_date.strftime('%Y-%m-%d') if contract.sign_date else '',
            contract.get_status_display(),
            contract.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    # 记录操作日志
    logger.info(
        f"用户 {request.user.username} 导出合同记录 | 请求ID数: {len(contract_ids)} | 实际导出: {contracts.count()} | 文件: {filename}"
    )
    
    messages.success(
        request, 
        f'✅ 成功导出 {contracts.count()} 条合同记录（共选择 {len(contract_ids)} 个ID）'
    )
    
    return respons

@user_passes_test(is_superuser)
def contract_export(request):
    ids_param = request.GET.get('ids', '')
    if ids_param:
        contract_ids = [cid.strip() for cid in ids_param.split(',') if cid.strip().isdigit()]
        if contract_ids:
            contracts = Contract.objects.filter(id__in=contract_ids).order_by('-signing_time')
        else:
            contracts = Contract.objects.all().prefetch_related().order_by('-signing_time')
    else:
        status_filter = request.GET.get("status", "")
        type_filter = request.GET.get("contract_type", "")
        start_date = request.GET.get("start_date", "")
        end_date = request.GET.get("end_date", "")
        keyword = request.GET.get("keyword", "")
        contracts = Contract.objects.all().prefetch_related().order_by('-signing_time')
        if status_filter:
            contracts = contracts.filter(status=status_filter)
        if type_filter:
            contracts = contracts.filter(contract_type=type_filter)
        if start_date:
            contracts = contracts.filter(signing_time__gte=start_date)
        if end_date:
            contracts = contracts.filter(signing_time__lte=end_date)
        if keyword:
            contracts = contracts.filter(
                Q(project_name__icontains=keyword) |
                Q(contract_code__icontains=keyword) |
                Q(contract_party_a__icontains=keyword) |
                Q(contract_party_b__icontains=keyword) |
                Q(project_address__icontains=keyword) |
                Q(remark__icontains=keyword) |
                Q(contract_type__icontains=keyword) |
                Q(contract_text__icontains=keyword) |
                Q(payment_agreement__icontains=keyword)
            )
    wb = Workbook()
    ws = wb.active
    ws.title = "合同信息表"
    headers = [
        "序号", "合同类型", "项目编号", "合同编号", "项目名称", "合同甲方", "合同乙方",
        "合同总价（万元）", "付款约定", "项目规模", "项目投资（万元）", "项目地址",
        "约定人员配备", "订立时间", "服务期", "服务截止期", "延期约定", "计划开工时间",
        "预计竣工时间", "合同状态", "备注"
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx, header in enumerate(headers, 1):
        if header in ["项目名称", "付款约定", "项目地址", "备注", "延期约定"]:
            ws.column_dimensions[chr(64+col_idx)].width = 25
        elif header in ["合同甲方", "合同乙方", "约定人员配备", "服务期"]:
            ws.column_dimensions[chr(64+col_idx)].width = 20
        else:
            ws.column_dimensions[chr(64+col_idx)].width = 15
    for row_idx, contract in enumerate(contracts, 2):
        ws.cell(row=row_idx, column=1, value=row_idx-1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=2, value=contract.get_contract_type_display() or "-").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=3, value=contract.project_code or "-").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=4, value=contract.contract_code or "-").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=5, value=contract.project_name or "-").alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row_idx, column=6, value=contract.contract_party_a or "-").alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row_idx, column=7, value=contract.contract_party_b or "-").alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row_idx, column=8, value=(contract.contract_amount or 0)/10000).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_idx, column=9, value=contract.payment_agreement or "-").alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row_idx, column=10, value=contract.project_scale or "-").alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row_idx, column=11, value=contract.project_investment or "-").alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_idx, column=12, value=contract.project_address or "-").alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row_idx, column=13, value=contract.agreed_staffing or "-").alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row_idx, column=14, value=contract.signing_time.strftime("%Y-%m-%d") if contract.signing_time else "-").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=15, value=contract.service_period_months or "-").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=16, value=contract.service_deadline.strftime("%Y-%m-%d") if contract.service_deadline else "-").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=17, value=contract.extension_agreement or "-").alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row_idx, column=18, value=contract.planned_start_time.strftime("%Y-%m-%d") if contract.planned_start_time else "-").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=19, value=contract.estimated_completion_time.strftime("%Y-%m-%d") if contract.estimated_completion_time else "-").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=20, value=contract.get_status_display() or "-").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=21, value=contract.remark or "-").alignment = Alignment(horizontal="left", vertical="center")
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = FileResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=合同信息表_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response


@login_required
def contract_approval_chain(request):
    """合同审批流程列表"""
    from eims_app.models.model_contract_approval import ContractApproval
    from django.db.models import Q
    
    # 获取筛选参数
    status = request.GET.get('status', '')
    keyword = request.GET.get('keyword', '')
    
    # 基础查询集
    queryset = ContractApproval.objects.select_related('applicant', 'department', 'current_approver').all()
    
    # 筛选
    if status:
        queryset = queryset.filter(status=status)
    if keyword:
        queryset = queryset.filter(
            Q(title__icontains=keyword) |
            Q(contract_name__icontains=keyword) |
            Q(party_b__icontains=keyword)
        )
    
    # 分页
    from django.core.paginator import Paginator
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'status': status,
        'keyword': keyword,
        'APPROVAL_STATUS_CHOICES': ContractApproval.APPROVAL_STATUS_CHOICES,
    }
    
    return render(request, 'contract_management/approval_chain_list.html', context)


@login_required
def contract_approval_add(request):
    """新增合同审批 - 必须上传附件"""
    from eims_app.models.model_contract_approval import ContractApproval, ContractAttachment
    from eims_app.forms.form_contract_approval import ContractApprovalForm, ContractAttachmentForm
    from django.utils import timezone
    
    if request.method == 'POST':
        form = ContractApprovalForm(request.POST)
        
        # 检查是否上传了附件（新增时必须上传）
        files = request.FILES.getlist('new_attachments')
        if not files or len(files) == 0:
            messages.error(request, '⚠️ 请至少上传一个附件才能提交')
            return render(request, 'contract_management/approval_form.html', {
                'form': form,
                'title': '新增合同审批',
            })
        
        if form.is_valid():
            # 保存审批信息
            approval = form.save(commit=False)
            approval.applicant = request.user
            approval.status = 'draft'  # 默认为草稿
            approval.save()
            
            # 处理文件上传
            file_types = request.POST.getlist('new_file_types')
            
            for i, file in enumerate(files):
                attachment = ContractAttachment(
                    approval=approval,
                    file=file,
                    file_type=file_types[i] if i < len(file_types) else 'contract'
                )
                attachment.save()
            
            messages.success(request, '合同审批创建成功！')
            return redirect('eims_app:contract_approval_detail', pk=approval.pk)
    else:
        form = ContractApprovalForm()
    
    context = {
        'form': form,
        'title': '新增合同审批',
    }
    
    return render(request, 'contract_management/approval_form.html', context)


@login_required
def contract_approval_detail(request, pk):
    """合同审批详情"""
    from eims_app.models.model_contract_approval import ContractApproval, ContractApprovalRecord
    from django.contrib.auth import get_user_model
    
    approval = get_object_or_404(ContractApproval, pk=pk)
    
    # 获取审批历史记录
    records = approval.approval_records.select_related('operator').all()
    
    # 获取可用的审批人列表（用于转发功能）
    User = get_user_model()
    available_approvers = User.objects.filter(is_active=True).exclude(pk=request.user.pk)
    
    context = {
        'approval': approval,
        'records': records,
        'available_approvers': available_approvers,
        'title': '合同审批详情',
    }
    
    return render(request, 'contract_management/approval_detail.html', context)


@login_required
def contract_approval_edit(request, pk):
    """编辑合同审批（草稿、已退回或已撤回状态可编辑，且仅申请人可编辑）"""
    from eims_app.models.model_contract_approval import ContractApproval, ContractAttachment
    from eims_app.forms.form_contract_approval import ContractApprovalForm, ContractAttachmentForm
    
    approval = get_object_or_404(ContractApproval, pk=pk)
    
    # 只有申请人可以编辑
    if approval.applicant != request.user:
        messages.error(request, '只有发起人可以编辑该审批')
        return redirect('eims_app:contract_approval_detail', pk=approval.pk)
    
    # 草稿、已退回或已撤回状态可以编辑
    if approval.status not in ['draft', 'rejected', 'cancelled']:
        messages.error(request, '只有草稿、已退回或已撤回状态的审批可以编辑')
        return redirect('eims_app:contract_approval_detail', pk=approval.pk)
    
    # 获取已有附件
    existing_attachments = approval.attachments.filter(is_deleted=False)
    
    if request.method == 'POST':
        form = ContractApprovalForm(request.POST, instance=approval)
        if form.is_valid():
            approval = form.save()
            
            # 处理附件：删除旧附件（如果用户勾选了删除），添加新附件
            # 删除操作
            delete_ids = request.POST.getlist('delete_attachments')
            if delete_ids:
                ContractAttachment.objects.filter(
                    id__in=delete_ids, 
                    approval=approval
                ).update(is_deleted=True)
            
            # 上传新附件
            new_files = request.FILES.getlist('new_attachments')
            new_file_types = request.POST.getlist('new_file_types')
            
            if new_files:
                for i, file in enumerate(new_files):
                    file_type = new_file_types[i] if i < len(new_file_types) else 'contract'
                    ContractAttachment.objects.create(
                        approval=approval,
                        file=file,
                        file_type=file_type,
                        uploaded_by=request.user
                    )
            
            # 根据操作类型处理
            action = request.POST.get('action', 'save')
            submit_comment = request.POST.get('submit_comment', '')
            
            if action == 'submit':
                # 保存并提交审批
                from eims_app.models.model_contract_approval import ContractApprovalRecord
                from django.utils import timezone
                
                # 自动填充发起人、发起时间和申请部门（如果是首次提交）
                if not approval.initiator:
                    approval.initiator = request.user
                    approval.initiation_time = timezone.now()
                    
                    # 如果申请部门为空，使用用户所在部门
                    if not approval.department and hasattr(request.user, 'department') and request.user.department:
                        approval.department = request.user.department
                
                # 根据审批流程类型指派审批人
                try:
                    assigned_approver = approval.assign_current_approver()
                    if assigned_approver:
                        approval.save()
                    else:
                        messages.warning(request, '未找到合适的审批人，请手动选择或联系管理员配置')
                        return redirect('eims_app:contract_approval_edit', pk=approval.pk)
                except Exception as e:
                    messages.error(request, f'指派审批人失败：{str(e)}')
                    return redirect('eims_app:contract_approval_edit', pk=approval.pk)
                
                # 更新状态
                approval.status = 'pending'
                approval.submitted_at = timezone.now()
                approval.save()
                
                # 记录操作
                ContractApprovalRecord.objects.create(
                    approval=approval,
                    action='submit',
                    operator=request.user,
                    comment=submit_comment or '提交审批'
                )
                
                messages.success(request, '合同审批已提交，等待审核')
            else:
                # 仅保存
                messages.success(request, '合同审批修改成功！')
            
            return redirect('eims_app:contract_approval_detail', pk=approval.pk)
    else:
        form = ContractApprovalForm(instance=approval)
    
    context = {
        'form': form,
        'approval': approval,
        'existing_attachments': existing_attachments,
        'title': '编辑合同审批',
    }
    
    return render(request, 'contract_management/approval_form.html', context)


@login_required
def contract_approval_submit(request, pk):
    """提交合同审批"""
    from eims_app.models.model_contract_approval import ContractApproval, ContractApprovalRecord
    from django.utils import timezone
    
    approval = get_object_or_404(ContractApproval, pk=pk)
    
    # 只有草稿或已退回状态可以提交
    if approval.status not in ['draft', 'rejected']:
        messages.error(request, '当前状态不能提交审批')
        return redirect('eims_app:contract_approval_detail', pk=approval.pk)
    
    # 自动填充发起人、发起时间和申请部门（如果是首次提交）
    if not approval.initiator:  # 只在第一次提交时设置
        approval.initiator = request.user
        approval.initiation_time = timezone.now()
        
        # 如果申请部门为空，使用用户所在部门
        if not approval.department and hasattr(request.user, 'department') and request.user.department:
            approval.department = request.user.department
    
    # 根据审批流程类型指派审批人
    try:
        assigned_approver = approval.assign_current_approver()
        if assigned_approver:
            approval.save()  # 保存指派的审批人和状态
        else:
            messages.warning(request, '未找到合适的审批人，请手动选择或联系管理员配置')
            return redirect('eims_app:contract_approval_edit', pk=approval.pk)
    except Exception as e:
        messages.error(request, f'指派审批人失败：{str(e)}')
        return redirect('eims_app:contract_approval_edit', pk=approval.pk)
    
    # 更新状态
    approval.status = 'pending'
    approval.submitted_at = timezone.now()
    approval.save()
    
    # 记录操作
    ContractApprovalRecord.objects.create(
        approval=approval,
        action='submit',
        operator=request.user,
        comment='提交审批'
    )
    
    messages.success(request, '合同审批已提交，等待审核')
    return redirect('eims_app:contract_approval_detail', pk=approval.pk)


@login_required
def contract_approval_approve(request, pk):
    """批准合同审批 - 可选择转发到下一步或终结审批"""
    from eims_app.models.model_contract_approval import ContractApproval, ContractApprovalRecord
    from eims_app.models.model_project_detail import ProjectDetail
    from eims_app.models.model_employee import Employee
    from django.utils import timezone
    import uuid
    
    approval = get_object_or_404(ContractApproval, pk=pk)
    
    # 检查权限：当前审批人才能审批
    if approval.current_approver != request.user:
        messages.error(request, '您不是当前审批人，无法进行审批操作')
        return redirect('eims_app:contract_approval_detail', pk=approval.pk)
    
    if request.method == 'POST':
        comment = request.POST.get('comment', '')
        submit_comment = request.POST.get('submit_comment', '')  # 转发时的意见
        action_type = request.POST.get('action_type', 'finalize')  # 'forward' 或 'finalize'
        
        # 记录审批操作
        ContractApprovalRecord.objects.create(
            approval=approval,
            action='approve',
            operator=request.user,
            comment=comment or '同意'
        )
        
        if action_type == 'forward':
            # 转发到下一步审批人
            assign_method = request.POST.get('assign_method', 'auto')  # 'auto' 或 'manual'
            
            if assign_method == 'manual':
                # 自主选择审批人
                next_approver_id = request.POST.get('next_approver')
                if next_approver_id:
                    from django.contrib.auth import get_user_model
                    User = get_user_model()
                    try:
                        next_approver = User.objects.get(pk=next_approver_id)
                        # 升级审批级别
                        approval.approval_level += 1
                        approval.current_approver = next_approver
                        approval.status = 'reviewing'  # 保持审核中状态
                        approval.save()
                        
                        # 记录操作并保存下一步审批人
                        ContractApprovalRecord.objects.create(
                            approval=approval,
                            action='approve',
                            operator=request.user,
                            comment=submit_comment or comment or f'同意并转发给 {next_approver.username}',
                            next_approver=next_approver
                        )
                        
                        messages.success(request, f'已转发给 {next_approver.username} 进行下一步审批')
                    except User.DoesNotExist:
                        messages.error(request, '选择的审批人不存在')
                        return redirect('eims_app:contract_approval_detail', pk=approval.pk)
                else:
                    messages.error(request, '请选择下一步审批人')
                    return redirect('eims_app:contract_approval_detail', pk=approval.pk)
            else:
                # 系统指定审批人 - 根据流程自动指派
                try:
                    # 升级审批级别
                    approval.approval_level += 1
                    assigned_approver = approval.assign_current_approver()
                    
                    if assigned_approver:
                        approval.status = 'reviewing'  # 保持审核中状态
                        approval.save()
                        
                        # 记录操作
                        ContractApprovalRecord.objects.create(
                            approval=approval,
                            action='approve',
                            operator=request.user,
                            comment=submit_comment or comment or f'同意并由系统指派给 {assigned_approver.username}'
                        )
                        
                        messages.success(request, f'审批通过，系统已指派给 {assigned_approver.username} 进行下一步审批')
                    else:
                        messages.warning(request, '未找到合适的下一位审批人，请手动选择或联系管理员配置')
                        return redirect('eims_app:contract_approval_detail', pk=approval.pk)
                except Exception as e:
                    messages.error(request, f'系统指派审批人失败：{str(e)}')
                    return redirect('eims_app:contract_approval_detail', pk=approval.pk)
        else:
            # 终结审批，生成合同台账
            approval.status = 'approved'
            approval.approved_at = timezone.now()
            approval.approval_result = 'pending'  # 待签订
            approval.save()
            
            # 记录操作
            ContractApprovalRecord.objects.create(
                approval=approval,
                action='approve',
                operator=request.user,
                comment=submit_comment or comment or '同意'
            )
            
            # 自动生成合同台账记录
            project_code = f"HT{timezone.now().strftime('%Y%m%d%H%M%S')}"
            contract_code = f"HT-{timezone.now().strftime('%Y%m%d')}-{approval.pk}"
            
            project_detail = ProjectDetail.objects.create(
                project_code=project_code,
                contract_code=contract_code,
                project_name=approval.contract_name,
                contract_category=approval.contract_category,
                contract_amount=approval.contract_amount,
                contract_party_a=approval.party_a,
                contract_party_b=approval.party_b,
                service_period_months=approval.service_period_months,
                service_deadline=approval.service_deadline,
                signing_date=None,  # 未签订
                project_status='not_started',
                contract_status='pending_review',
                settlement_status='unsettled',
                remark=f"来自审批：{approval.title}. {approval.remark or ''}",
            )
            
            # 关联审批和台账
            approval.generated_contract = project_detail
            approval.save()
            
            messages.success(request, '合同审批已通过，已自动生成合同台账记录')
    
    return redirect('eims_app:contract_approval_detail', pk=approval.pk)


@login_required
def contract_approval_reject(request, pk):
    """退回合同审批"""
    from eims_app.models.model_contract_approval import ContractApproval, ContractApprovalRecord
    
    approval = get_object_or_404(ContractApproval, pk=pk)
    
    # 检查权限：当前审批人才能退回
    if approval.current_approver != request.user:
        messages.error(request, '您不是当前审批人，无法进行退回操作')
        return redirect('eims_app:contract_approval_detail', pk=approval.pk)
    
    if request.method == 'POST':
        comment = request.POST.get('comment', '')
        
        # 更新状态
        approval.status = 'rejected'
        approval.save()
        
        # 记录操作
        ContractApprovalRecord.objects.create(
            approval=approval,
            action='reject',
            operator=request.user,
            comment=comment or '退回'
        )
        
        messages.success(request, '合同审批已退回')
    
    return redirect('eims_app:contract_approval_detail', pk=approval.pk)


@login_required
def contract_approval_cancel(request, pk):
    """撤销合同审批（申请人可撤销自己提交的审批）"""
    from eims_app.models.model_contract_approval import ContractApproval, ContractApprovalRecord
    
    approval = get_object_or_404(ContractApproval, pk=pk)
    
    # 只有申请人可以撤销
    if approval.applicant != request.user:
        messages.error(request, '只有申请人可以撤销审批')
        return redirect('eims_app:contract_approval_detail', pk=approval.pk)
    
    # 只有待审核或审核中状态可以撤销
    if approval.status not in ['pending', 'reviewing']:
        messages.error(request, '当前状态不能撤销')
        return redirect('eims_app:contract_approval_detail', pk=approval.pk)
    
    # 更新状态
    approval.status = 'cancelled'
    approval.save()
    
    # 记录操作
    ContractApprovalRecord.objects.create(
        approval=approval,
        action='cancel',
        operator=request.user,
        comment='撤销审批'
    )
    
    messages.success(request, '合同审批已撤销')
    return redirect('eims_app:contract_approval_detail', pk=approval.pk)