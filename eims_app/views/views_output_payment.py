# E:\EIMS2026\eims_app\views\views_output_payment.py
# 产值回款视图

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.generic import ListView, CreateView, UpdateView, DetailView
from django.urls import reverse_lazy
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils.decorators import method_decorator
from decimal import Decimal, ROUND_HALF_UP
from ..models.model_output_payment import OutputPayment
from ..models.model_project_detail import ProjectDetail
from ..forms.form_output_payment import OutputForm


def is_superuser(user):
    """检查用户是否为超级管理员"""
    return user.is_superuser


class OutputPaymentListView(ListView):
    """产值回款列表页"""
    model = OutputPayment
    template_name = 'output_payment/output_payment_list.html'
    context_object_name = 'output_payments'
    ordering = ['-month', '-create_time']
    paginate_by = 15
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = '产值回款'
        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()
        # 搜索功能
        search = self.request.GET.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(project_code__icontains=search) |
                Q(month__icontains=search)
            )
        return queryset


class OutputPaymentCreateView(CreateView):
    """创建产值回款"""
    model = OutputPayment
    form_class = OutputForm
    template_name = 'output_payment/output_payment_add.html'
    success_url = reverse_lazy('eims_app:output_payment_list')
    
    @method_decorator(login_required)
    @method_decorator(user_passes_test(is_superuser))
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request)
    
    def form_valid(self, form):
        form.instance.operator = self.request.user.username
        messages.success(self.request, '成功添加产值回款')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, '添加失败，请检查表单数据')
        return super().form_invalid(form)


class OutputPaymentDetailView(DetailView):
    """产值回款详情页"""
    model = OutputPayment
    template_name = 'output_payment/output_payment_detail.html'
    context_object_name = 'output_payment'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = '产值回款详情'
        return context


class OutputPaymentUpdateView(UpdateView):
    """更新产值回款"""
    model = OutputPayment
    form_class = OutputForm
    template_name = 'output_payment/output_payment_edit.html'
    success_url = reverse_lazy('eims_app:output_payment_list')
    
    @method_decorator(login_required)
    @method_decorator(user_passes_test(is_superuser))
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request)
    
    def form_valid(self, form):
        form.instance.operator = self.request.user.username
        messages.success(self.request, '成功更新产值回款')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, '更新失败，请检查表单数据')
        return super().form_invalid(form)


@login_required
@user_passes_test(is_superuser)
def add_output(request, pk):
    """添加产值回款 - 新页面，自动从上个月数据填充"""
    project = get_object_or_404(ProjectDetail, pk=pk)
    
    if request.method == 'POST':
        # 获取表单数据
        last_month_cumulative_output = parse_decimal(request.POST.get('last_month_cumulative_output', 0))
        current_month_output = parse_decimal(request.POST.get('current_month_output', 0))
        last_month_cumulative_payment = parse_decimal(request.POST.get('last_month_cumulative_payment', 0))
        current_month_payment = parse_decimal(request.POST.get('current_month_payment', 0))
        month = request.POST.get('month', '2026-01')
        
        # 自动计算
        current_month_cumulative_output = last_month_cumulative_output + current_month_output
        current_month_cumulative_payment = last_month_cumulative_payment + current_month_payment
        contract_balance = parse_decimal(request.POST.get('contract_total', 0)) - current_month_cumulative_payment
        
        # 检查是否已存在该月份的记录
        existing_output = OutputPayment.objects.filter(project=project, month=month).first()
        
        if existing_output:
            messages.warning(request, f'{month} 的产值回款记录已存在，将为您更新记录')
            # 更新现有记录
            existing_output.monthly_output = current_month_output
            existing_output.cumulative_output = current_month_cumulative_output
            existing_output.contract_total = parse_decimal(request.POST.get('contract_total', 0))
            existing_output.cumulative_received = current_month_cumulative_payment
            existing_output.contract_receivable = parse_decimal(request.POST.get('contract_total', 0))
            existing_output.near_term_receivable = contract_balance
            existing_output.actual_payment = current_month_payment
            existing_output.recent_payment_request = request.POST.get('recent_payment_request', '')
            existing_output.payment_measures = request.POST.get('payment_measures', '')
            existing_output.next_month_request = request.POST.get('next_month_request', '')
            existing_output.need_assistance = request.POST.get('need_assistance', '')
            existing_output.operator = request.user.username
            existing_output.save()
        else:
            # 创建新记录
            output = OutputPayment(
                project=project,
                project_code=project.project_code,
                month=month,
                monthly_output=current_month_output,
                cumulative_output=current_month_cumulative_output,
                contract_total=parse_decimal(request.POST.get('contract_total', 0)),
                cumulative_received=current_month_cumulative_payment,
                contract_receivable=parse_decimal(request.POST.get('contract_total', 0)),
                near_term_receivable=contract_balance,
                actual_payment=current_month_payment,
                recent_payment_request=request.POST.get('recent_payment_request', ''),
                payment_measures=request.POST.get('payment_measures', ''),
                next_month_request=request.POST.get('next_month_request', ''),
                need_assistance=request.POST.get('need_assistance', ''),
                operator=request.user.username
            )
            output.save()
            messages.success(request, '成功添加产值回款')
        
        # 更新项目信息中的累计回款和合同余款
        project.cumulative_payment = current_month_cumulative_payment
        project.contract_balance = contract_balance
        project.save(update_fields=['cumulative_payment', 'contract_balance'])
        
        return redirect('eims_app:project_ledger_detail', pk=pk)
    
    # GET 请求时，获取上月产值回款记录用于预填充
    last_month_output = OutputPayment.objects.filter(
        project=project
    ).order_by('-month', '-create_time').first()
    
    context = {
        'project': project,
        # 从上月产值回款预填充（如果有）
        'last_monthly_output': last_month_output.monthly_output if last_month_output else 0,
        'last_cumulative_output': last_month_output.cumulative_output if last_month_output else 0,
        'last_monthly_payment': last_month_output.actual_payment if last_month_output else 0,
        'last_cumulative_payment': last_month_output.cumulative_received if last_month_output else 0,
        'contract_total': project.contract_amount if hasattr(project, 'contract_amount') else 0,
        'last_recent_payment_request': last_month_output.recent_payment_request if last_month_output else '',
        'last_payment_measures': last_month_output.payment_measures if last_month_output else '',
        'last_next_month_request': last_month_output.next_month_request if last_month_output else '',
        'last_need_assistance': last_month_output.need_assistance if last_month_output else '',
    }
    return render(request, 'project_ledger/add_output.html', context)


@login_required
def delete_output(request, pk):
    """删除产值回款 - 仅超级管理员可用"""
    
    # 检查超级管理员权限
    if not request.user.is_superuser:
        messages.error(request, '⚠️ 权限不足：只有超级管理员才能删除记录')
        return HttpResponseRedirect(reverse_lazy('eims_app:project_ledger_detail', args=[pk]))
    
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if ids:
            OutputPayment.objects.filter(pk__in=ids).delete()
        messages.success(request, f'成功删除 {len(ids)} 条产值回款')
    
    return HttpResponseRedirect(reverse_lazy('eims_app:project_ledger_detail', args=[pk]))


@login_required
@user_passes_test(is_superuser)
def import_output_payment(request, pk):
    """导入产值回款"""
    from openpyxl import load_workbook
    
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, '请选择要上传的文件')
            return HttpResponseRedirect(reverse_lazy('eims_app:project_ledger_detail', args=[pk]))
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            required_fields = []
            
            for field in required_fields:
                if field not in headers:
                    messages.error(request, f'Excel 文件缺少必填列：{field}')
                    return HttpResponseRedirect(reverse_lazy('eims_app:project_ledger_detail', args=[pk]))
            
            project = get_object_or_404(ProjectDetail, pk=pk)
            success_count = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                row_data = dict(zip(headers, row))
                
                output = OutputPayment(
                    project=project,
                    project_code=project.project_code,
                    month=str(row_data.get('月份', '2026-01')),
                    monthly_output=parse_decimal(row_data.get('当月产值 (万元)', 0)),
                    cumulative_output=parse_decimal(row_data.get('累计产值 (万元)', 0)),
                    contract_total=parse_decimal(row_data.get('合同总额 (元)', 0)),
                    cumulative_received=parse_decimal(row_data.get('累计已收款 (元)', 0)),
                    contract_receivable=parse_decimal(row_data.get('合同应收款 (元)', 0)),
                    near_term_receivable=parse_decimal(row_data.get('近期待收款 (元)', 0)),
                    payment_basis=str(row_data.get('合同付款依据', '')) or '',
                    last_payment_situation=str(row_data.get('上次回款情况', '')) or '',
                    recent_payment_request=str(row_data.get('近期请款情况', '')) or '',
                    actual_payment=parse_decimal(row_data.get('本月实际回款 (元)', 0)),
                    next_month_request=str(row_data.get('下个月请款', '')) or '',
                    next_month_plan=parse_decimal(row_data.get('下月计划收款 (元)', 0)),
                    payment_measures=str(row_data.get('请款措施', '')) or '',
                    need_assistance=str(row_data.get('需要协助', '')) or '',
                    remark=str(row_data.get('备注', '')) or '',
                    payment_date=parse_date(row_data.get('回款日期')) if row_data.get('回款日期') else None,
                    payment_method=str(row_data.get('回款方式', '')) or '',
                    output_amount=parse_decimal(row_data.get('当月产值 (万元)', 0)),
                    payment_amount=parse_decimal(row_data.get('本月实际回款 (元)', 0)),
                    operator=request.user.username if request.user.is_authenticated else ''
                )
                output.save()
                success_count += 1
            
            messages.success(request, f'成功导入 {success_count} 条产值回款')
        except Exception as e:
            messages.error(request, f'导入失败：{str(e)}')
        
        return HttpResponseRedirect(reverse_lazy('eims_app:project_ledger_detail', args=[pk]))
    
    return HttpResponseRedirect(reverse_lazy('eims_app:project_ledger_detail', args=[pk]))


def parse_decimal(value, default=0):
    """解析小数值"""
    try:
        if value is None or value == '':
            return Decimal(default)
        return Decimal(str(value))
    except:
        return Decimal(default)


def parse_date(value):
    """解析日期值"""
    from datetime import datetime
    if not value:
        return None
    try:
        if isinstance(value, str):
            return datetime.strptime(value, '%Y-%m-%d').date()
        return value
    except:
        return None
