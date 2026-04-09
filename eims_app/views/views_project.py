from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import Q
from django.utils import timezone
from django.http import HttpResponseRedirect
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from eims_app.forms.form_project import ProjectForm
from eims_app.forms.form_project_ledger import ProjectLedgerForm  # 新增：项目台账表单
from eims_app.models.model_project_detail import ProjectDetail  # 改用 ProjectDetail 模型
from eims_app.models.model_project_dynamic import ProjectDynamic  # 项目动态模型
from eims_app.models.model_contract import Contract
from django.core.paginator import Paginator
from django.contrib.auth.decorators import user_passes_test, login_required
from django.utils.decorators import method_decorator
from django.http import JsonResponse
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
from datetime import datetime
from eims_app.utils.tenant_utils import filter_queryset_by_tenant  # 租户过滤工具

def is_superuser(user):
    return user.is_superuser

def parse_date(value):
    """解析日期"""
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y%m%d']:
            try:
                return datetime.strptime(value, fmt).date()
            except:
                continue
    return None

def parse_decimal(value):
    """解析数字"""
    if value is None or value == '':
        return 0
    try:
        return float(value)
    except:
        return 0

class ProjectListView(ListView):
    """项目列表 - 重定向到项目台账页面（使用 ProjectDetail 统一数据源）"""
    model = ProjectDetail
    template_name = 'project_ledger/list.html'  # 使用项目台账模板
    context_object_name = 'page_obj'
    
    def get(self, request, *args, **kwargs):
        response = super().get(request, *args, **kwargs)
        # 防止浏览器缓存
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # 使用 ProjectDetail 的字段
        context['status_choices'] = ProjectDetail.PROJECT_STATUS_CHOICES
        # 注意：ProjectDetail 没有 category 字段，使用 contract_category
        context['category_choices'] = getattr(ProjectDetail, 'CONTRACT_CATEGORY_CHOICES', [])
        
        selected_status = self.request.GET.get('status', '')
        selected_category = self.request.GET.get('category', '')
        keyword = self.request.GET.get('keyword', '')
        show_detail = self.request.GET.get('show_detail', '')
        
        # 查询所有记录（不区分模块），并应用租户过滤
        queryset = ProjectDetail.objects.select_related().all()
        queryset = filter_queryset_by_tenant(queryset, self.request)  # 租户过滤
        
        if selected_status:
            queryset = queryset.filter(project_status=selected_status)
            context['selected_status_label'] = dict(ProjectDetail.PROJECT_STATUS_CHOICES).get(selected_status, '')
        if selected_category:
            queryset = queryset.filter(contract_category=selected_category)
            context['selected_category_label'] = dict(getattr(ProjectDetail, 'CONTRACT_CATEGORY_CHOICES', [])).get(selected_category, '')
        if keyword:
            queryset = queryset.filter(
                Q(project_name__icontains=keyword) |
                Q(project_code__icontains=keyword) |
                Q(contract_code__icontains=keyword) |
                Q(project_address__icontains=keyword) |
                Q(project_director__icontains=keyword) |
                Q(project_manager__icontains=keyword) |
                Q(remark__icontains=keyword)
            )
        
        paginator = Paginator(queryset, 10)
        page_number = self.request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
        # 获取总记录数
        total_count = queryset.count()
        
        # 如果 show_detail 参数存在，获取第一个项目用于显示详情
        first_project = None
        if show_detail:
            first_project = queryset.order_by('-created_at').first()
        
        context.update({
            'page_obj': page_obj,
            'selected_status': selected_status,
            'selected_category': selected_category,
            'keyword': keyword,
            'show_detail': show_detail,
            'first_project': first_project,
            'total_count': total_count,  # 添加总记录数
        })
        
        return context

class ProjectCreateView(CreateView):
    """项目创建 - 改用 ProjectDetail 模型"""
    model = ProjectDetail
    form_class = ProjectLedgerForm  # 使用项目台账表单
    template_name = 'project_ledger/form.html'  # 使用项目台账模板
    success_url = reverse_lazy('eims_app:project_ledger_list')  # 重定向到项目台账列表
    
    @method_decorator(user_passes_test(is_superuser))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def form_valid(self, form):
        form.instance.created_at = timezone.now()
        form.instance.updated_at = timezone.now()
        messages.success(self.request, '✓ 项目创建成功！')
        return super().form_valid(form)

class ProjectUpdateView(UpdateView):
    """项目更新 - 改用 ProjectDetail 模型"""
    model = ProjectDetail
    form_class = ProjectLedgerForm  # 使用项目台账表单
    template_name = 'project_ledger/form.html'  # 使用项目台账模板
    success_url = reverse_lazy('eims_app:project_ledger_list')  # 重定向到项目台账列表
    
    @method_decorator(user_passes_test(is_superuser))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def form_valid(self, form):
        form.instance.updated_at = timezone.now()
        messages.success(self.request, '✓ 项目更新成功！')
        return super().form_valid(form)

class ProjectDeleteView(DeleteView):
    """项目删除 - 改用 ProjectDetail 模型"""
    model = ProjectDetail
    template_name = 'project_ledger/delete.html'  # 使用项目台账模板
    success_url = reverse_lazy('eims_app:project_ledger_list')  # 重定向到项目台账列表
    
    @method_decorator(user_passes_test(is_superuser))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def form_valid(self, form):
        messages.success(self.request, '✓ 项目删除成功！')
        return super().form_valid(form)

class ProjectDetailView(DetailView):
    """项目详情 - 改用 ProjectDetail 模型"""
    model = ProjectDetail
    template_name = 'project/detail.html'  # 模板路径保持不变
    context_object_name = 'project'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project = self.get_object()
        # related_contracts 属性在 ProjectDetail 模型中可能不存在，需要检查
        # 如果不存在，可以注释掉或移除
        # context['related_contracts'] = getattr(project, 'related_contracts', None)
        context['all_projects_list'] = ProjectDetail.objects.order_by('project_code')
        context['all_contracts_list'] = Contract.objects.exclude(contract_code__isnull=True).exclude(contract_code='').order_by('contract_code')
        
        all_projects = list(ProjectDetail.objects.order_by('project_code'))
        project_ids = [p.id for p in all_projects]
        current_idx = project_ids.index(project.id)
        
        context['total_projects'] = len(all_projects)
        context['current_index'] = current_idx + 1
        
        if all_projects:
            context['first_project'] = all_projects[0]
            context['last_project'] = all_projects[-1]
        
        if current_idx > 0:
            context['prev_project'] = all_projects[current_idx - 1]
        if current_idx < len(all_projects) - 1:
            context['next_project'] = all_projects[current_idx + 1]
        
        return context

@user_passes_test(is_superuser)
def project_batch_delete(request):
    """批量删除项目 - 改用 ProjectDetail 模型"""
    if request.method == 'POST':
        project_ids = request.POST.getlist('project_ids')
        if project_ids:
            ProjectDetail.objects.filter(id__in=project_ids).delete()
            messages.success(request, f'✓ 成功删除 {len(project_ids)} 个项目！')
        return redirect('eims_app:project_ledger_list')  # 重定向到项目台账
    return redirect('eims_app:project_ledger_list')

def project_by_contract(request):
    """根据合同编号获取项目 ID - 改用 ProjectDetail 模型"""
    contract_code = request.GET.get('code', '')
    if contract_code:
        project = ProjectDetail.objects.filter(project_code=contract_code).first()
        if project:
            return JsonResponse({'id': project.id, 'name': project.project_name})
    return JsonResponse({'error': 'Project not found'}, status=404)

@user_passes_test(is_superuser)
def project_export(request):
    """项目数据导出 - 改用 ProjectDetail 模型"""
    ids_param = request.GET.get('ids', '')
    
    if ids_param:
        project_ids = [int(id) for id in ids_param.split(',') if id.isdigit()]
        projects = ProjectDetail.objects.filter(id__in=project_ids)
    else:
        projects = ProjectDetail.objects.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "项目数据"
    
    # 使用 ProjectDetail 的字段
    headers = ['项目编号', '合同编号', '项目名称', '合同类别', '项目状态', '合同状态', 
               '合同甲方', '合同乙方', '签订日期', '合同总价 (元)', '项目地址', 
               '现场负责人', '项目总监', '备注']
    ws.append(headers)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    status_map = dict(ProjectDetail.PROJECT_STATUS_CHOICES)
    contract_status_map = dict(ProjectDetail.CONTRACT_STATUS_CHOICES)
    category_map = dict(getattr(ProjectDetail, 'CONTRACT_CATEGORY_CHOICES', {}))
    
    for p in projects:
        ws.append([
            p.project_code or '',
            p.contract_code or '',
            p.project_name or '',
            category_map.get(p.contract_category, p.contract_category or ''),
            status_map.get(p.project_status, p.project_status or ''),
            contract_status_map.get(p.contract_status, p.contract_status or ''),
            p.contract_party_a or '',
            p.contract_party_b or '',
            p.signing_date.strftime('%Y-%m-%d') if p.signing_date else '',
            str(p.contract_amount) if p.contract_amount else '',
            p.project_address or '',
            p.project_manager or '',
            p.project_director or '',
            p.remark or ''
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=项目台账数据.xlsx'
    return response


@user_passes_test(is_superuser)
def project_import(request):
    """项目导入 - 已弃用，重定向到项目台账导入"""
    messages.info(request, 'ℹ️ 请使用项目台账导入功能')
    return redirect('eims_app:project_ledger_import')


@login_required
def import_project_dynamic(request, pk):
    """导入项目动态"""
    from eims_app.models.model_project_dynamic import ProjectDynamic
    from openpyxl import load_workbook
    
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, '请选择要上传的文件')
            return HttpResponseRedirect(reverse_lazy('eims_app:project_ledger_detail', args=[pk]))
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            required_fields = ['项目编号', '项目进度', '项目状态']
            
            for field in required_fields:
                if field not in headers:
                    messages.error(request, f'Excel 文件缺少必填列：{field}')
                    return HttpResponseRedirect(reverse_lazy('eims_app:project_ledger_detail', args=[pk]))
            
            project = get_object_or_404(Project, pk=pk)
            success_count = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                row_data = dict(zip(headers, row))
                
                dynamic = ProjectDynamic(
                    project=project,
                    project_code=project.project_code,
                    project_progress=str(row_data.get('项目进度', '')) or '',
                    project_status=str(row_data.get('项目状态', '')) or '',
                    notice_entry=parse_date(row_data.get('通知进场')),
                    delay_status=str(row_data.get('延期情况', '')) or '',
                    planned_start_time=parse_date(row_data.get('计划开工时间')),
                    actual_start_time=parse_date(row_data.get('实际开工时间')),
                    planned_completion=parse_date(row_data.get('预计竣工时间')),
                    remark=str(row_data.get('备注', '')) or '',
                    operator=request.user.username if request.user.is_authenticated else ''
                )
                dynamic.save()
                success_count += 1
            
            messages.success(request, f'成功导入 {success_count} 条项目动态')
        except Exception as e:
            messages.error(request, f'导入失败：{str(e)}')
        
        return HttpResponseRedirect(reverse_lazy('eims_app:project_ledger_detail', args=[pk]))
    
    return HttpResponseRedirect(reverse_lazy('eims_app:project_ledger_detail', args=[pk]))


@user_passes_test(is_superuser)
def import_personnel(request, pk):
    """导入项目人员"""
    from eims_app.models import Personnel
    from openpyxl import load_workbook
    
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, '请选择要上传的文件')
            return HttpResponseRedirect(reverse_lazy('eims_app:project_ledger_detail', args=[pk]))
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            required_fields = ['人员编号', '姓名']
            
            for field in required_fields:
                if field not in headers:
                    messages.error(request, f'Excel文件缺少必填列：{field}')
                    return HttpResponseRedirect(reverse_lazy('eims_app:project_ledger_detail', args=[pk]))
            
            project = get_object_or_404(Project, pk=pk)
            success_count = 0
            
            gender_map = {'男': 0, '女': 1, '其他': 2}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                row_data = dict(zip(headers, row))
                
                phone = str(row_data.get('手机号码', ''))
                if not phone or phone == 'None':
                    phone = f'138{pk:04d}{success_count:04d}'
                
                personnel = Personnel(
                    personnel_code=str(row_data.get('人员编号', '')),
                    project=project,
                    project_code=project.project_code,
                    name=str(row_data.get('姓名', '')),
                    gender=gender_map.get(row_data.get('性别', '男'), 0),
                    position=str(row_data.get('岗位', '')) or '',
                    phone=phone,
                    department=str(row_data.get('部门', '')) or '',
                    entry_time=parse_date(row_data.get('入岗时间')),
                    leave_time=parse_date(row_data.get('离岗时间')),
                    email=str(row_data.get('邮箱', '')) or '',
                    remark=str(row_data.get('备注', '')) or '',
                    operator=request.user.username if request.user.is_authenticated else ''
                )
                try:
                    personnel.save()
                    success_count += 1
                except:
                    phone = f'138{pk:04d}{success_count:05d}'
                    personnel.phone = phone
                    personnel.save()
                    success_count += 1
            
            messages.success(request, f'成功导入 {success_count} 条项目人员')
        except Exception as e:
            messages.error(request, f'导入失败：{str(e)}')
        
        return HttpResponseRedirect(reverse_lazy('eims_app:project_ledger_detail', args=[pk]))
    
    return HttpResponseRedirect(reverse_lazy('eims_app:project_ledger_detail', args=[pk]))


@login_required
def delete_dynamic(request, pk):
    """删除项目动态 - 仅超级管理员可用"""
    from eims_app.models.model_project_dynamic import ProjectDynamic
    
    # 检查超级管理员权限
    if not request.user.is_superuser:
        messages.error(request, '⚠️ 权限不足：只有超级管理员才能删除记录')
        return HttpResponseRedirect(reverse_lazy('eims_app:project_ledger_detail', args=[pk]))
    
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if ids:
            ProjectDynamic.objects.filter(pk__in=ids).delete()
        messages.success(request, f'成功删除 {len(ids)} 条项目动态')
    
    return HttpResponseRedirect(reverse_lazy('eims_app:project_ledger_detail', args=[pk]))


@login_required
def delete_personnel(request, pk):
    """删除项目人员 - 仅超级管理员可用"""
    from eims_app.models.model_personnel import Personnel
    
    # 检查超级管理员权限
    if not request.user.is_superuser:
        messages.error(request, '⚠️ 权限不足：只有超级管理员才能删除记录')
        return HttpResponseRedirect(reverse_lazy('eims_app:project_ledger_detail', args=[pk]))
    
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if ids:
            Personnel.objects.filter(pk__in=ids).delete()
        messages.success(request, f'成功删除 {len(ids)} 条项目人员')
    
    return HttpResponseRedirect(reverse_lazy('eims_app:project_ledger_detail', args=[pk]))

@login_required
def add_dynamic(request, pk):
    """
    新增项目动态 - 动态信息保存到 ProjectDynamic，
    但进场时间等字段直接更新到 ProjectDetail（项目信息总表）
    """
    
    project = get_object_or_404(ProjectDetail, pk=pk)
    
    if request.method == 'POST':
        # 1. 创建项目动态记录
        dynamic = ProjectDynamic(
            project=project,
            project_code=project.project_code,
            project_progress=request.POST.get('project_progress', ''),
            project_status=request.POST.get('project_status', ''),
            notice_entry=parse_date(request.POST.get('notice_entry')),
            delay_status=request.POST.get('delay_status', ''),
            planned_start_time=parse_date(request.POST.get('planned_start_time')),
            actual_start_time=parse_date(request.POST.get('actual_start_time')),
            planned_completion=parse_date(request.POST.get('planned_completion')),
            operator=request.user.username
        )
        dynamic.save()
        
        # 2. 同时更新 ProjectDetail（项目信息总表）中的对应字段
        # 通知进场日期
        notice_entry_date = parse_date(request.POST.get('notice_entry'))
        if notice_entry_date:
            project.entry_time = notice_entry_date
        
        # 计划开工日期
        planned_start = parse_date(request.POST.get('planned_start_time'))
        if planned_start:
            project.planned_start_date = planned_start
        
        # 实际开工日期
        actual_start = parse_date(request.POST.get('actual_start_time'))
        if actual_start:
            project.actual_start_date = actual_start
        
        # 预计竣工日期
        planned_end = parse_date(request.POST.get('planned_completion'))
        if planned_end:
            project.estimated_completion_date = planned_end
        
        # 保存更新到总表
        project.save(update_fields=['entry_time', 'planned_start_date', 
                                    'actual_start_date', 'estimated_completion_date'])
        
        messages.success(request, '✓ 成功添加项目动态，并已更新项目信息总表')
        return redirect('eims_app:project_ledger_detail', pk=pk)
    
    # GET 请求时，从 ProjectDetail 和上月动态预填充时间字段
    # 获取上月的动态记录（用于预填充）
    last_month_dynamic = ProjectDynamic.objects.filter(
        project_code=project.project_code
    ).order_by('-create_time').first()
    
    context = {
        'project': project,
        # 从项目总表预填充
        'entry_time': project.entry_time,
        'planned_start_date': project.planned_start_date,
        'actual_start_date': project.actual_start_date,
        'estimated_completion_date': project.estimated_completion_date,
        # 从上月动态预填充（如果有）
        'last_project_progress': last_month_dynamic.project_progress if last_month_dynamic else '',
        'last_project_status': last_month_dynamic.project_status if last_month_dynamic else '',
        'last_delay_status': last_month_dynamic.delay_status if last_month_dynamic else '',
        # 系统信息
        'current_time': timezone.now(),
    }
    return render(request, 'project_ledger/add_dynamic.html', context)


@login_required
def add_personnel(request, pk):
    """添加项目人员 - 新页面"""
    from eims_app.models.model_personnel import Personnel
    from eims_app.models.model_project_detail import ProjectDetail
    from eims_app.models.model_employee import Employee  # 导入员工模型
    
    project = get_object_or_404(ProjectDetail, pk=pk)
    
    # 检查是否是编辑模式
    edit_id = request.GET.get('edit_id')
    editing_personnel = None
    if edit_id:
        try:
            editing_personnel = Personnel.objects.get(pk=edit_id)
        except Personnel.DoesNotExist:
            pass
    
    # GET 请求时，获取人员花名册中的所有姓名（按姓名升序排列，去重）
    # 从 Personnel 表（人员花名册）获取姓名，而不是 Employee 表
    personnel_names = Personnel.objects.filter(
        is_deleted=False  # 只获取未删除的记录
    ).order_by('name').values_list('name', flat=True).distinct()
    
    # 转换为列表（去重后的姓名列表）
    employee_names = list(personnel_names)
    
    if request.method == 'POST':
        try:
            # 获取项目信息
            project_code = project.project_code
            
            # 🔍 调试：打印所有 POST 数据
            print("\n=== POST Data ===")
            print(f"POST keys: {list(request.POST.keys())}")
            for key, value in request.POST.items():
                if 'has_change' in key or 'name_' in key or 'csrf' in key:
                    print(f"{key}: {value}")
            print("=================\n")
            
            # 生成人员编号和电话
            personnel_code = request.POST.get('personnel_code', f'RY{project_code}_{Personnel.objects.filter(project_code=project_code).count() + 1:03d}')
            
            phone = request.POST.get('phone', '')
            if not phone:
                phone = f'138{pk:04d}0000'
            
            # 创建人员记录（监理团队各岗位）
            positions = [
                ('director', '总监'),
                ('deputy_director', '总代'),
                ('civil_supervisor', '土建专监'),
                ('electrical_supervisor', '水电专监'),
                ('supervisor', '监理员'),
                ('document_controller', '资料员'),
                ('witness', '见证员'),
                ('safety_officer', '安全员')
            ]
            
            created_count = 0
            for position_key, position_name in positions:
                # 🔍 调试：检查每个岗位的提交数据
                has_change_raw = request.POST.get(f'has_change_{position_key}', '')
                name = request.POST.get(f'name_{position_key}', '')
                print(f"{position_key}: has_change='{has_change_raw}', name='{name}'")
                
                # ✅ 复选框 checked 时会发送 'on' 或其他非空值
                has_change = bool(has_change_raw and has_change_raw.strip())
                
                if has_change and name:  # 勾选了变化且填写了姓名
                    personnel = Personnel(
                        personnel_code=personnel_code,
                        project=project,  # 关联项目对象
                        project_code=project_code,  # 项目编号字符串
                        name=name,
                        gender=int(request.POST.get(f'gender_{position_key}', 0)),
                        phone=phone,
                        position=position_name,
                        department=request.POST.get('department', ''),
                        entry_time=parse_date(request.POST.get(f'entry_time_{position_key}')),
                        leave_time=parse_date(request.POST.get(f'leave_time_{position_key}')),
                        email=request.POST.get(f'email_{position_key}', ''),
                        remark=request.POST.get(f'remark_{position_key}', ''),
                        operator=request.user.username
                    )
                    personnel.save()
                    created_count += 1
                    print(f"✓ Created personnel: {name}, project_id={project.id}, project_code={project_code}")
            
            if created_count == 0:
                messages.warning(request, '未勾选任何岗位的变化，或未填写姓名')
            else:
                messages.success(request, f'成功添加 {created_count} 名项目人员')
        except Exception as e:
            import traceback
            print(f"❌ 错误详情：{str(e)}")
            print(traceback.format_exc())
            messages.error(request, f'添加失败：{str(e)}')
        
        # 无论成功还是失败，都跳转到项目详情页
        return redirect('eims_app:project_ledger_detail', pk=pk)
    
    context = {
        'project': project,
        'employee_names': employee_names,  # 人员姓名列表（倒序）
        'editing_personnel': editing_personnel,  # 正在编辑的人员记录
        'edit_id': edit_id,  # 编辑 ID
        # 系统信息
        'current_time': timezone.now(),
    }
    return render(request, 'project_ledger/add_personnel.html', context)
