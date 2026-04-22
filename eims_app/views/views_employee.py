import logging
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator
from eims_app.models import Employee
from eims_app.forms.form_employee import EmployeeForm
from django.contrib.auth.decorators import user_passes_test
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime

def is_superuser(user):
    return user.is_superuser

@user_passes_test(is_superuser)
def employee_list(request):
    """员工信息列表页面 - 支持筛选和搜索"""
    
    # 如果是 /root/ 路径且没有选择公司，重定向到公司选择页面
    if hasattr(request, 'current_system') and request.current_system == 'root':
        if not hasattr(request, 'tenant') or not request.tenant:
            messages.warning(request, '请先选择要查看的公司')
            return redirect('eims_app:tenant_select')
    
    # 1. 获取筛选参数
    search_key = request.GET.get('keyword', '')
    education = request.GET.get('education', '')
    ethnic = request.GET.get('ethnic', '')
    
    # 2. 基础查询集（按租户过滤）
    if hasattr(request, 'tenant') and request.tenant:
        # 使用 tenant_id 而不是 tenant 对象来避免跨数据库 JOIN
        employee_list = Employee.objects.filter(is_deleted=False, tenant_id=request.tenant.id).order_by('personnel_code')
    else:
        employee_list = Employee.objects.filter(is_deleted=False).order_by('personnel_code')
    
    # 3. 多条件筛选
    if search_key:
        employee_list = employee_list.filter(
            Q(name__icontains=search_key) |
            Q(personnel_code__icontains=search_key) |
            Q(mobile__icontains=search_key) |
            Q(id_card__icontains=search_key) |
            Q(native_place__icontains=search_key) |
            Q(admin_position__icontains=search_key) |
            Q(tech_position__icontains=search_key)
        ).distinct()
    
    if education:
        employee_list = employee_list.filter(education=education)
    
    if ethnic:
        employee_list = employee_list.filter(ethnic=ethnic)
    
    # 4. 分页
    page = request.GET.get('page', 1)
    paginator = Paginator(employee_list, 10)  # 每页显示 10 条
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    # 5. 获取筛选选项
    EDUCATION_CHOICES = Employee.EDUCATION_CHOICES
    ETHNIC_CHOICES = Employee.ETHNIC_CHOICES
    
    context = {
        'page_obj': page_obj,
        'search_key': search_key,
        'selected_education': education,
        'selected_ethnic': ethnic,
        'EDUCATION_CHOICES': EDUCATION_CHOICES,
        'ETHNIC_CHOICES': ETHNIC_CHOICES,
        'page_title': '员工信息管理'
    }
    
    return render(request, 'employee/list.html', context)

@user_passes_test(is_superuser)
def employee_add(request):
    """添加员工信息"""
    
    if request.method == "POST":
        form = EmployeeForm(request.POST)
        if form.is_valid():
            employee = form.save(commit=False)
            employee.operator = request.user.username
            # 自动分配租户
            if hasattr(employee, 'tenant') and hasattr(request, 'tenant'):
                employee.tenant = request.tenant
            employee.save()
            messages.success(request, "员工信息添加成功！")
            return redirect("eims_app:employee_list")
        else:
            messages.error(request, "员工信息添加失败，请检查红色标注的输入项！")
    else:
        form = EmployeeForm()
    
    return render(request, "employee/add.html", {
        "form": form,
        "page_title": "添加员工"
    })

@user_passes_test(is_superuser)
def employee_detail(request, pk):
    """员工详情页面"""
    
    # 如果是 /root/ 路径且没有选择公司，重定向到公司选择页面
    if hasattr(request, 'current_system') and request.current_system == 'root':
        if not hasattr(request, 'tenant') or not request.tenant:
            messages.warning(request, '请先选择要查看的公司')
            return redirect('eims_app:tenant_select')
    
    # 按租户过滤，防止跨租户访问
    if hasattr(request, 'tenant') and request.tenant:
        employee = get_object_or_404(Employee, pk=pk, is_deleted=False, tenant_id=request.tenant.id)
    else:
        employee = get_object_or_404(Employee, pk=pk, is_deleted=False)
    
    # 提取字段信息用于模板显示
    field_data = []
    for field in employee._meta.fields:
        value = getattr(employee, field.name)
        
        # 处理选择字段
        display_value = value
        if field.choices and value:
            choices_dict = dict(field.choices)
            display_value = choices_dict.get(value, value)
        
        label = getattr(field, 'verbose_name', field.name).title()
        field_data.append({
            'name': field.name,
            'label': label,
            'value': display_value,
            'type': field.get_internal_type(),
            'has_choices': bool(getattr(field, 'choices', None))
        })
    
    return render(request, 'employee/detail.html', {
        'employee': employee,
        'field_data': field_data
    })

@user_passes_test(is_superuser)
def employee_edit(request, pk):
    """编辑员工信息"""
    
    # 如果是 /root/ 路径且没有选择公司，重定向到公司选择页面
    if hasattr(request, 'current_system') and request.current_system == 'root':
        if not hasattr(request, 'tenant') or not request.tenant:
            messages.warning(request, '请先选择要查看的公司')
            return redirect('eims_app:tenant_select')
    
    # 按租户过滤，防止跨租户访问
    if hasattr(request, 'tenant') and request.tenant:
        employee = get_object_or_404(Employee, pk=pk, is_deleted=False, tenant_id=request.tenant.id)
    else:
        employee = get_object_or_404(Employee, pk=pk, is_deleted=False)
    
    if request.method == "POST":
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            employee = form.save(commit=False)
            employee.operator = request.user.username
            employee.save()
            messages.success(request, "员工信息修改成功！")
            # 从人员花名册进入的，返回人员花名册；否则返回员工列表
            next_url = request.GET.get('next', 'eims_app:personnel_list')
            return redirect(next_url)
        else:
            messages.error(request, "员工信息修改失败，请检查红色标注的输入项！")
    else:
        form = EmployeeForm(instance=employee)
    
    return render(request, "employee/edit.html", {
        "form": form,
        "page_title": f"编辑员工：{employee.name}",
        "employee": employee
    })

@user_passes_test(is_superuser)
def employee_delete(request, pk):
    """删除员工（软删除）"""
    
    # 如果是 /root/ 路径且没有选择公司，重定向到公司选择页面
    if hasattr(request, 'current_system') and request.current_system == 'root':
        if not hasattr(request, 'tenant') or not request.tenant:
            messages.warning(request, '请先选择要查看的公司')
            return redirect('eims_app:tenant_select')
    
    # 按租户过滤，防止误删其他公司的员工
    if hasattr(request, 'tenant') and request.tenant:
        employee = get_object_or_404(Employee, pk=pk, is_deleted=False, tenant_id=request.tenant.id)
    else:
        employee = get_object_or_404(Employee, pk=pk, is_deleted=False)
    
    try:
        employee.is_deleted = True
        employee.save()
        messages.success(request, f"员工 {employee.name} 删除成功！")
    except Exception as e:
        messages.error(request, f"删除失败：{str(e)}")
    
    # 从当前URL提取系统前缀（如 /root/, /dingce/, /shengchang/, /jiachengda/）
    path_parts = request.path.strip('/').split('/')
    system_prefix = path_parts[0] if path_parts and path_parts[0] else 'root'
    
    # 构建人员花名册的URL路径
    next_url = f'/{system_prefix}/personnel/'
    return redirect(next_url)

@user_passes_test(is_superuser)
def employee_batch_delete(request):
    """批量删除员工"""
    if request.method == "POST":
        employee_ids = request.POST.getlist('employee_ids')
        
        if not employee_ids:
            messages.error(request, "请至少选择一个员工！")
            return redirect("eims_app:employee_list")
        
        try:
            # 按租户过滤，防止误删其他公司的员工
            if hasattr(request, 'tenant') and request.tenant:
                count = Employee.objects.filter(id__in=employee_ids, tenant_id=request.tenant.id).update(is_deleted=True)
            else:
                count = Employee.objects.filter(id__in=employee_ids).update(is_deleted=True)
            messages.success(request, f"成功删除 {count} 个员工记录！")
        except Exception as e:
            messages.error(request, f"批量删除失败：{str(e)}")
    
    return redirect("eims_app:employee_list")


@user_passes_test(is_superuser)
def employee_export(request):
    """导出员工信息为 Excel（包含模型所有字段）- 支持全部导出或按 ID 列表导出"""
    
    # 获取筛选参数
    search_key = request.GET.get('keyword', '')
    education = request.GET.get('education', '')
    ethnic = request.GET.get('ethnic', '')
    
    # 获取选中的 ID 列表（POST 请求）
    selected_ids = None
    if request.method == 'POST':
        selected_ids = request.POST.getlist('employee_ids')
    
    # 基础查询集（只导出未删除的，按租户过滤）
    if hasattr(request, 'tenant') and request.tenant:
        queryset = Employee.objects.filter(is_deleted=False, tenant_id=request.tenant.id).order_by('personnel_code')
    else:
        queryset = Employee.objects.filter(is_deleted=False).order_by('personnel_code')
    
    # 如果有选中的 ID，只导出这些
    if selected_ids:
        queryset = queryset.filter(id__in=selected_ids)
    else:
        # 否则应用筛选条件
        if search_key:
            queryset = queryset.filter(
                Q(name__icontains=search_key) |
                Q(personnel_code__icontains=search_key) |
                Q(mobile__icontains=search_key) |
                Q(id_card__icontains=search_key) |
                Q(native_place__icontains=search_key) |
                Q(admin_position__icontains=search_key) |
                Q(tech_position__icontains=search_key)
            ).distinct()
        
        if education:
            queryset = queryset.filter(education=education)
        
        if ethnic:
            queryset = queryset.filter(ethnic=ethnic)
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "员工花名册"
    
    # 定义表头（按模型字段顺序）
    headers = [
        '员工编号', '姓名', '性别', '身份证号', '籍贯', '民族', '学历',
        '住址', '固定电话', '手机号', '应急联系人', '应急电话', '微信',
        '行政职务', '技术职务', '执业资格', '职称', '任职资格',
        '入职时间', '离职时间',
        '操作人', '创建时间', '更新时间', '备注'
    ]
    
    # 设置表头样式
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 字段映射（英文 → 中文）
    field_mapping = {
        'personnel_code': '人员编号',
        'name': '姓名',
        'gender': '性别',
        'id_card': '身份证号',
        'native_place': '籍贯',
        'ethnic': '民族',
        'education': '学历',
        'address': '住址',
        'home_phone': '固定电话',
        'mobile': '手机号',
        'emergency_contact': '应急联系人',
        'emergency_phone': '应急电话',
        'wechat': '微信',
        'admin_position': '行政职务',
        'tech_position': '技术职务',
        'professional_qualification': '执业资格',
        'professional_title': '职称',
        'job_qualification': '任职资格',
        'entry_time': '入职时间',
        'leave_time': '离职时间',
        'operator': '操作人',
        'create_time': '创建时间',
        'update_time': '更新时间',
        'remark': '备注',
    }
    
    # 获取选择字段的字典
    gender_dict = dict(Employee.GENDER_CHOICES)
    ethnic_dict = dict(Employee.ETHNIC_CHOICES)
    education_dict = dict(Employee.EDUCATION_CHOICES)
    
    # 填充数据
    for row_idx, employee in enumerate(queryset, 2):
        # 员工编号
        ws.cell(row=row_idx, column=1, value=employee.personnel_code or '')
        # 姓名
        ws.cell(row=row_idx, column=2, value=employee.name or '')
        # 性别（转换为中文）
        gender_value = gender_dict.get(employee.gender, employee.gender) if employee.gender is not None else ''
        ws.cell(row=row_idx, column=3, value=gender_value)
        # 身份证号
        ws.cell(row=row_idx, column=4, value=employee.id_card or '')
        # 籍贯
        ws.cell(row=row_idx, column=5, value=employee.native_place or '')
        # 民族（转换为中文）
        ethnic_value = ethnic_dict.get(employee.ethnic, employee.ethnic) if employee.ethnic else ''
        ws.cell(row=row_idx, column=6, value=ethnic_value)
        # 学历（转换为中文）
        education_value = education_dict.get(employee.education, employee.education) if employee.education else ''
        ws.cell(row=row_idx, column=7, value=education_value)
        # 住址
        ws.cell(row=row_idx, column=8, value=employee.address or '')
        # 固定电话
        ws.cell(row=row_idx, column=9, value=employee.home_phone or '')
        # 手机号
        ws.cell(row=row_idx, column=10, value=employee.mobile or '')
        # 应急联系人
        ws.cell(row=row_idx, column=11, value=employee.emergency_contact or '')
        # 应急电话
        ws.cell(row=row_idx, column=12, value=employee.emergency_phone or '')
        # 微信
        ws.cell(row=row_idx, column=13, value=employee.wechat or '')
        # 行政职务
        ws.cell(row=row_idx, column=14, value=employee.admin_position or '')
        # 技术职务
        ws.cell(row=row_idx, column=15, value=employee.tech_position or '')
        # 执业资格
        ws.cell(row=row_idx, column=16, value=employee.professional_qualification or '')
        # 职称
        ws.cell(row=row_idx, column=17, value=employee.professional_title or '')
        # 任职资格
        ws.cell(row=row_idx, column=18, value=employee.job_qualification or '')
        # 入职时间
        entry_time_str = employee.entry_time.strftime('%Y-%m-%d') if employee.entry_time else ''
        ws.cell(row=row_idx, column=19, value=entry_time_str)
        # 离职时间
        leave_time_str = employee.leave_time.strftime('%Y-%m-%d') if employee.leave_time else ''
        ws.cell(row=row_idx, column=20, value=leave_time_str)
        # 操作人
        ws.cell(row=row_idx, column=21, value=employee.operator or '')
        # 创建时间
        create_time_str = employee.create_time.strftime('%Y-%m-%d %H:%M:%S') if employee.create_time else ''
        ws.cell(row=row_idx, column=22, value=create_time_str)
        # 更新时间
        update_time_str = employee.update_time.strftime('%Y-%m-%d %H:%M:%S') if employee.update_time else ''
        ws.cell(row=row_idx, column=23, value=update_time_str)
        # 备注
        ws.cell(row=row_idx, column=24, value=employee.remark or '')
        
        # 设置整行样式
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
    
    # 调整列宽
    column_widths = [
        15,  # 员工编号
        12,  # 姓名
        8,   # 性别
        20,  # 身份证号
        15,  # 籍贯
        8,   # 民族
        10,  # 学历
        30,  # 住址
        15,  # 固定电话
        15,  # 手机号
        12,  # 应急联系人
        15,  # 应急电话
        15,  # 微信
        20,  # 行政职务
        20,  # 技术职务
        25,  # 执业资格
        20,  # 职称
        25,  # 任职资格
        12,  # 入职时间
        12,  # 离职时间
        12,  # 操作人
        20,  # 创建时间
        20,  # 更新时间
        30,  # 备注
    ]
    
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 写入内存并返回
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 设置响应头
    filename = f'员工花名册_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
