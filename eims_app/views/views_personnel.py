import logging
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, FileResponse
from django.utils import timezone
from django.db.models import Q
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
import json
import datetime
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from eims_app.models import Personnel, Employee, Department
from eims_app.utils.tenant_utils import filter_queryset_by_tenant
from eims_app.models.model_project_detail import ProjectDetail  # 改用 ProjectDetail
from eims_app.forms.form_personnel import PersonnelForm
from django.urls import reverse 
from django.db import transaction
import csv
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test

# 导入拼音转换库
try:
    from pypinyin import pinyin, Style
    
    def get_pinyin_key(text):
        """
        将中文文本转换为拼音排序关键字
        按照拼音字母顺序排序（先按声母，再按韵母）
        """
        if not text:
            return ''
        # 将中文字符串转换为拼音列表
        # Style.NORMAL: 不带声调，如 'zhong'
        # heteronym=False: 不返回多音字的所有读音
        pinyin_list = pinyin(text, style=Style.NORMAL, heteronym=False)
        # 将拼音列表扁平化为字符串
        # 例如：['张'], ['三'] -> [['zhang'], ['san']] -> 'zhangsan'
        pinyin_str = ''.join([p[0] for p in pinyin_list])
        return pinyin_str.lower()
except ImportError:
    # 如果pypinyin未安装，使用默认排序
    def get_pinyin_key(text):
        """备选方案：使用Unicode编码排序"""
        return text if text else ''

def is_superuser(user):
    return user.is_superuser

def has_personnel_permission(user):
    """检查用户是否具有人员管理权限（超级管理员或具有相应权限）"""
    if user.is_superuser:
        return True
    # 检查是否具有 Personnel 模型的查看或修改权限
    return user.has_perm('eims_app.view_personnel') or user.has_perm('eims_app.change_personnel')

@login_required
@user_passes_test(has_personnel_permission)
def personnel_list(request):
    """人员花名册页面 - 显示公司所有人员信息（包含Employee和Personnel）"""
    from eims_app.models import Employee, Personnel
    
    # 如果是 /root/ 路径且没有选择公司，重定向到公司选择页面
    if hasattr(request, 'current_system') and request.current_system == 'root':
        if not hasattr(request, 'tenant') or not request.tenant:
            from django.contrib import messages
            messages.warning(request, '请先选择要查看的公司')
            return redirect('eims_app:tenant_select')
    
    # 1. 获取筛选参数
    search_key = request.GET.get('keyword', '')
    project_code = request.GET.get('project_code', '')
    department = request.GET.get('department', '')
    position = request.GET.get('position', '')
    
    # 2. 获取排序参数
    sort_field = request.GET.get('sort_field', 'personnel_code')  # 默认按编号排序
    sort_order = request.GET.get('sort_order', 'asc')  # 默认升序
    
    # 3. 获取租户信息
    tenant_id = None
    if hasattr(request, 'tenant') and request.tenant:
        tenant_id = request.tenant.id
    
    # 3. 构建合并的人员列表
    # 策略：以Personnel（人员去向/项目分配）为主，补充未分配的Employee
    
    # 3.1 获取所有Employee
    if tenant_id:
        all_employees = Employee.objects.filter(is_deleted=False, tenant_id=tenant_id).order_by('personnel_code')
    else:
        all_employees = Employee.objects.filter(is_deleted=False).order_by('personnel_code')
    
    # 构建Employee字典，方便快速查找
    employee_dict = {emp.id: emp for emp in all_employees}
    
    # 3.2 获取当前租户的所有Personnel（人员去向表）
    if tenant_id:
        all_personnels = Personnel.objects.filter(
            is_deleted=False,
            tenant_id=tenant_id
        ).order_by('personnel_code')
    else:
        all_personnels = Personnel.objects.filter(
            is_deleted=False
        ).order_by('personnel_code')
    
    # 3.3 获取已分配项目的Employee IDs
    assigned_employee_ids = set(
        all_personnels.exclude(employee__isnull=True).values_list('employee_id', flat=True).distinct()
    )
    
    # 4. 合并数据：以Personnel为主，补充未分配的Employee
    # 创建一个包装类来统一数据结构
    class PersonnelWrapper:
        """包装类，统一Employee和Personnel的数据结构"""
        def __init__(self, source, source_type='employee'):
            self._source = source
            self._source_type = source_type
            
            if source_type == 'employee':
                # Employee模型
                self.id = source.id
                self.pk = source.id
                self.source_type = 'employee'  # 公开属性，用于模板访问
                self.personnel_code = source.personnel_code
                self.employee_code = source.personnel_code  # 别名，用于兼容模板和搜索
                self.name = source.name
                self.gender = source.gender
                self.id_card = source.id_card
                self.native_place = source.native_place
                self.ethnic = source.ethnic
                self.education = source.education
                self.department = '-'  # Employee没有部门字段
                self.admin_position = source.admin_position
                self.tech_position = source.tech_position
                self.professional_qualification = source.professional_qualification
                self.professional_title = source.professional_title
                self.job_qualification = source.job_qualification
                self.mobile = source.mobile
                self.phone = source.mobile  # 别名
                self.home_phone = source.home_phone
                self.entry_time = source.entry_time
                self.leave_time = source.leave_time
                self.address = source.address
                self.emergency_contact = source.emergency_contact
                self.emergency_phone = source.emergency_phone
                self.wechat = source.wechat
                self.email = source.email
                self.remark = source.remark
                self.tenant = source.tenant
                self.project = None
                self.project_code = ''
            else:
                # Personnel模型（独立人员）
                self.id = source.id
                self.pk = source.id
                self.source_type = 'personnel'  # 公开属性，用于模板访问
                self.personnel_code = source.personnel_code
                self.employee_code = source.personnel_code  # 别名，用于兼容模板和搜索
                self.name = source.name
                self.gender = source.gender
                self.id_card = ''  # Personnel没有身份证
                self.native_place = ''
                self.ethnic = 'han'
                self.education = 'bachelor'
                self.department = source.department
                self.admin_position = ''
                self.tech_position = ''
                self.professional_qualification = ''
                self.professional_title = ''
                self.job_qualification = ''
                self.mobile = source.phone
                self.phone = source.phone
                self.home_phone = ''
                self.entry_time = source.entry_time
                self.leave_time = source.leave_time
                self.address = ''
                self.emergency_contact = ''
                self.emergency_phone = ''
                self.wechat = ''
                self.email = source.email
                self.remark = source.remark
                self.tenant = source.tenant
                self.project = source.project
                self.project_code = source.project_code
        
        def __str__(self):
            return f"{self.personnel_code} - {self.name}"
    
    # 5. 合并列表
    personnel_list = []
    
    # 5.1 添加所有Personnel（人员去向表中的所有人员）
    for per in all_personnels:
        wrapper = PersonnelWrapper(per, 'personnel')
        personnel_list.append(wrapper)
    
    # 5.2 添加未分配项目的Employee（不在Personnel中的纯员工）
    for emp in all_employees:
        if emp.id not in assigned_employee_ids:
            # 检查是否已通过Personnel显示（通过姓名判断）
            is_duplicate = False
            for existing in personnel_list:
                if existing.name == emp.name:
                    is_duplicate = True
                    break
            
            if not is_duplicate:
                personnel_list.append(PersonnelWrapper(emp, 'employee'))
    
    # 6. 搜索过滤
    if search_key:
        filtered_list = []
        for person in personnel_list:
            if (search_key.lower() in person.name.lower() or
                search_key.lower() in person.employee_code.lower() or
                search_key.lower() in str(person.mobile) or
                search_key.lower() in str(person.id_card) or
                search_key.lower() in str(person.email or '') or
                search_key.lower() in str(person.remark or '') or
                search_key.lower() in str(person.department or '')):
                filtered_list.append(person)
        personnel_list = filtered_list
    
    # 7. 部门筛选
    if department:
        personnel_list = [p for p in personnel_list if p.department == department]
    
    # 8. 排序处理
    def get_sort_key(person):
        """获取排序键值，处理None值和中文拼音排序"""
        field_map = {
            'employee_code': person.employee_code or '',
            'name': person.name or '',
            'gender': person.gender if person.gender is not None else 999,
            'id_card': person.id_card or '',
            'native_place': person.native_place or '',
            'ethnic': person.ethnic or '',
            'education': person.education or '',
            'department': person.department or '',
            'admin_position': person.admin_position or '',
            'tech_position': person.tech_position or '',
            'professional_qualification': person.professional_qualification or '',
            'professional_title': person.professional_title or '',
            'job_qualification': person.job_qualification or '',
            'mobile': person.mobile or '',
            'home_phone': person.home_phone or '',
            'entry_time': str(person.entry_time) if person.entry_time else '',
            'leave_time': str(person.leave_time) if person.leave_time else '',
            'address': person.address or '',
            'emergency_contact': person.emergency_contact or '',
            'emergency_phone': person.emergency_phone or '',
            'wechat': person.wechat or '',
            'email': person.email or '',
        }
        return field_map.get(sort_field, '')
    
    # 执行排序（支持中文拼音排序）
    reverse_order = (sort_order == 'desc')
    try:
        personnel_list.sort(key=get_sort_key, reverse=reverse_order)
    except Exception as e:
        # 如果排序失败，保持原有顺序
        pass
    
    # 9. 分页处理
    paginator = Paginator(personnel_list, 13)  # 每页显示 13 条
    page = request.GET.get('page')
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    except Exception:
        page_obj = paginator.page(1)
    
    # 10. 统计信息
    total_personnel = len(personnel_list)
    
    context = {
        "page_obj": page_obj,
        "selected_keyword": search_key,
        "selected_project_code": project_code,
        "selected_department": department,
        "selected_position": position,
        'sort_field': sort_field,  # 当前排序字段
        'sort_order': sort_order,  # 当前排序方向
        'home_url': reverse('eims_app:eims_index'),
        'eims_index_url': reverse('eims_app:eims_index'),
        'total_personnel': total_personnel,
        'active_personnel': 0,  # Employee 模型不使用 active 状态
        'all_projects': ProjectDetail.objects.order_by('project_code'),
        'all_departments': Department.objects.filter(is_deleted=False, status='active').order_by('department_code'),
    }
    return render(request, "personnel/list.html", context)


@login_required
@user_passes_test(has_personnel_permission)
def personnel_navigation(request):
    """人员管理模块导航页面"""
    context = {
        'home_url': reverse('eims_app:eims_index'),
        'eims_index_url': reverse('eims_app:eims_index'),
    }
    return render(request, "personnel/navigation.html", context)

@login_required
@user_passes_test(has_personnel_permission)
def personnel_add(request):
    """添加人员"""
    # 获取当前租户信息
    tenant = getattr(request, 'tenant', None)
    
    # 根据公司代码设置人员编号前缀和公司全称
    personnel_prefix = ''
    company_full_name = ''
    
    if tenant:
        company_code = tenant.code
        company_full_name = tenant.name
        
        # 根据公司代码设置前缀
        if company_code == 'dingce' or '鼎策' in company_full_name:
            personnel_prefix = 'DCRY-'
        elif company_code == 'shengchang' or '晟昌' in company_full_name:
            personnel_prefix = 'SCRY-'
        elif company_code == 'jiachengda' or '嘉诚达' in company_full_name:
            personnel_prefix = 'JCDRY-'
        else:
            # 默认前缀
            personnel_prefix = 'RY-'
    
    if request.method == 'POST':
        form = PersonnelForm(request.POST, tenant=getattr(request, 'tenant', None))
        if form.is_valid():
            personnel = form.save(commit=False)
            # 自动分配租户
            if hasattr(personnel, 'tenant') and hasattr(request, 'tenant'):
                personnel.tenant = request.tenant
            personnel.save()
            messages.success(request, "人员添加成功！")
            return redirect('eims_app:personnel_list')
        else:
            messages.error(request, "人员添加失败，请检查红色标注的输入项！")
    else:
        form = PersonnelForm(tenant=getattr(request, 'tenant', None))
    
    context = {
        'form': form,
        'personnel_prefix': personnel_prefix,
        'company_full_name': company_full_name,
    }
    return render(request, 'personnel/add.html', context)

@login_required
@user_passes_test(has_personnel_permission)
def personnel_detail(request, pk):
    """人员详情"""
    personnel = get_object_or_404(Personnel, pk=pk, is_deleted=False)
    
    field_data = []
    for field in personnel._meta.get_fields():
        if field.concrete and not field.many_to_many and not field.one_to_many:
            if field.name in ['id', 'create_time', 'update_time', 'is_deleted']:
                continue
            if hasattr(field, 'choices') and field.choices:
                display_value = getattr(personnel, f'get_{field.name}_display')()
            else:
                display_value = getattr(personnel, field.name, None)
            label = getattr(field, 'verbose_name', field.name).title()
            field_data.append({
                'name': field.name,
                'label': label,
                'value': display_value,
                'type': field.get_internal_type(),
                'has_choices': bool(getattr(field, 'choices', None))
            })
    
    return render(request, 'personnel/detail.html', {
        'personnel': personnel,
        'field_data': field_data
    })

@login_required
@user_passes_test(has_personnel_permission)
def personnel_edit(request, pk):
    """编辑人员"""
    personnel = get_object_or_404(Personnel, pk=pk, is_deleted=False)
    
    if request.method == "POST":
        form = PersonnelForm(request.POST, instance=personnel, tenant=getattr(request, 'tenant', None))
        if form.is_valid():
            form.save()
            messages.success(request, "人员信息修改成功！")
            return redirect("eims_app:personnel_list")
        else:
            messages.error(request, "人员信息修改失败，请检查红色标注的输入项！")
    else:
        form = PersonnelForm(instance=personnel, tenant=getattr(request, 'tenant', None))
    
    return render(request, "personnel/edit.html", {
        "form": form,
        "page_title": f"编辑人员：{personnel.name}",
        "personnel": personnel
    })

@login_required
@user_passes_test(has_personnel_permission)
def personnel_delete(request, pk):
    """删除人员（软删除）"""
    personnel = get_object_or_404(Personnel, pk=pk, is_deleted=False)
    personnel_name = personnel.name
    personnel.is_deleted = True
    personnel.save()
    messages.success(request, f"人员【{personnel_name}】删除成功！")
    return redirect("eims_app:personnel_list")

@login_required
@user_passes_test(has_personnel_permission)
def personnel_batch_delete(request):
    """批量删除人员"""
    if request.method == 'POST':
        personnel_ids = request.POST.getlist('personnel_ids')
    else:
        personnel_ids = request.GET.get('ids', '').split(',')
    
    personnel_ids = [pid for pid in personnel_ids if pid.isdigit()]
    
    if not personnel_ids:
        messages.warning(request, '未选择任何人员记录')
        return redirect('eims_app:personnel_list')
    
    try:
        with transaction.atomic():
            # 按租户过滤，防止误删其他公司的人员
            delete_filter = {'id__in': personnel_ids}
            if hasattr(request, 'tenant') and request.tenant:
                delete_filter['tenant_id'] = request.tenant.id
            
            personnels_to_delete = Personnel.objects.filter(**delete_filter)
            count = personnels_to_delete.count()
            
            if count == 0:
                messages.warning(request, '未找到要删除的人员记录')
                return redirect('eims_app:personnel_list')
            
            # 软删除
            for p in personnels_to_delete:
                p.is_deleted = True
                p.save()
            
            messages.success(request, f'成功删除 {count} 条人员记录')
            
    except Exception as e:
        messages.error(request, f'删除失败：{str(e)}')
    
    return redirect('eims_app:personnel_list')

@login_required
@user_passes_test(has_personnel_permission)
def personnel_import(request):
    """导入人员花名册信息（导入到 Employee 模型）"""
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]
        
        if not excel_file.name.endswith(".xlsx"):
            messages.error(request, "导入失败！仅支持.xlsx 格式的 Excel 文件")
            return redirect("eims_app:personnel_list")
        
        if excel_file.size > 10 * 1024 * 1024:
            messages.error(request, "导入失败！文件过大（最大支持 10MB）")
            return redirect("eims_app:personnel_list")
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            if ws is None or ws.max_row == 0:
                messages.error(request, "导入失败！Excel 文件为空或无有效工作表")
                return redirect("eims_app:personnel_list")
            
            headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
            required_headers = ["人员编号", "姓名"]
            missing = [h for h in required_headers if h not in headers]
            if missing:
                messages.error(request, f"导入失败！缺少必需表头：{', '.join(missing)}")
                return redirect("eims_app:personnel_list")
            
            if ws.max_row < 2:
                messages.error(request, "导入失败！Excel 文件无有效数据行")
                return redirect("eims_app:personnel_list")
            
            def safe_str(val, default=""):
                if val is None or str(val).strip() == "":
                    return default
                return str(val).strip()
            
            def parse_date(val):
                if not val:
                    return None
                try:
                    if isinstance(val, str):
                        val = val.strip()
                        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"]:
                            try:
                                return datetime.datetime.strptime(val, fmt).date()
                            except ValueError:
                                continue
                    elif hasattr(val, 'date'):
                        return val.date()
                except Exception:
                    pass
                return None
            
            # 数据映射
            gender_map = {'男': 0, '女': 1, '其他': 2}
            ethnic_map = {
                '汉族': 'han', '回族': 'hui', '满族': 'man', '蒙古族': 'mongol',
                '藏族': 'tibetan', '维吾尔族': 'uyghur', '其他': 'other'
            }
            education_map = {
                '小学': 'primary', '初中': 'junior', '高中': 'senior',
                '大专': 'college', '本科': 'bachelor', '硕士': 'master', '博士': 'doctor'
            }
            
            success_count = 0
            fail_count = 0
            fail_samples = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row or all(cell is None for cell in row):
                    continue
                
                try:
                    row_data = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
                    
                    employee_code = safe_str(row_data.get("人员编号"))
                    name = safe_str(row_data.get("姓名"))
                    
                    if not employee_code:
                        raise ValueError("人员编号为空（必需字段）")
                    if not name:
                        raise ValueError("姓名为空（必需字段）")
                    
                    # 准备 Employee 模型数据
                    employee_data = {
                        "employee_code": employee_code,
                        "name": name,
                        "gender": gender_map.get(safe_str(row_data.get("性别")), 0),
                        "id_card": safe_str(row_data.get("身份证号")),
                        "native_place": safe_str(row_data.get("籍贯")),
                        "ethnic": ethnic_map.get(safe_str(row_data.get("民族")), 'han'),
                        "education": education_map.get(safe_str(row_data.get("学历")), 'bachelor'),
                        "admin_position": safe_str(row_data.get("行政职务")),
                        "tech_position": safe_str(row_data.get("技术职务")),
                        "professional_qualification": safe_str(row_data.get("执业资格")),
                        "professional_title": safe_str(row_data.get("职称")),
                        "job_qualification": safe_str(row_data.get("任职资格")),
                        "mobile": safe_str(row_data.get("手机号码")),
                        "home_phone": safe_str(row_data.get("固定电话")),
                        "address": safe_str(row_data.get("住址")),
                        "emergency_contact": safe_str(row_data.get("应急联系人")),
                        "emergency_phone": safe_str(row_data.get("应急电话")),
                        "wechat": safe_str(row_data.get("微信")),
                        "email": safe_str(row_data.get("邮箱")),
                        "entry_time": parse_date(row_data.get("入职时间")),
                        "leave_time": parse_date(row_data.get("离职时间")),
                        "remark": safe_str(row_data.get("备注")),
                    }
                    
                    # 如果人员编号已存在则更新，否则创建
                    employee, created = Employee.objects.update_or_create(
                        employee_code=employee_code,
                        defaults=employee_data
                    )
                    success_count += 1
                    
                    # 创建对应的 Personnel 记录（二次分配）
                    personnel_data = {
                        "personnel_code": employee_code,
                        "name": name,
                        "gender": gender_map.get(safe_str(row_data.get("性别")), 0),
                        "department": safe_str(row_data.get("部门")) or "未分配",
                        "position": safe_str(row_data.get("岗位")),
                        "phone": safe_str(row_data.get("手机号码")),
                        "email": safe_str(row_data.get("邮箱")),
                        "entry_time": parse_date(row_data.get("入职时间")),
                        "leave_time": parse_date(row_data.get("离职时间")),
                        "employee": employee,  # 关联到 Employee
                        "operator": request.user.username,
                    }
                    
                    # 检查是否已存在 Personnel 记录
                    personnel, personnel_created = Personnel.objects.update_or_create(
                        personnel_code=employee_code,
                        defaults=personnel_data
                    )
                    
                except Exception as e:
                    fail_count += 1
                    if len(fail_samples) < 3:
                        fail_samples.append(f"第{row_idx}行：{str(e)} | 人员编号='{employee_code}'")
            
            msg = f"导入完成！成功{success_count}条，失败{fail_count}条"
            if fail_samples:
                msg += "\n\n失败示例（前 3 条）:\n" + "\n".join(fail_samples)
            messages.success(request, msg)
            return redirect("eims_app:personnel_list")
            
        except Exception as e:
            import traceback
            print("\n" + "="*70)
            print("【IMPORT CRITICAL ERROR】")
            traceback.print_exc()
            print("="*70 + "\n")
            messages.error(request, f"导入异常：{str(e)}（详见控制台日志）")
            return redirect("eims_app:personnel_list")
    
    messages.error(request, "导入失败！未检测到上传文件")
    return redirect("eims_app:personnel_list")

@login_required
@user_passes_test(has_personnel_permission)
def personnel_import_template(request):
    """下载人员花名册导入模板（Employee 模型）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "人员花名册导入模板"
    
    # 完整的员工信息表头（23 个字段：Employee 21 个 + Personnel 额外 2 个）
    headers = [
        "人员编号", "姓名", "性别", "身份证号", "籍贯", "民族", "学历", 
        "行政职务", "技术职务", "执业资格", "职称", "任职资格", 
        "手机号码", "固定电话", "入职时间", "离职时间", "住址", 
        "应急联系人", "应急电话", "微信", "邮箱", "备注",
        "部门", "岗位"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 示例数据（23 个字段）
    sample_data = [
        "EMP001", "张三", "男", "110101199001011234", "北京", "汉族", "本科", 
        "经理", "工程师", "一级建造师", "高级工程师", "合格", 
        "13800138001", "010-12345678", "2026-01-01", "", "北京市朝阳区", 
        "李四", "13900139002", "zhangsan", "zhangsan@example.com", "示例人员",
        "技术部", "技术员"
    ]
    
    for col_idx, value in enumerate(sample_data, 1):
        ws.cell(row=2, column=col_idx, value=value)
    
    # 设置列宽
    column_widths = {
        "人员编号": 15, "姓名": 15, "身份证号": 20, "籍贯": 15, "民族": 10, "学历": 10,
        "行政职务": 15, "技术职务": 15, "执业资格": 20, "职称": 15, "任职资格": 20,
        "手机号码": 15, "固定电话": 15, "入职时间": 15, "离职时间": 15, "住址": 25,
        "应急联系人": 15, "应急电话": 15, "微信": 15, "邮箱": 25, "备注": 30,
        "部门": 15, "岗位": 15
    }
    
    for col_idx, header in enumerate(headers, 1):
        width = column_widths.get(header, 12)
        col_letter = chr(64 + col_idx) if col_idx <= 26 else f"{chr(64 + col_idx // 26)}{chr(64 + col_idx % 26)}"
        ws.column_dimensions[col_letter].width = width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = FileResponse(
        buffer,
        as_attachment=True,
        filename='人员花名册导入模板.xlsx',
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    return response

@login_required
@user_passes_test(has_personnel_permission)
def personnel_export(request):
    """导出人员花名册 - 导出员工基本信息 + 项目分配信息（全部字段）"""
    
    # 如果是 /root/ 路径且没有选择公司，重定向到公司选择页面
    if hasattr(request, 'current_system') and request.current_system == 'root':
        if not hasattr(request, 'tenant') or not request.tenant:
            from django.contrib import messages
            messages.warning(request, '请先选择要查看的公司')
            return redirect('eims_app:tenant_select')
    
    # 获取选中的 ID 列表（POST 请求）
    selected_ids = None
    if request.method == 'POST':
        selected_ids = request.POST.getlist('personnel_ids')
    
    if selected_ids:
        # 从 Personnel 表查询，并预加载关联的 Employee 数据
        filter_dict = {'id__in': selected_ids, 'is_deleted': False}
        if hasattr(request, 'tenant') and request.tenant:
            filter_dict['tenant_id'] = request.tenant.id
        personnel_list = Personnel.objects.filter(**filter_dict).select_related('employee')
    else:
        # GET 请求：导出全部
        filter_dict = {'is_deleted': False}
        if hasattr(request, 'tenant') and request.tenant:
            filter_dict['tenant_id'] = request.tenant.id
        personnel_list = Personnel.objects.filter(**filter_dict).select_related('employee').order_by('personnel_code')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "人员花名册"
    
    # 完整的员工信息字段（27 个）
    headers = [
        "序号", "人员编号", "姓名", "性别", "身份证号", "籍贯", "民族", 
        "学历", "部门", "行政职务", "技术职务", "执业资格", "职称", "任职资格", 
        "手机号码", "固定电话", "入职时间", "离职时间", "住址", "应急联系人", 
        "应急电话", "微信", "邮箱", "操作人", "创建时间", "更新时间", "备注"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 设置列宽
    column_widths = {
        "姓名": 15, "身份证号": 20, "籍贯": 15, "住址": 25, "备注": 30, 
        "手机号码": 15, "固定电话": 15, "应急联系人": 15, "应急电话": 15, 
        "微信": 15, "邮箱": 25, "人员编号": 15, "部门": 15, "行政职务": 15, "技术职务": 15, 
        "执业资格": 20, "职称": 15, "任职资格": 20, "学历": 10, "民族": 10,
        "操作人": 12, "创建时间": 20, "更新时间": 20
    }
    
    for col_idx, header in enumerate(headers, 1):
        width = column_widths.get(header, 12)  # 默认宽度 12
        col_letter = chr(64 + col_idx) if col_idx <= 26 else f"{chr(64 + col_idx // 26)}{chr(64 + col_idx % 26)}"
        ws.column_dimensions[col_letter].width = width
    
    # 预加载部门信息（如果需要）
    # 数据映射
    gender_map = {0: '男', 1: '女', 2: '其他'}
    ethnic_map = dict(Employee.ETHNIC_CHOICES)
    education_map = dict(Employee.EDUCATION_CHOICES)
    
    for row_idx, personnel in enumerate(personnel_list, 2):
        # 获取关联的 Employee 对象（可能为 None）
        employee = personnel.employee
        
        # 列 1: 序号
        ws.cell(row=row_idx, column=1, value=row_idx-1).alignment = Alignment(horizontal="center", vertical="center")
        
        # 列 2: 人员编号
        ws.cell(row=row_idx, column=2, value=personnel.personnel_code or "-").alignment = Alignment(horizontal="center", vertical="center")
        
        # 列 3: 姓名
        ws.cell(row=row_idx, column=3, value=personnel.name or "-").alignment = Alignment(horizontal="left", vertical="center")
        
        # 列 4: 性别
        ws.cell(row=row_idx, column=4, value=gender_map.get(personnel.gender, "-")).alignment = Alignment(horizontal="center", vertical="center")
        
        # 列 5: 身份证号（从 Employee 获取）
        ws.cell(row=row_idx, column=5, value=employee.id_card if employee else "-").alignment = Alignment(horizontal="left", vertical="center")
        
        # 列 6: 籍贯（从 Employee 获取）
        ws.cell(row=row_idx, column=6, value=employee.native_place if employee else "-").alignment = Alignment(horizontal="left", vertical="center")
        
        # 列 7: 民族（从 Employee 获取）
        ws.cell(row=row_idx, column=7, value=ethnic_map.get(employee.ethnic, "-") if employee else "-").alignment = Alignment(horizontal="center", vertical="center")
        
        # 列 8: 学历（从 Employee 获取）
        ws.cell(row=row_idx, column=8, value=education_map.get(employee.education, "-") if employee else "-").alignment = Alignment(horizontal="center", vertical="center")
        
        # 列 9: 部门（从 Personnel 获取）
        ws.cell(row=row_idx, column=9, value=personnel.department or "-").alignment = Alignment(horizontal="left", vertical="center")
        
        # 列 10: 行政职务（从 Employee 获取）
        ws.cell(row=row_idx, column=10, value=employee.admin_position if employee else "-").alignment = Alignment(horizontal="left", vertical="center")
        
        # 列 11: 技术职务（从 Employee 获取）
        ws.cell(row=row_idx, column=11, value=employee.tech_position if employee else "-").alignment = Alignment(horizontal="left", vertical="center")
        
        # 列 12: 执业资格（从 Employee 获取）
        ws.cell(row=row_idx, column=12, value=employee.professional_qualification if employee else "-").alignment = Alignment(horizontal="left", vertical="center")
        
        # 列 13: 职称（从 Employee 获取）
        ws.cell(row=row_idx, column=13, value=employee.professional_title if employee else "-").alignment = Alignment(horizontal="left", vertical="center")
        
        # 列 14: 任职资格（从 Employee 获取）
        ws.cell(row=row_idx, column=14, value=employee.job_qualification if employee else "-").alignment = Alignment(horizontal="left", vertical="center")
        
        # 列 15: 手机号码（从 Employee 获取）
        ws.cell(row=row_idx, column=15, value=employee.mobile if employee else "-").alignment = Alignment(horizontal="center", vertical="center")
        
        # 列 16: 固定电话（从 Employee 获取）
        ws.cell(row=row_idx, column=16, value=employee.home_phone if employee else "-").alignment = Alignment(horizontal="left", vertical="center")
        
        # 列 17: 入职时间（从 Employee 获取）
        entry_time_str = employee.entry_time.strftime("%Y-%m-%d") if employee and employee.entry_time else "-"
        ws.cell(row=row_idx, column=17, value=entry_time_str).alignment = Alignment(horizontal="center", vertical="center")
        
        # 列 18: 离职时间（从 Employee 获取）
        leave_time_str = employee.leave_time.strftime("%Y-%m-%d") if employee and employee.leave_time else "-"
        ws.cell(row=row_idx, column=18, value=leave_time_str).alignment = Alignment(horizontal="center", vertical="center")
        
        # 列 19: 住址（从 Employee 获取）
        ws.cell(row=row_idx, column=19, value=employee.address if employee else "-").alignment = Alignment(horizontal="left", vertical="center")
        
        # 列 20: 应急联系人（从 Employee 获取）
        ws.cell(row=row_idx, column=20, value=employee.emergency_contact if employee else "-").alignment = Alignment(horizontal="left", vertical="center")
        
        # 列 21: 应急电话（从 Employee 获取）
        ws.cell(row=row_idx, column=21, value=employee.emergency_phone if employee else "-").alignment = Alignment(horizontal="center", vertical="center")
        
        # 列 22: 微信（从 Employee 获取）
        ws.cell(row=row_idx, column=22, value=employee.wechat if employee else "-").alignment = Alignment(horizontal="left", vertical="center")
        
        # 列 23: 邮箱（从 Employee 获取）
        ws.cell(row=row_idx, column=23, value=employee.email if employee else "-").alignment = Alignment(horizontal="left", vertical="center")
        
        # 列 24: 操作人（从 Personnel 获取）
        ws.cell(row=row_idx, column=24, value=personnel.operator or "-").alignment = Alignment(horizontal="left", vertical="center")
        
        # 列 25: 创建时间（从 Personnel 获取）
        create_time_str = personnel.create_time.strftime("%Y-%m-%d %H:%M:%S") if personnel.create_time else "-"
        ws.cell(row=row_idx, column=25, value=create_time_str).alignment = Alignment(horizontal="left", vertical="center")
        
        # 列 26: 更新时间（从 Personnel 获取）
        update_time_str = personnel.update_time.strftime("%Y-%m-%d %H:%M:%S") if personnel.update_time else "-"
        ws.cell(row=row_idx, column=26, value=update_time_str).alignment = Alignment(horizontal="left", vertical="center")
        
        # 列 27: 备注（从 Employee 获取）
        ws.cell(row=row_idx, column=27, value=employee.remark if employee else "-").alignment = Alignment(horizontal="left", vertical="center")
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = FileResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=人员信息表_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response


@login_required
@user_passes_test(has_personnel_permission)
def personnel_destination(request):
    """人员去向页面 - 展示所有人员的部门和项目分配情况"""
    
    # 获取筛选参数
    search_key = request.GET.get('keyword', '')
    department_filter = request.GET.get('department', '')
    
    # 获取列筛选参数（支持多个字段）
    filter_fields = {
        'personnel_code': request.GET.get('filter_personnel_code', ''),
        'name': request.GET.get('filter_name', ''),
        'department': request.GET.get('filter_department', ''),
        'position': request.GET.get('filter_position', ''),
        'gender': request.GET.get('filter_gender', ''),
        'phone': request.GET.get('filter_phone', ''),
    }
    
    # 获取排序参数
    sort_field = request.GET.get('sort_field', 'personnel_code')  # 默认排序字段
    sort_order = request.GET.get('sort_order', 'asc')  # 默认排序方向
    
    # 验证排序字段是否合法
    valid_sort_fields = ['personnel_code', 'name', 'department', 'position', 'gender', 'phone']
    if sort_field not in valid_sort_fields:
        sort_field = 'personnel_code'
    
    # 验证排序方向是否合法
    if sort_order not in ['asc', 'desc']:
        sort_order = 'asc'
    
    # 构建排序字符串
    order_by = sort_field if sort_order == 'asc' else f'-{sort_field}'
    
    # 查询所有人员（先不过滤，后续在Python层面处理拼音排序）
    personnel_filter = {'is_deleted': False}
    if hasattr(request, 'tenant') and request.tenant:
        personnel_filter['tenant_id'] = request.tenant.id
    
    # Determine the correct database based on URL path
    current_system = getattr(request, 'current_system', 'default') or 'default'
    
    # Map URL paths to databases
    # Note: Personnel is tenant-isolated, so /root/ should use default database
    # (or get tenant from session for more precise routing)
    db_mapping = {
        'dingce': 'dingce',
        'shengchang': 'shengchang',
        'jiachengda': 'jiachengda',
        'root': 'default',  # Personnel distributed in company databases, use default
    }
    target_db = db_mapping.get(current_system, 'default')
    
    # Use explicit database routing to avoid confusion with ProjectDetail
    personnel_list = list(Personnel.objects.using(target_db).filter(**personnel_filter))
    
    # 全局搜索筛选
    if search_key:
        personnel_list = [p for p in personnel_list if (
            search_key.lower() in (p.name or '').lower() or
            search_key.lower() in (p.personnel_code or '').lower() or
            search_key.lower() in (p.department or '').lower()
        )]
    
    # 部门筛选
    if department_filter:
        personnel_list = [p for p in personnel_list if p.department == department_filter]
    
    # 列筛选处理（支持多字段同时筛选）
    for field, value in filter_fields.items():
        if value:  # 如果筛选值不为空
            if field == 'gender':
                # 性别特殊处理：支持中文输入
                if '男' in value:
                    personnel_list = [p for p in personnel_list if p.gender == 0]
                elif '女' in value:
                    personnel_list = [p for p in personnel_list if p.gender == 1]
                elif '其他' in value or '其它' in value:
                    personnel_list = [p for p in personnel_list if p.gender == 2]
            else:
                # 其他字段使用模糊匹配
                personnel_list = [p for p in personnel_list if value.lower() in (getattr(p, field, '') or '').lower()]
    
    # 排序处理：对于需要拼音排序的字段，在Python层面处理
    # 定义需要拼音排序的字段
    pinyin_sort_fields = ['name', 'department', 'position']
    
    if sort_field in pinyin_sort_fields:
        # 使用拼音排序
        def get_pinyin_sort_key(personnel):
            """获取人员的拼音排序关键字"""
            field_value = getattr(personnel, sort_field, '') or ''
            # 将中文字符串转换为拼音
            pinyin_list = pinyin(field_value, style=Style.NORMAL, heteronym=False)
            # 将拼音列表扁平化为字符串（例如：['zhang'], ['san'] -> 'zhangsan'）
            pinyin_str = ''.join([p[0] for p in pinyin_list])
            return pinyin_str.lower()
        
        # 按拼音排序
        personnel_list = sorted(personnel_list, key=get_pinyin_sort_key, reverse=(sort_order == 'desc'))
    else:
        # 对于非拼音字段（如personnel_code, gender, phone），使用Python排序
        def get_sort_key(personnel):
            """获取普通字段的排序关键字"""
            value = getattr(personnel, sort_field, '')
            # 处理None值，将其放到最后
            if value is None:
                return '' if sort_order == 'asc' else '\uffff'
            return value
        
        personnel_list = sorted(personnel_list, key=get_sort_key, reverse=(sort_order == 'desc'))
    
    # 预加载所有项目信息（按租户过滤）
    # 注意：ProjectDetail模型没有is_deleted字段，直接查询所有项目
    project_filter = {}
    if hasattr(request, 'tenant') and request.tenant:
        project_filter['tenant_id'] = request.tenant.id
    
    project_info = {p.project_code: p for p in ProjectDetail.objects.filter(**project_filter)}
    
    # 为每个人员加载项目分配信息
    for personnel in personnel_list:
        # 主要项目
        if personnel.project_code and personnel.project_code in project_info:
            personnel.project1_info = project_info[personnel.project_code]
        else:
            personnel.project1_info = None
        
        # 项目 2
        if personnel.project_code2 and personnel.project_code2 in project_info:
            personnel.project2_info = project_info[personnel.project_code2]
        else:
            personnel.project2_info = None
            
        # 项目 3
        if personnel.project_code3 and personnel.project_code3 in project_info:
            personnel.project3_info = project_info[personnel.project_code3]
        else:
            personnel.project3_info = None
            
        # 项目 4
        if personnel.project_code4 and personnel.project_code4 in project_info:
            personnel.project4_info = project_info[personnel.project_code4]
        else:
            personnel.project4_info = None
            
        # 项目 5
        if personnel.project_code5 and personnel.project_code5 in project_info:
            personnel.project5_info = project_info[personnel.project_code5]
        else:
            personnel.project5_info = None
    
    # 获取所有部门
    all_departments = Department.objects.filter(
        is_deleted=False,
        status='active'
    ).order_by('department_code')
    
    # 分页处理
    paginator = Paginator(personnel_list, 20)  # 每页显示 20 条
    page = request.GET.get('page')
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    except Exception:
        page_obj = paginator.page(1)
    
    context = {
        'page_obj': page_obj,
        'all_personnel': personnel_list,
        'all_departments': all_departments,
        'selected_department': department_filter,
        'search_keyword': search_key,
        'sort_field': sort_field,
        'sort_order': sort_order,
        # 列筛选参数
        'filter_personnel_code': filter_fields['personnel_code'],
        'filter_name': filter_fields['name'],
        'filter_department': filter_fields['department'],
        'filter_position': filter_fields['position'],
        'filter_gender': filter_fields['gender'],
        'filter_phone': filter_fields['phone'],
        'home_url': reverse('eims_app:eims_index'),
        'eims_index_url': reverse('eims_app:eims_index'),
    }
    return render(request, "personnel/destination.html", context)
