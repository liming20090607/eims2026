from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator
from ..models import ProjectDetail
from ..forms.form_contract_management import ContractManagementForm
import os
from django.conf import settings
from django.http import FileResponse, Http404
import openpyxl
from django.utils import timezone
from datetime import datetime


def forgot_password(request):
    """忘记密码页面"""
    return render(request, 'sms_auth/forgot_password.html')


@login_required
def contract_management_list(request):
    """合同管理列表 - 显示所有合同信息"""
    
    # 获取搜索条件
    search_key = request.GET.get('search', '')
    contract_category = request.GET.get('contract_category', '')
    contract_status = request.GET.get('contract_status', '')
    settlement_status = request.GET.get('settlement_status', '')
    
    # 基础查询集 - 查询所有记录，不区分模块
    queryset = ProjectDetail.objects.select_related().all()
    
    # 应用筛选条件
    if search_key:
        queryset = queryset.filter(
            Q(project_name__icontains=search_key) |
            Q(contract_code__icontains=search_key) |
            Q(contract_party_a__icontains=search_key) |
            Q(contract_party_b__icontains=search_key)
        )
    
    if contract_category:
        queryset = queryset.filter(contract_category=contract_category)
    
    if contract_status:
        queryset = queryset.filter(contract_status=contract_status)
    
    if settlement_status:
        queryset = queryset.filter(settlement_status=settlement_status)
    
    # 排序
    sort_field = request.GET.get('sort', '-created_at')
    queryset = queryset.order_by(sort_field)
    
    # 分页
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 获取筛选后的总记录数
    total_count = queryset.count()
    
    context = {
        'page_obj': page_obj,
        'search_key': search_key,
        'contract_category': contract_category,
        'contract_status': contract_status,
        'settlement_status': settlement_status,
        'total_count': total_count,  # 添加总记录数
        'CONTRACT_CATEGORY_CHOICES': ProjectDetail.CONTRACT_CATEGORY_CHOICES,
        'CONTRACT_STATUS_CHOICES': ProjectDetail.CONTRACT_STATUS_CHOICES,
        'SETTLEMENT_STATUS_CHOICES': ProjectDetail.SETTLEMENT_STATUS_CHOICES,
    }
    
    response = render(request, 'contract_management/list.html', context)
    # 防止浏览器缓存，确保数据同步
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@login_required
def contract_management_add(request):
    """新增合同"""
    
    if request.method == 'POST':
        form = ContractManagementForm(request.POST, request.FILES)
        if form.is_valid():
            contract = form.save(commit=False)
            contract.save()
            messages.success(request, '✓ 合同添加成功！')
            return redirect('eims_app:contract_management_list')
        else:
            messages.error(request, '请检查输入内容是否正确')
    else:
        form = ContractManagementForm()
    
    context = {
        'form': form,
        'title': '新增合同',
        'action': 'add',
    }
    
    return render(request, 'contract_management/form.html', context)


@login_required
def contract_management_edit(request, pk):
    """编辑合同 - 仅超级管理员可用"""
    
    # 检查超级管理员权限
    if not request.user.is_superuser:
        messages.error(request, '⚠️ 权限不足：只有超级管理员才能编辑记录')
        return redirect('eims_app:contract_management_list')
    
    contract = get_object_or_404(ProjectDetail, pk=pk)
    
    if request.method == 'POST':
        form = ContractManagementForm(request.POST, request.FILES, instance=contract)
        if form.is_valid():
            form.save()
            messages.success(request, '✓ 合同更新成功！')
            return redirect('eims_app:contract_management_list')
        else:
            messages.error(request, '请检查输入内容是否正确')
    else:
        form = ContractManagementForm(instance=contract)
    
    context = {
        'form': form,
        'contract': contract,
        'title': '编辑合同',
        'action': 'edit',
    }
    
    return render(request, 'contract_management/form.html', context)


@login_required
def contract_management_detail(request, pk):
    """合同详情"""
    
    contract = get_object_or_404(ProjectDetail, pk=pk)
    
    context = {
        'contract': contract,
        'title': '合同详情',
    }
    
    return render(request, 'contract_management/detail.html', context)


@login_required
def contract_management_delete(request, pk):
    """删除合同 - 仅超级管理员可用"""
    
    # 检查超级管理员权限
    if not request.user.is_superuser:
        messages.error(request, '⚠️ 权限不足：只有超级管理员才能删除记录')
        return redirect('eims_app:contract_management_list')
    
    contract = get_object_or_404(ProjectDetail, pk=pk)
    
    if request.method == 'POST':
        contract.delete()
        messages.success(request, '✓ 合同已删除！')
        return redirect('eims_app:contract_management_list')
    
    context = {
        'contract': contract,
        'title': '确认删除',
    }
    
    return render(request, 'contract_management/delete.html', context)


@login_required
def preview_contract_text_contract(request, pk):
    """预览合同文本（合同管理模块）"""
    
    contract = get_object_or_404(ProjectDetail, pk=pk)
    
    if not contract.contract_text:
        raise Http404("合同文本不存在")
    
    try:
        response = FileResponse(contract.contract_text.open(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{os.path.basename(contract.contract_text.name)}"'
        return response
    except FileNotFoundError:
        raise Http404("文件未找到")


@login_required
def contract_management_import(request):
    """合同管理导入 - 从 Excel 导入数据到 ProjectDetail 表"""
    
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        
        if not excel_file:
            messages.error(request, '请选择要导入的 Excel 文件')
            return redirect('eims_app:contract_management_import')
        
        try:
            # 读取 Excel 文件
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            # 读取表头
            headers = [cell.value for cell in sheet[1]]
            
            # 字段映射（中文 → 英文）
            field_mapping = {
                '项目月报': 'monthly_report_required',
                '合同类别': 'contract_category',
                '项目编号': 'project_code',
                '合同编号': 'contract_code',
                '项目名称': 'project_name',
                '项目状态': 'project_status',
                '合同状态': 'contract_status',
                '结算情况': 'settlement_status',
                '合同甲方': 'contract_party_a',
                '合同乙方': 'contract_party_b',
                '签订日期': 'signing_date',
                '合同文本': 'contract_text',
                '合同总价 (元)': 'contract_amount',
                '合同总价（元）': 'contract_amount',  # 兼容旧格式
                '合同总价': 'contract_amount',  # 兼容旧格式
                '付款约定': 'payment_agreement',
                '累计回款 (元)': 'cumulative_payment',
                '累计回款': 'cumulative_payment',  # 兼容旧格式
                '合同余额 (元)': 'contract_balance',
                '合同余款': 'contract_balance',  # 兼容旧格式
                '项目规模': 'project_scale',
                '项目总投资（万元）': 'project_investment',
                '项目总投资': 'project_investment',  # 兼容旧格式
                '项目地址': 'project_address',
                '约定人员配备': 'agreed_staffing',
                '服务开始日期': 'service_start_date',
                '服务开始时间': 'service_start_date',  # 兼容旧格式
                '服务周期': 'service_period_months',
                '服务到期时间': 'service_deadline',
                '服务到期日期': 'service_deadline',  # 兼容旧格式
                '延期约定': 'extension_agreement',
                '实际延期情况': 'actual_extension_status',
                '报建情况': 'construction_permit_status',
                '施工许可证': 'construction_permit',
                '施工许可证状态': 'construction_permit',  # 兼容旧格式
                '进场通知': 'entry_notice',
                '进场通知书': 'entry_notice_document',  # 分开映射
                '进场时间': 'entry_time',
                '计划开工时间': 'planned_start_date',
                '计划开工日期': 'planned_start_date',  # 兼容旧格式
                '实际开工时间': 'actual_start_date',
                '实际开工日期': 'actual_start_date',  # 兼容旧格式
                '预计竣工时间': 'estimated_completion_date',
                '预计竣工日期': 'estimated_completion_date',  # 兼容旧格式
                '项目总监': 'project_director',
                '现场负责人': 'project_manager',
                '联系电话': 'contact_phone',
                '备注': 'remark',
            }
            
            success_count = 0
            error_count = 0
            error_rows = []
            
            # 从第 2 行开始读取数据
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 跳过空行
                    if not any(row):
                        continue
                    
                    # 创建数据字典
                    data = {}
                    for col_idx, value in enumerate(row, start=1):
                        if col_idx <= len(headers):
                            header = headers[col_idx - 1]
                            if header in field_mapping:
                                field_name = field_mapping[header]
                                if value is not None and str(value).strip():
                                    # 特殊字段处理
                                    if 'date' in field_name and value:
                                        try:
                                            if hasattr(value, 'strftime'):
                                                # 如果是 datetime 或 date 对象，直接提取日期部分
                                                if hasattr(value, 'date'):
                                                    # datetime 对象
                                                    data[field_name] = value.date()
                                                else:
                                                    # date 对象
                                                    data[field_name] = value
                                            else:
                                                # 字符串，尝试解析
                                                date_str = str(value).strip()
                                                if ' ' in date_str:
                                                    date_str = date_str.split(' ')[0]
                                                data[field_name] = datetime.strptime(date_str, '%Y-%m-%d').date()
                                        except Exception as e:
                                            messages.warning(request, f'第{row_idx}行 {header} 字段日期格式错误：{value}，已自动修正')
                                            data[field_name] = None  # 设为 None，稍后自动修复逻辑会处理
                                            # 不跳过，继续处理
                                    elif 'amount' in field_name or 'investment' in field_name:
                                        try:
                                            data[field_name] = float(str(value).replace(',', ''))
                                        except:
                                            data[field_name] = 0
                                    elif field_name == 'monthly_report_required':
                                        # 项目月报（支持多种写法）
                                        monthly_mapping = {
                                            '需要': True,
                                            '不需要': False,
                                            '是': True,
                                            '否': False,
                                            'True': True,
                                            'False': False,
                                            '1': True,
                                            '0': False,
                                            'YES': True,
                                            'NO': False,
                                        }
                                        data[field_name] = monthly_mapping.get(str(value).strip().upper(), True)
                                    elif field_name == 'contract_category':
                                        # 合同类别映射（支持多种写法）
                                        category_mapping = {
                                            '工程监理': 'engineering_supervision',
                                            '造价咨询': 'cost_consulting',
                                            '工程检测': 'testing',
                                            '检测': 'testing',  # 兼容旧格式
                                            '全过程咨询': 'whole_process_consulting',
                                        }
                                        data[field_name] = category_mapping.get(str(value).strip(), 'engineering_supervision')
                                    elif field_name == 'project_status':
                                        # 项目状态映射（支持多种写法和同义词）
                                        status_mapping = {
                                            '未开工': 'not_started',
                                            '在施工': 'under_construction',
                                            '停工中': 'stopped',
                                            '在停工': 'stopped',  # 兼容旧格式
                                            '已完工': 'completed',
                                            '完工': 'completed',  # 同义词
                                        }
                                        data[field_name] = status_mapping.get(str(value).strip(), 'not_started')
                                    elif field_name == 'contract_status':
                                        # 合同状态映射（支持多种写法和同义词）
                                        contract_status_mapping = {
                                            '待审核': 'pending_review',
                                            '在执行': 'executing',
                                            '执行中': 'executing',  # 同义词
                                            '已终止': 'terminated',
                                            '已解除': 'released',
                                        }
                                        data[field_name] = contract_status_mapping.get(str(value).strip(), 'pending_review')
                                    elif field_name == 'settlement_status':
                                        # 结算情况映射
                                        settlement_mapping = {
                                            '已结算': 'settled',
                                            '未结算': 'unsettled',
                                        }
                                        data[field_name] = settlement_mapping.get(str(value).strip(), 'unsettled')
                                    elif field_name == 'signing_date':
                                        # 签订日期（允许为空）
                                        if value:
                                            try:
                                                if hasattr(value, 'strftime'):
                                                    data[field_name] = value
                                                else:
                                                    data[field_name] = datetime.strptime(str(value), '%Y-%m-%d').date()
                                            except:
                                                data[field_name] = None
                                        else:
                                            data[field_name] = None
                                    elif field_name == 'entry_notice':
                                        # 进场通知映射
                                        entry_notice_mapping = {
                                            '有': 'yes',
                                            '无': 'no',
                                        }
                                        data[field_name] = entry_notice_mapping.get(str(value).strip(), 'no')
                                    else:
                                        data[field_name] = str(value).strip()
                    
                    # 必填字段检查 - 自动修复缺失的必填字段
                    required_fields = ['contract_category', 'contract_code', 'project_name', 'contract_party_a', 'contract_party_b']
                    missing_fields = [f for f in required_fields if not data.get(f)]
                    
                    if missing_fields:
                        # 自动修复：生成默认值
                        for field in missing_fields:
                            if field == 'contract_code' and data.get('project_code'):
                                # 用项目编号生成合同编号
                                data['contract_code'] = f"CTR-{data.get('project_code')}"
                                messages.warning(request, f'第{row_idx}行：合同编号缺失，已自动生成：{data["contract_code"]}')
                            elif field == 'project_name':
                                # 使用其他字段组合生成项目名称
                                if data.get('contract_code'):
                                    data['project_name'] = f"项目-{data.get('contract_code')}"
                                else:
                                    data['project_name'] = f"未命名项目-{row_idx}"
                                messages.warning(request, f'第{row_idx}行：项目名称缺失，已自动生成：{data["project_name"]}')
                            elif field in ['contract_party_a', 'contract_party_b']:
                                # 设置为默认值
                                data[field] = '待补充'
                                messages.warning(request, f'第{row_idx}行：{field}缺失，已设置为"待补充"')
                            elif field == 'contract_category':
                                # 设置为默认值
                                data[field] = 'engineering_supervision'
                                messages.warning(request, f'第{row_idx}行：合同类别缺失，已设置为默认值')
                        
                        logger.warning(f'第{row_idx}行：自动修复了缺失的必填字段 {missing_fields}')
                    
                    # 🔧 自动修复：处理特殊字段的数据验证
                    try:
                        # 修复金额字段 - 处理 '--' 等无效值，并确保不为 None
                        amount_fields = ['contract_amount', 'cumulative_payment', 'contract_balance', 'project_investment']
                        for amt_field in amount_fields:
                            # 确保金额字段有默认值 0.0
                            if amt_field not in data:
                                data[amt_field] = 0.0
                            else:
                                value = data.get(amt_field)
                                if value is None or value == '--' or value == '' or str(value).strip() in ['--', '-', '']:
                                    data[amt_field] = 0.0
                                    logger.debug(f'第{row_idx}行：{amt_field} 无效值已修正为 0')
                                else:
                                    try:
                                        data[amt_field] = float(str(value).replace(',', ''))
                                    except (ValueError, TypeError):
                                        data[amt_field] = 0.0
                                        logger.warning(f'第{row_idx}行：{amt_field} 无法解析，已设为 0')
                        
                        # 修复日期字段 - 确保是有效的 date 对象或 None
                        date_fields = ['signing_date', 'service_deadline', 'entry_time', 'planned_start_date', 
                                      'actual_start_date', 'estimated_completion_date', 'service_start_date']
                        for date_field in date_fields:
                            if date_field in data and data[date_field] is None:
                                # 保持为 None，Django 会处理
                                pass
                            elif date_field in data and isinstance(data[date_field], str):
                                # 如果是字符串，尝试解析
                                try:
                                    date_str = data[date_field].strip()
                                    if ' ' in date_str:
                                        date_str = date_str.split(' ')[0]
                                    data[date_field] = datetime.strptime(date_str, '%Y-%m-%d').date()
                                except:
                                    data[date_field] = None
                                    logger.warning(f'第{row_idx}行：{date_field} 日期格式错误，已设为 None')
                        
                        # 修复布尔字段 - 确保是 True/False
                        bool_fields = ['monthly_report_required']
                        for bool_field in bool_fields:
                            if bool_field in data:
                                value = data[bool_field]
                                if isinstance(value, bool):
                                    pass  # 已经是布尔值
                                elif value in [True, 'True', 'true', '1', 1, '是', 'YES', 'Yes', 'yes']:
                                    data[bool_field] = True
                                elif value in [False, 'False', 'false', '0', 0, '否', 'NO', 'No', 'no']:
                                    data[bool_field] = False
                                else:
                                    data[bool_field] = True  # 默认设为 True
                        
                        # 修复枚举字段 - 确保值在允许范围内
                        # 合同类别
                        if 'contract_category' in data:
                            valid_categories = ['engineering_supervision', 'cost_consulting', 'testing', 'whole_process_consulting']
                            if data['contract_category'] not in valid_categories:
                                logger.warning(f'第{row_idx}行：合同类别 {data["contract_category"]} 无效，已设为默认值')
                                data['contract_category'] = 'engineering_supervision'
                        
                        # 项目状态
                        if 'project_status' in data:
                            valid_statuses = ['not_started', 'under_construction', 'stopped', 'completed']
                            if data['project_status'] not in valid_statuses:
                                logger.warning(f'第{row_idx}行：项目状态 {data["project_status"]} 无效，已设为默认值')
                                data['project_status'] = 'not_started'
                        
                        # 合同状态
                        if 'contract_status' in data:
                            valid_contract_statuses = ['pending_review', 'executing', 'terminated', 'released']
                            if data['contract_status'] not in valid_contract_statuses:
                                logger.warning(f'第{row_idx}行：合同状态 {data["contract_status"]} 无效，已设为默认值')
                                data['contract_status'] = 'executing'
                        
                        # 结算情况
                        if 'settlement_status' in data:
                            valid_settlements = ['settled', 'unsettled']
                            if data['settlement_status'] not in valid_settlements:
                                logger.warning(f'第{row_idx}行：结算情况 {data["settlement_status"]} 无效，已设为默认值')
                                data['settlement_status'] = 'unsettled'
                        
                        # 进场通知
                        if 'entry_notice' in data:
                            valid_notices = ['yes', 'no']
                            if data['entry_notice'] not in valid_notices:
                                logger.warning(f'第{row_idx}行：进场通知 {data["entry_notice"]} 无效，已设为默认值')
                                data['entry_notice'] = 'no'
                                
                    except Exception as e:
                        logger.warning(f'第{row_idx}行：数据修复时出错：{e}，但继续尝试导入')
                    
                    # 创建或更新记录
                    # 优先使用 contract_code 查找，如果不存在则使用 project_code 查找
                    project = None
                    created = False
                    
                    if data.get('contract_code'):
                        try:
                            project = ProjectDetail.objects.get(contract_code=data.get('contract_code'))
                            # 找到了，更新数据
                            for key, value in data.items():
                                setattr(project, key, value)
                            project.save()
                            messages.info(request, f'合同编号 {data.get("contract_code")} 已存在，已更新')
                        except ProjectDetail.DoesNotExist:
                            # 没找到，尝试用 project_code 查找
                            if data.get('project_code'):
                                try:
                                    project = ProjectDetail.objects.get(project_code=data.get('project_code'))
                                    # 找到了，更新数据并补充 contract_code
                                    for key, value in data.items():
                                        setattr(project, key, value)
                                    project.save()
                                    messages.info(request, f'项目编号 {data.get("project_code")} 已存在（合同编号不同），已更新并关联')
                                except ProjectDetail.DoesNotExist:
                                    # 都不存在，创建新记录
                                    project = ProjectDetail.objects.create(**data)
                                    created = True
                            else:
                                # 没有 project_code，创建新记录
                                project = ProjectDetail.objects.create(**data)
                                created = True
                    elif data.get('project_code'):
                        # 只有 project_code，尝试查找
                        try:
                            project = ProjectDetail.objects.get(project_code=data.get('project_code'))
                            # 找到了，更新数据
                            for key, value in data.items():
                                setattr(project, key, value)
                            project.save()
                            messages.info(request, f'项目编号 {data.get("project_code")} 已存在，已更新')
                        except ProjectDetail.DoesNotExist:
                            # 创建新记录
                            project = ProjectDetail.objects.create(**data)
                            created = True
                    
                    if created:
                        success_count += 1
                        
                except Exception as e:
                    error_rows.append(f'第{row_idx}行：{str(e)}')
                    error_count += 1
                    continue
            
            # 显示导入结果
            if success_count > 0:
                messages.success(request, f'✓ 成功导入 {success_count} 条记录')
            if error_count > 0:
                messages.warning(request, f'⚠ 导入失败 {error_count} 条记录')
                for error in error_rows[:5]:  # 只显示前 5 个错误
                    messages.error(request, error)
            
            return redirect('eims_app:contract_management_list')
            
        except Exception as e:
            messages.error(request, f'导入失败：{str(e)}')
            return redirect('eims_app:contract_management_import')
    
    # GET 请求显示导入页面
    context = {
        'title': '合同管理导入',
        'template_fields': [
            '项目月报', '合同类别', '项目编号', '合同编号', '项目名称', '项目状态', '合同状态', '结算情况',
            '合同甲方', '合同乙方', '签订日期', '合同文本', '合同总价（元）', '付款约定', '累计回款', '合同余款',
            '项目规模', '项目总投资（万元）', '项目地址', '约定人员配备', '服务周期', '服务到期时间',
            '延期约定', '实际延期情况', '报建情况', '施工许可证', '进场通知', '进场时间', '计划开工时间',
            '实际开工时间', '预计竣工时间', '项目总监', '现场负责人', '联系电话', '备注'
        ]
    }
    
    return render(request, 'contract_management/import.html', context)


@login_required
def contract_management_batch_delete(request):
    """批量删除合同管理 - 仅超级管理员可用"""
    # 检查超级管理员权限
    if not request.user.is_superuser:
        messages.error(request, '⚠️ 权限不足：只有超级管理员才能批量删除记录')
        return redirect('eims_app:contract_management_list')
    
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if not ids:
            messages.warning(request, '⚠️ 未选择任何记录')
            return redirect('eims_app:contract_management_list')
        
        try:
            count = ProjectDetail.objects.filter(id__in=ids).count()
            ProjectDetail.objects.filter(id__in=ids).delete()
            messages.success(request, f'✓ 成功删除 {count} 条记录')
        except Exception as e:
            messages.error(request, f'❌ 删除失败：{str(e)}')
        
        return redirect('eims_app:contract_management_list')
    
    return redirect('eims_app:contract_management_list')


@login_required
def contract_management_export(request):
    """合同管理数据导出"""
    import openpyxl
    from django.http import HttpResponse
    from io import BytesIO
    from datetime import datetime
    
    # 获取选中的 ID（如果有）
    ids_param = request.GET.get('ids', '')
    
    if ids_param:
        project_ids = [int(id) for id in ids_param.split(',') if id.isdigit()]
        queryset = ProjectDetail.objects.filter(id__in=project_ids)
    else:
        # 按租户过滤
        filter_dict = {'is_deleted': False}
        if hasattr(request, 'tenant') and request.tenant:
            filter_dict['tenant_id'] = request.tenant.id
        queryset = ProjectDetail.objects.filter(**filter_dict)
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合同管理数据"
    
    # 设置表头（与项目信息总表字段顺序一致）
    headers = [
        '项目月报', '项目编号', '合同编号', '项目名称', '合同类别', '项目状态', '合同状态',
        '结算情况', '合同甲方', '合同乙方', '签订日期', '合同文本', '合同总价 (元)',
        '付款约定', '累计回款', '合同余额 (元)', '项目规模', '项目总投资',
        '项目地址', '约定人员配备', '服务开始日期', '服务周期', '服务到期日期', '延期约定',
        '实际延期情况', '施工许可证', '进场通知', '进场时间', '计划开工日期', '实际开工日期',
        '预计竣工日期', '项目总监', '现场负责人', '联系电话', '备注'
    ]
    
    # 设置表头样式
    from openpyxl.styles import Font, Alignment, Border, Side
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 填充数据
    status_map = dict(ProjectDetail.PROJECT_STATUS_CHOICES)
    contract_status_map = dict(ProjectDetail.CONTRACT_STATUS_CHOICES)
    settlement_map = dict(getattr(ProjectDetail, 'SETTLEMENT_STATUS_CHOICES', []))
    category_map = dict(getattr(ProjectDetail, 'CONTRACT_CATEGORY_CHOICES', []))
    permit_map = dict(getattr(ProjectDetail, 'PERMIT_STATUS_CHOICES', []))
    
    for row_idx, project in enumerate(queryset, 2):
        data_row = [
            '需要' if project.monthly_report_required else '不需要',
            project.project_code or '',
            project.contract_code or '',
            project.project_name or '',
            category_map.get(project.contract_category, project.contract_category or ''),
            status_map.get(project.project_status, project.project_status or ''),
            contract_status_map.get(project.contract_status, project.contract_status or ''),
            settlement_map.get(project.settlement_status, project.settlement_status or ''),
            project.contract_party_a or '',
            project.contract_party_b or '',
            project.signing_date.strftime('%Y-%m-%d') if project.signing_date else '',
            str(project.contract_text) if project.contract_text and hasattr(project.contract_text, '__str__') else '',
            str(project.contract_amount) if project.contract_amount else '',
            project.payment_agreement or '',
            str(project.cumulative_payment) if hasattr(project, 'cumulative_payment') and project.cumulative_payment else '',
            str(project.contract_balance) if hasattr(project, 'contract_balance') and project.contract_balance else '',
            project.project_scale or '',
            str(project.project_investment) if project.project_investment else '',
            project.project_address or '',
            project.agreed_staffing or '',
            project.service_start_date.strftime('%Y-%m-%d') if project.service_start_date else '',
            project.service_period_months or '',
            project.service_deadline.strftime('%Y-%m-%d') if project.service_deadline else '',
            project.extension_agreement or '',
            project.actual_extension_status or '',
            str(project.construction_permit) if project.construction_permit and hasattr(project.construction_permit, '__str__') else '',
            str(project.entry_notice) if project.entry_notice and hasattr(project.entry_notice, '__str__') else '',
            project.entry_time.strftime('%Y-%m-%d') if project.entry_time else '',
            project.planned_start_date.strftime('%Y-%m-%d') if project.planned_start_date else '',
            project.actual_start_date.strftime('%Y-%m-%d') if project.actual_start_date else '',
            project.estimated_completion_date.strftime('%Y-%m-%d') if project.estimated_completion_date else '',
            project.project_director or '',
            project.project_manager or '',
            project.contact_phone or '',
            project.remark or ''
        ]
        
        for col_idx, value in enumerate(data_row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
    
    # 调整列宽（按照新的字段顺序）
    column_widths = [15] * len(headers)
    column_widths[1] = 12  # 项目月报
    column_widths[2] = 15  # 项目编号
    column_widths[3] = 15  # 合同编号
    column_widths[4] = 30  # 项目名称
    column_widths[10] = 15  # 签订日期
    column_widths[11] = 20  # 合同文本
    column_widths[12] = 15  # 合同总价
    column_widths[13] = 25  # 付款约定
    column_widths[14] = 15  # 累计回款
    column_widths[15] = 15  # 合同余额
    column_widths[16] = 15  # 项目规模
    column_widths[17] = 15  # 项目总投资
    column_widths[18] = 25  # 项目地址
    column_widths[20] = 15  # 服务开始日期
    column_widths[21] = 12  # 服务周期
    column_widths[22] = 15  # 服务到期日期
    column_widths[27] = 15  # 进场时间
    column_widths[28] = 15  # 实际开工日期
    column_widths[29] = 15  # 预计竣工日期
    
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    # 写入内存并返回
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'合同管理数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
