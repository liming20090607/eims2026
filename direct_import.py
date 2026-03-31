"""
直接导入 Excel 数据 - 绕过验证问题
使用方法：python direct_import.py <excel 文件路径>
"""
import os
import sys
import django
from datetime import datetime
from decimal import Decimal

# 设置 Django 环境
sys.path.append(r'E:\EIMS2026')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'settings')
django.setup()

from openpyxl import load_workbook
from eims_app.models import ProjectDetail

def clean_date(value):
    """清理日期值"""
    if not value:
        return None
    if hasattr(value, 'strftime'):
        return value.date() if hasattr(value, 'date') else value
    # 字符串处理
    date_str = str(value).strip()
    if ' ' in date_str:
        date_str = date_str.split(' ')[0]
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except:
        return None

def clean_decimal(value):
    """清理数值"""
    if not value:
        return Decimal('0')
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    # 字符串处理
    try:
        return Decimal(str(value).replace(',', '').replace('--', '0'))
    except:
        return Decimal('0')

def map_choice_field(field_name, value):
    """映射选项字段"""
    value = str(value).strip()
    
    mappings = {
        'contract_category': {
            '工程监理': 'engineering_supervision',
            '造价咨询': 'cost_consulting',
            '工程检测': 'testing',
            '检测': 'testing',
            '全过程咨询': 'whole_process_consulting',
        },
        'project_status': {
            '未开工': 'not_started',
            '在施工': 'under_construction',
            '停工中': 'stopped',
            '在停工': 'stopped',
            '已完工': 'completed',
            '完工': 'completed',
        },
        'contract_status': {
            '待审核': 'pending_review',
            '在执行': 'executing',
            '执行中': 'executing',
            '已终止': 'terminated',
            '已解除': 'released',
        },
        'settlement_status': {
            '已结算': 'settled',
            '未结算': 'unsettled',
        },
        'entry_notice': {
            '有': 'yes',
            '无': 'no',
        },
    }
    
    mapping_dict = mappings.get(field_name, {})
    return mapping_dict.get(value, None)

def direct_import(excel_path):
    """直接导入 Excel 数据"""
    print("=" * 80)
    print("📥 开始直接导入数据")
    print("=" * 80)
    
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    # 读取表头
    headers = [cell.value.strip() if cell.value else '' for cell in ws[1]]
    print(f"\n📋 发现 {len(headers)} 列")
    
    # 字段映射
    field_mapping = {
        '项目月报': 'monthly_report_required',
        '合同类别': 'contract_category',
        '项目编号': 'project_code',
        '合同编号': 'contract_code',
        '项目名称': 'project_name',
        '项目状态': 'project_status',
        '合同状态': 'contract_status',
        '结算情况': 'settlement_status',
        '合同甲方': 'contract_party_a',
        '合同乙方': 'contract_party_b',
        '签订日期': 'signing_date',
        '合同总价 (元)': 'contract_amount',
        '合同总价（元）': 'contract_amount',  # 兼容中文括号
        '累计回款 (元)': 'cumulative_payment',
        '累计回款': 'cumulative_payment',  # 兼容旧格式
        '合同余额 (元)': 'contract_balance',
        '合同余款': 'contract_balance',  # 兼容旧格式
        '项目规模': 'project_scale',
        '项目总投资（万元）': 'project_investment',
        '项目地址': 'project_address',
        '约定人员配备': 'agreed_staffing',
        '服务开始日期': 'service_start_date',
        '服务周期': 'service_period_months',
        '服务到期时间': 'service_deadline',
        '延期约定': 'extension_agreement',
        '进场通知': 'entry_notice',
        '进场时间': 'entry_time',
        '计划开工时间': 'planned_start_date',
        '实际开工时间': 'actual_start_date',
        '预计竣工时间': 'estimated_completion_date',
        '项目总监': 'project_director',
        '现场负责人': 'project_manager',
        '联系电话': 'contact_phone',
        '备注': 'remark',  # 修正为 remark
    }
    
    success_count = 0
    error_count = 0
    update_count = 0
    
    # 处理数据行
    for row_idx in range(2, ws.max_row + 1):
        print(f"\n处理第{row_idx}行...")
        
        # 构建数据字典
        data = {}
        for col_idx, header in enumerate(headers, 1):
            if header in field_mapping:
                field_name = field_mapping[header]
                value = ws.cell(row=row_idx, column=col_idx).value
                
                # 跳过空值
                if value is None or str(value).strip() == '':
                    continue
                
                # 特殊字段处理
                if field_name in ['contract_category', 'project_status', 'contract_status', 'settlement_status', 'entry_notice']:
                    mapped = map_choice_field(field_name, value)
                    if mapped:
                        data[field_name] = mapped
                    else:
                        print(f"  ⚠️ 无法映射 {field_name}: {value}")
                        data[field_name] = None
                elif 'date' in field_name or field_name in ['signing_date', 'service_deadline', 'entry_time', 'planned_start_date', 'actual_start_date', 'estimated_completion_date', 'service_start_date']:
                    data[field_name] = clean_date(value)
                elif 'amount' in field_name or 'investment' in field_name:
                    data[field_name] = clean_decimal(value)
                elif field_name == 'service_period_months':
                    # 服务周期：提取数字部分
                    if isinstance(value, str):
                        # 去除"个月"等文字
                        import re
                        numbers = re.findall(r'\d+', value)
                        data[field_name] = int(numbers[0]) if numbers else 0
                    else:
                        data[field_name] = int(value) if value else 0
                elif field_name == 'monthly_report_required':
                    data[field_name] = str(value).strip().upper() in ['需要', '是', 'TRUE', '1', 'YES']
                else:
                    data[field_name] = str(value).strip()
        
        # 检查必填字段
        required = ['project_code', 'contract_code', 'project_name', 'contract_party_a', 'contract_party_b']
        if not all(data.get(f) for f in required):
            missing = [f for f in required if not data.get(f)]
            print(f"  ❌ 缺少必填字段：{missing}")
            error_count += 1
            continue
        
        # 创建或更新
        try:
            project = ProjectDetail.objects.filter(project_code=data['project_code']).first()
            
            if project:
                # 更新
                for key, value in data.items():
                    setattr(project, key, value)
                project.save()
                print(f"  ✓ 更新成功：{data['project_name']}")
                update_count += 1
            else:
                # 创建
                project = ProjectDetail.objects.create(**data)
                print(f"  ✓ 创建成功：{data['project_name']}")
                success_count += 1
                
        except Exception as e:
            print(f"  ❌ 错误：{e}")
            error_count += 1
    
    print("\n" + "=" * 80)
    print(f"✅ 导入完成！")
    print(f"   新建：{success_count} 条")
    print(f"   更新：{update_count} 条")
    print(f"   失败：{error_count} 条")
    print("=" * 80)

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("使用方法：python direct_import.py <excel 文件路径>")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    if not os.path.exists(excel_path):
        print(f"❌ 文件不存在：{excel_path}")
        sys.exit(1)
    
    direct_import(excel_path)
