"""
创建项目和合同导入模板 Excel 文件
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def create_project_ledger_template():
    """创建项目台账导入模板"""
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "监理项目台账"
    
    # 定义样式
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    required_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    optional_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 表头（按照你的 36 个字段顺序）
    headers = [
        '项目月报', '合同类别', '项目编号', '合同编号', '项目名称', 
        '项目状态', '合同状态', '结算情况', '合同甲方', '合同乙方', 
        '签订日期', '合同文本', '合同总价（元）', '付款约定', '累计回款', 
        '合同余款', '项目规模', '项目总投资（万元）', '项目地址', 
        '约定人员配备', '服务周期', '服务到期时间', '延期约定', 
        '实际延期情况', '报建情况', '施工许可证', '进场通知', 
        '进场通知书', '进场时间', '计划开工时间', '实际开工时间', 
        '预计竣工时间', '项目总监', '现场负责人', '联系电话', '备注'
    ]
    
    # 必填字段标记
    required_fields = {'项目编号', '合同编号', '项目名称', '合同甲方', '合同乙方'}
    
    # 写入表头
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # 标记必填字段
        if header in required_fields:
            note_cell = sheet.cell(row=2, column=col, value='*必填')
            note_cell.fill = required_fill
            note_cell.font = Font(color='FF0000', bold=True, size=9)
            note_cell.alignment = Alignment(horizontal="center", vertical="center")
            note_cell.border = thin_border
        else:
            note_cell = sheet.cell(row=2, column=col, value='可选')
            note_cell.fill = optional_fill
            note_cell.font = Font(color='1976D2', size=9)
            note_cell.alignment = Alignment(horizontal="center", vertical="center")
            note_cell.border = thin_border
    
    # 示例数据行
    example_data = [
        '是',  # 项目月报
        '工程监理',  # 合同类别
        'TEST2026001',  # 项目编号
        'HT-2026-001',  # 合同编号
        '测试项目名称',  # 项目名称
        '未开工',  # 项目状态
        '待审核',  # 合同状态
        '未结算',  # 结算情况
        '甲方公司名称',  # 合同甲方
        '乙方公司名称',  # 合同乙方
        '2026-01-15',  # 签订日期
        '',  # 合同文本（留空）
        '1000000.00',  # 合同总价（元）
        '按月支付工程进度款',  # 付款约定
        '0.00',  # 累计回款
        '1000000.00',  # 合同余款
        '建筑面积 5000 平方米',  # 项目规模
        '500.00',  # 项目总投资（万元）
        'XX 市 XX 区 XX 路',  # 项目地址
        '配置 10 名专业人员',  # 约定人员配备
        '12 个月',  # 服务周期
        '2027-01-14',  # 服务到期时间
        '无',  # 延期约定
        '无',  # 实际延期情况
        '已完成',  # 报建情况
        '',  # 施工许可证（留空）
        '有',  # 进场通知
        '',  # 进场通知书（留空）
        '2026-02-01',  # 进场时间
        '2026-03-01',  # 计划开工时间
        '',  # 实际开工时间（留空）
        '2027-03-01',  # 预计竣工时间
        '张三',  # 项目总监
        '李四',  # 现场负责人
        '13800138000',  # 联系电话
        '这是测试备注',  # 备注
    ]
    
    # 写入示例数据
    for col, value in enumerate(example_data, start=1):
        sheet.cell(row=3, column=col, value=value)
    
    # 设置列宽
    column_widths = {
        1: 10,   # 项目月报
        2: 12,   # 合同类别
        3: 15,   # 项目编号
        4: 15,   # 合同编号
        5: 30,   # 项目名称
        6: 10,   # 项目状态
        7: 10,   # 合同状态
        8: 10,   # 结算情况
        9: 20,   # 合同甲方
        10: 20,  # 合同乙方
        11: 12,  # 签订日期
        12: 15,  # 合同文本
        13: 15,  # 合同总价
        14: 25,  # 付款约定
        15: 15,  # 累计回款
        16: 15,  # 合同余款
        17: 20,  # 项目规模
        18: 18,  # 项目总投资
        19: 25,  # 项目地址
        20: 18,  # 约定人员配备
        21: 15,  # 服务周期
        22: 12,  # 服务到期时间
        23: 15,  # 延期约定
        24: 18,  # 实际延期情况
        25: 12,  # 报建情况
        26: 15,  # 施工许可证
        27: 12,  # 进场通知
        28: 15,  # 进场通知书
        29: 12,  # 进场时间
        30: 12,  # 计划开工时间
        31: 12,  # 实际开工时间
        32: 12,  # 预计竣工时间
        33: 12,  # 项目总监
        34: 15,  # 现场负责人
        35: 15,  # 联系电话
        36: 30,  # 备注
    }
    
    for col, width in column_widths.items():
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # 添加说明工作表
    guide_sheet = workbook.create_sheet(title="填写说明")
    
    guide_content = [
        ['项目台账导入模板填写说明'],
        ['', ''],
        ['一、必填字段（5 个）'],
        ['字段名称', '说明', '示例'],
        ['项目编号', '唯一标识，不能重复', 'TEST2026001'],
        ['合同编号', '合同编号', 'HT-2026-001'],
        ['项目名称', '项目全称', 'XX 工程监理项目'],
        ['合同甲方', '甲方单位全称', 'XX 公司'],
        ['合同乙方', '乙方单位全称', 'XX 监理公司'],
        ['', ''],
        ['二、选择字段'],
        ['字段名称', '可选值'],
        ['项目月报', '需要/不需要'],
        ['合同类别', '工程监理/造价咨询/工程检测/全过程咨询'],
        ['项目状态', '未开工/在施工/停工中/已完工'],
        ['合同状态', '待审核/在执行/已终止/已解除'],
        ['结算情况', '已结算/未结算'],
        ['报建情况', '已完成/未完成'],
        ['进场通知', '有/无'],
        ['', ''],
        ['三、日期字段'],
        ['字段名称', '格式要求', '示例'],
        ['签订日期', 'YYYY-MM-DD 或留空', '2026-01-15'],
        ['服务到期时间', 'YYYY-MM-DD', '2027-01-14'],
        ['进场时间', 'YYYY-MM-DD', '2026-02-01'],
        ['计划开工时间', 'YYYY-MM-DD', '2026-03-01'],
        ['实际开工时间', 'YYYY-MM-DD', '2026-03-01 或留空'],
        ['预计竣工时间', 'YYYY-MM-DD', '2027-03-01'],
        ['', ''],
        ['四、金额字段'],
        ['字段名称', '单位', '说明'],
        ['合同总价（元）', '元', '数字格式，如：1000000.00'],
        ['累计回款', '元', '数字格式，如：500000.00'],
        ['合同余款', '元', '数字格式，如：500000.00'],
        ['项目总投资（万元）', '万元', '数字格式，如：500.00'],
        ['', ''],
        ['五、注意事项'],
        ['1. 第一行为表头，不要修改或删除'],
        ['2. 第二行为必填标记，第三行为示例数据，从第四行开始填写实际数据'],
        ['3. 带 * 号的字段为必填项，不能为空'],
        ['4. 日期字段请使用 YYYY-MM-DD 格式'],
        ['5. 金额字段请填写数字，不要包含货币符号'],
        ['6. 选择字段请从下拉列表中选择（如果已设置）或按指定文字填写'],
        ['7. 文件上传后请保留此说明工作表，方便下次参考'],
    ]
    
    for row_idx, row_data in enumerate(guide_content, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = guide_sheet.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
                cell.alignment = Alignment(horizontal="center")
            elif row_idx == 3 or (row_idx > 3 and row_data[0] != ''):
                cell.font = Font(bold=True)
    
    # 设置说明工作表列宽
    guide_sheet.column_dimensions['A'].width = 20
    guide_sheet.column_dimensions['B'].width = 30
    guide_sheet.column_dimensions['C'].width = 25
    
    # 保存文件
    file_path = r'e:\EIMS2026\static\templates\project_ledger_import_template.xlsx'
    workbook.save(file_path)
    
    print(f"✅ 项目台账导入模板已创建：{file_path}")
    return file_path


def create_contract_management_template():
    """创建合同管理导入模板"""
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "合同管理"
    
    # 定义样式
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    required_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    optional_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 表头（与项目台账相同的 36 个字段）
    headers = [
        '项目月报', '合同类别', '项目编号', '合同编号', '项目名称', 
        '项目状态', '合同状态', '结算情况', '合同甲方', '合同乙方', 
        '签订日期', '合同文本', '合同总价（元）', '付款约定', '累计回款', 
        '合同余款', '项目规模', '项目总投资（万元）', '项目地址', 
        '约定人员配备', '服务周期', '服务到期时间', '延期约定', 
        '实际延期情况', '报建情况', '施工许可证', '进场通知', 
        '进场通知书', '进场时间', '计划开工时间', '实际开工时间', 
        '预计竣工时间', '项目总监', '现场负责人', '联系电话', '备注'
    ]
    
    # 必填字段标记
    required_fields = {'项目编号', '合同编号', '项目名称', '合同甲方', '合同乙方'}
    
    # 写入表头
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # 标记必填字段
        if header in required_fields:
            note_cell = sheet.cell(row=2, column=col, value='*必填')
            note_cell.fill = required_fill
            note_cell.font = Font(color='FF0000', bold=True, size=9)
            note_cell.alignment = Alignment(horizontal="center", vertical="center")
            note_cell.border = thin_border
        else:
            note_cell = sheet.cell(row=2, column=col, value='可选')
            note_cell.fill = optional_fill
            note_cell.font = Font(color='1976D2', size=9)
            note_cell.alignment = Alignment(horizontal="center", vertical="center")
            note_cell.border = thin_border
    
    # 示例数据行
    example_data = [
        '是',  # 项目月报
        '工程监理',  # 合同类别
        'CONTRACT2026001',  # 项目编号
        'HT-2026-001',  # 合同编号
        '合同管理测试项目',  # 项目名称
        '未开工',  # 项目状态
        '在执行',  # 合同状态
        '未结算',  # 结算情况
        '甲方单位名称',  # 合同甲方
        '乙方单位名称',  # 合同乙方
        '2026-01-10',  # 签订日期
        '',  # 合同文本（留空）
        '1500000.00',  # 合同总价（元）
        '按季度支付',  # 付款约定
        '750000.00',  # 累计回款
        '750000.00',  # 合同余款
        '大型项目',  # 项目规模
        '800.00',  # 项目总投资（万元）
        'XX 省 XX 市',  # 项目地址
        '配置 15 人团队',  # 约定人员配备
        '18 个月',  # 服务周期
        '2027-07-09',  # 服务到期时间
        '可延期 3 个月',  # 延期约定
        '实际延期 1 个月',  # 实际延期情况
        '已完成',  # 报建情况
        '',  # 施工许可证（留空）
        '有',  # 进场通知
        '',  # 进场通知书（留空）
        '2026-01-20',  # 进场时间
        '2026-02-15',  # 计划开工时间
        '2026-02-15',  # 实际开工时间
        '2027-08-15',  # 预计竣工时间
        '王五',  # 项目总监
        '赵六',  # 现场负责人
        '13900139000',  # 联系电话
        '合同管理测试备注',  # 备注
    ]
    
    # 写入示例数据
    for col, value in enumerate(example_data, start=1):
        sheet.cell(row=3, column=col, value=value)
    
    # 设置列宽（与项目台账相同）
    column_widths = {
        1: 10, 2: 12, 3: 15, 4: 15, 5: 30,
        6: 10, 7: 10, 8: 10, 9: 20, 10: 20,
        11: 12, 12: 15, 13: 15, 14: 25, 15: 15,
        16: 15, 17: 20, 18: 18, 19: 25, 20: 18,
        21: 15, 22: 12, 23: 15, 24: 18, 25: 12,
        26: 15, 27: 12, 28: 15, 29: 12, 30: 12,
        31: 12, 32: 12, 33: 12, 34: 15, 35: 15,
        36: 30,
    }
    
    for col, width in column_widths.items():
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # 添加说明工作表
    guide_sheet = workbook.create_sheet(title="填写说明")
    
    guide_content = [
        ['合同管理导入模板填写说明'],
        ['', ''],
        ['一、必填字段（5 个）'],
        ['字段名称', '说明', '示例'],
        ['项目编号', '唯一标识，不能重复', 'CONTRACT2026001'],
        ['合同编号', '合同编号', 'HT-2026-001'],
        ['项目名称', '项目全称', 'XX 项目'],
        ['合同甲方', '甲方单位全称', 'XX 公司'],
        ['合同乙方', '乙方单位全称', 'XX 公司'],
        ['', ''],
        ['二、选择字段'],
        ['字段名称', '可选值'],
        ['项目月报', '需要/不需要'],
        ['合同类别', '工程监理/造价咨询/工程检测/全过程咨询'],
        ['项目状态', '未开工/在施工/停工中/已完工'],
        ['合同状态', '待审核/在执行/已终止/已解除'],
        ['结算情况', '已结算/未结算'],
        ['报建情况', '已完成/未完成'],
        ['进场通知', '有/无'],
        ['', ''],
        ['三、日期字段'],
        ['字段名称', '格式要求', '示例'],
        ['签订日期', 'YYYY-MM-DD 或留空', '2026-01-10'],
        ['服务到期时间', 'YYYY-MM-DD', '2027-07-09'],
        ['进场时间', 'YYYY-MM-DD', '2026-01-20'],
        ['计划开工时间', 'YYYY-MM-DD', '2026-02-15'],
        ['实际开工时间', 'YYYY-MM-DD', '2026-02-15'],
        ['预计竣工时间', 'YYYY-MM-DD', '2027-08-15'],
        ['', ''],
        ['四、金额字段'],
        ['字段名称', '单位', '说明'],
        ['合同总价（元）', '元', '数字格式，如：1500000.00'],
        ['累计回款', '元', '数字格式，如：750000.00'],
        ['合同余款', '元', '数字格式，如：750000.00'],
        ['项目总投资（万元）', '万元', '数字格式，如：800.00'],
        ['', ''],
        ['五、注意事项'],
        ['1. 第一行为表头，不要修改或删除'],
        ['2. 第二行为必填标记，第三行为示例数据，从第四行开始填写实际数据'],
        ['3. 带 * 号的字段为必填项，不能为空'],
        ['4. 日期字段请使用 YYYY-MM-DD 格式'],
        ['5. 金额字段请填写数字，不要包含货币符号'],
        ['6. 选择字段请按指定文字填写'],
        ['7. 文件上传后请保留此说明工作表，方便下次参考'],
    ]
    
    for row_idx, row_data in enumerate(guide_content, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = guide_sheet.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
                cell.alignment = Alignment(horizontal="center")
            elif row_idx == 3 or (row_idx > 3 and row_data[0] != ''):
                cell.font = Font(bold=True)
    
    # 设置说明工作表列宽
    guide_sheet.column_dimensions['A'].width = 20
    guide_sheet.column_dimensions['B'].width = 30
    guide_sheet.column_dimensions['C'].width = 25
    
    # 保存文件
    file_path = r'e:\EIMS2026\static\templates\contract_management_import_template.xlsx'
    workbook.save(file_path)
    
    print(f"✅ 合同管理导入模板已创建：{file_path}")
    return file_path


if __name__ == '__main__':
    print("开始创建导入模板文件...")
    print()
    
    # 创建项目台账模板
    project_path = create_project_ledger_template()
    
    # 创建合同管理模板
    contract_path = create_contract_management_template()
    
    print()
    print("=" * 60)
    print("✅ 所有模板创建完成！")
    print("=" * 60)
    print()
    print(f"项目台账模板：{project_path}")
    print(f"合同管理模板：{contract_path}")
    print()
    print("提示：请在导入页面添加下载链接")
