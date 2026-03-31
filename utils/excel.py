import openpyxl
from openpyxl.styles import Font, Border, Side
from django.http import HttpResponse
from io import BytesIO

def export_excel(queryset, field_map, sheet_name="数据导出"):
    """
    通用Excel导出函数
    :param queryset: 数据查询集
    :param field_map: 字段映射（{"表头": "模型字段"}）
    :param sheet_name: 工作表名称
    :return: HttpResponse（Excel文件）
    """
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 设置表头
    headers = list(field_map.keys())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

    # 填充数据
    fields = list(field_map.values())
    for row, obj in enumerate(queryset, 2):
        for col, field in enumerate(fields, 1):
            # 处理关联字段（如project__name）
            if "__" in field:
                rel_fields = field.split("__")
                value = getattr(obj, rel_fields[0], None)
                for rel_field in rel_fields[1:]:
                    value = getattr(value, rel_field, None) if value else None
            else:
                value = getattr(obj, field, None)
            # 空值处理
            cell_value = str(value) if value is not None else ""
            ws.cell(row=row, column=col, value=cell_value)
            ws.cell(row=row, column=col).border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    # 写入内存并返回
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"{sheet_name}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

def import_excel(file, model, field_map):
    """
    通用Excel导入函数
    :param file: 上传的Excel文件
    :param model: 数据模型
    :param field_map: 字段映射（{"表头": "模型字段"}）
    :return: (success_count, fail_count, fail_msg)
    """
    success_count = 0
    fail_count = 0
    fail_msg = []

    wb = openpyxl.load_workbook(file)
    ws = wb.active

    # 获取表头并验证
    headers = [cell.value for cell in ws[1]]
    required_headers = list(field_map.keys())
    missing_headers = [h for h in required_headers if h not in headers]
    if missing_headers:
        return 0, 1, [f"缺失必填表头：{', '.join(missing_headers)}"]

    # 读取数据行（跳过表头）
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        row_data = dict(zip(headers, row))
        model_data = {}
        valid = True

        # 映射字段并验证
        for excel_header, model_field in field_map.items():
            value = row_data.get(excel_header)
            # 处理必填字段（根据模型字段是否允许空值扩展）
            model_data[model_field] = value

        # 保存数据
        try:
            model.objects.create(**model_data)
            success_count += 1
        except Exception as e:
            fail_count += 1
            fail_msg.append(f"第{row_num}行导入失败：{str(e)}")

    return success_count, fail_count, fail_msg 
